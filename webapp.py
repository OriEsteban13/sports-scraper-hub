import asyncio
import io
import os
import re
import json
import queue
import sqlite3
import subprocess
import tempfile
import threading
import time
import uuid
import zipfile as _zipfile_mod
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime, date
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse, urljoin

# PostgreSQL support (optional — only needed when DATABASE_URL is set)
try:
    import psycopg2
    import psycopg2.extras
    _PSYCOPG2_AVAILABLE = True
except ImportError:
    _PSYCOPG2_AVAILABLE = False

# Supabase Storage support (optional — only needed when SUPABASE_URL/KEY are set)
try:
    from supabase import create_client as _supabase_create_client
    _SUPABASE_AVAILABLE = True
except ImportError:
    _SUPABASE_AVAILABLE = False

# Load .env from project root (won't overwrite values already in the environment)
try:
    from dotenv import load_dotenv, set_key, dotenv_values
    _ENV_FILE = Path(__file__).parent / ".env"
    if not _ENV_FILE.exists():
        _ENV_FILE.touch()
    load_dotenv(_ENV_FILE, override=False)
    _DOTENV_AVAILABLE = True
except ImportError:
    _DOTENV_AVAILABLE = False
    _ENV_FILE = None

import hashlib
import mimetypes
import urllib.request

from fastapi import FastAPI, Request, Form, UploadFile, File, BackgroundTasks
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

# Scrapling imported lazily to keep startup RAM low.
# On Render (512MB), Playwright/Chromium is not installed —
# DynamicFetcher/StealthyFetcher are replaced with no-op stubs that return None.
def _get_fetchers():
    from scrapling.fetchers import Fetcher
    try:
        from scrapling.fetchers import DynamicFetcher, StealthyFetcher
    except Exception:
        class _NoopFetcher:
            @staticmethod
            def fetch(*a, **kw): return None
        DynamicFetcher = StealthyFetcher = _NoopFetcher
    return Fetcher, DynamicFetcher, StealthyFetcher

app = FastAPI(title="Sports Scraper Hub")
templates = Jinja2Templates(directory="templates")

def _fmtnum(value, decimals=0):
    """Format a number with Spanish notation: dots as thousands sep, comma as decimal."""
    if value is None or value == '':
        return '—'
    try:
        v = float(value)
        fmt = f"{v:,.{decimals}f}"          # e.g. "1,234,567.89"
        return fmt.replace(',', 'X').replace('.', ',').replace('X', '.')  # → "1.234.567,89"
    except (TypeError, ValueError):
        return str(value)

templates.env.filters['fmtnum'] = _fmtnum

BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
DB_PATH       = os.path.join(BASE_DIR, "scraper_hub.db")
MEDIA_DIR     = os.path.join(BASE_DIR, "media")
os.makedirs(MEDIA_DIR, exist_ok=True)

# ── cloud config (set via env vars in production) ─────────────────────────────
DATABASE_URL    = os.environ.get("DATABASE_URL", "")          # Supabase Postgres URL
SUPABASE_URL    = os.environ.get("SUPABASE_URL", "")          # https://xxx.supabase.co
SUPABASE_KEY    = os.environ.get("SUPABASE_KEY", "")          # service_role key
SUPABASE_BUCKET = os.environ.get("SUPABASE_BUCKET", "media")  # storage bucket name

_supabase_client = None
def _get_supabase():
    global _supabase_client
    if _supabase_client is None and _SUPABASE_AVAILABLE and SUPABASE_URL and SUPABASE_KEY:
        _supabase_client = _supabase_create_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase_client


# ── database ──────────────────────────────────────────────────────────────────

class _PgRow:
    """sqlite3.Row-compatible wrapper: row['name'] and row[0] both work."""
    __slots__ = ('_d', '_vals')
    def __init__(self, description, values):
        self._vals = tuple(values)
        self._d    = {d[0]: v for d, v in zip(description, values)}
    def __getitem__(self, key):
        return self._vals[key] if isinstance(key, int) else self._d[key]
    def __iter__(self):
        return iter(self._vals)
    def keys(self):
        return list(self._d.keys())
    def get(self, key, default=None):
        return self._d.get(key, default)

class _PgCursor:
    def __init__(self, cur):
        self._c = cur
    def _wrap(self, row):
        return _PgRow(self._c.description, row) if row is not None else None
    def fetchone(self):
        row = self._c.fetchone()
        return self._wrap(row) if row else None
    def fetchall(self):
        desc = self._c.description
        return [_PgRow(desc, r) for r in self._c.fetchall()]
    def __iter__(self):
        desc = self._c.description
        for row in self._c:
            yield _PgRow(desc, row)
    @property
    def rowcount(self):
        return self._c.rowcount

class _PgConnection:
    """psycopg2 connection with a sqlite3-compatible surface."""
    def __init__(self, dsn: str):
        self._conn = psycopg2.connect(dsn)

    # Convert SQLite SQL dialects to PostgreSQL on the fly
    @staticmethod
    def _fix(sql: str) -> str:
        import re
        sql = sql.replace("?", "%s")
        sql = sql.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
        sql = sql.replace("(date('now'))", "CURRENT_DATE")
        # date('now', '-N days/months/years') → CURRENT_DATE - INTERVAL 'N days/months/years'
        sql = re.sub(
            r"date\('now'\s*,\s*['\"]([+-]?\d+)\s+(days?|months?|years?)['\"](?:\s*,\s*['\"][^'\"]*['\"])*\)",
            lambda m: f"(CURRENT_DATE + INTERVAL '{m.group(1)} {m.group(2)}')::text",
            sql, flags=re.IGNORECASE
        )
        sql = sql.replace("date('now')", "CURRENT_DATE::text")
        if "INSERT OR IGNORE INTO" in sql:
            sql = sql.replace("INSERT OR IGNORE INTO", "INSERT INTO")
            sql = sql.rstrip(";").rstrip() + " ON CONFLICT DO NOTHING"
        # julianday(x) - julianday(y) → (CURRENT_DATE - y::date) days as float
        # Pattern: julianday('now') - julianday(col)
        sql = re.sub(
            r"julianday\('now'\)\s*-\s*julianday\((\w+)\)",
            r"EXTRACT(EPOCH FROM (NOW() - \1::timestamp))/86400",
            sql
        )
        return sql

    def execute(self, sql: str, params=None):
        cur = self._conn.cursor()
        cur.execute(self._fix(sql), params or ())
        return _PgCursor(cur)

    def executemany(self, sql: str, params_list):
        cur = self._conn.cursor()
        cur.executemany(self._fix(sql), params_list)
        return _PgCursor(cur)

    def executescript(self, script: str):
        """Run a multi-statement script (PostgreSQL equivalent of sqlite3.executescript)."""
        script = self._fix(script)
        cur = self._conn.cursor()
        for stmt in script.split(";"):
            stmt = stmt.strip()
            if stmt:
                try:
                    cur.execute(stmt)
                except Exception:
                    self._conn.rollback()
                    raise
        return _PgCursor(cur)

    def commit(self):  self._conn.commit()
    def close(self):   self._conn.close()
    def __enter__(self): return self
    def __exit__(self, *_): self.close()


def get_db():
    if DATABASE_URL and _PSYCOPG2_AVAILABLE:
        return _PgConnection(DATABASE_URL)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS sites (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            name                TEXT    NOT NULL,
            url                 TEXT    NOT NULL UNIQUE,
            css_selector        TEXT    DEFAULT '',
            article_url_pattern TEXT    DEFAULT '',
            use_stealth         INTEGER DEFAULT 0,
            active              INTEGER DEFAULT 1,
            last_scraped        TEXT,
            created_at          TEXT    DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS articles (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id          INTEGER REFERENCES sites(id) ON DELETE CASCADE,
            title            TEXT    NOT NULL,
            article_url      TEXT    DEFAULT '',
            author           TEXT    DEFAULT '',
            category         TEXT    DEFAULT '',
            company_property TEXT    DEFAULT '',
            company_brand    TEXT    DEFAULT '',
            company_agency   TEXT    DEFAULT '',
            summary          TEXT    DEFAULT '',
            body             TEXT    DEFAULT '',
            scraped_at       TEXT    DEFAULT CURRENT_TIMESTAMP,
            scrape_date      TEXT    DEFAULT (date('now'))
        );
        CREATE TABLE IF NOT EXISTS site_traffic (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id           INTEGER REFERENCES sites(id) ON DELETE CASCADE,
            measured_at       TEXT DEFAULT CURRENT_TIMESTAMP,
            monthly_visits    INTEGER DEFAULT 0,
            bounce_rate       REAL    DEFAULT 0,
            pages_per_visit   REAL    DEFAULT 0,
            avg_duration_sec  INTEGER DEFAULT 0,
            global_rank       INTEGER DEFAULT 0,
            country_rank      INTEGER DEFAULT 0,
            source            TEXT    DEFAULT '',
            error             TEXT    DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS saved_searches (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL,
            query      TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS email_logs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            subject     TEXT    DEFAULT '',
            recipients  TEXT    DEFAULT '',
            article_ids TEXT    DEFAULT '',
            body        TEXT    DEFAULT '',
            sent_at     TEXT    DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS clients (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL UNIQUE,
            notes      TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS client_searches (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id  INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
            name       TEXT NOT NULL,
            notes      TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS client_brands (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id  INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
            search_id  INTEGER REFERENCES client_searches(id) ON DELETE CASCADE,
            name       TEXT NOT NULL,
            query      TEXT NOT NULL,
            notes      TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS scrape_queue (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            site_id      INTEGER NOT NULL REFERENCES sites(id) ON DELETE CASCADE,
            job_date     TEXT NOT NULL,
            status       TEXT NOT NULL DEFAULT 'pending',
            priority     INTEGER NOT NULL DEFAULT 5,
            enqueued_at  TEXT DEFAULT CURRENT_TIMESTAMP,
            started_at   TEXT,
            completed_at TEXT,
            worker_id    TEXT,
            new_articles INTEGER DEFAULT 0,
            error        TEXT DEFAULT ''
        );
        CREATE UNIQUE INDEX IF NOT EXISTS uq_scrape_queue_site_date
            ON scrape_queue(site_id, job_date);
        CREATE INDEX IF NOT EXISTS idx_scrape_queue_status
            ON scrape_queue(status, priority DESC, enqueued_at);
    """)
    _pg = DATABASE_URL and _PSYCOPG2_AVAILABLE
    _ifne = "IF NOT EXISTS " if _pg else ""   # PostgreSQL supports IF NOT EXISTS on ALTER
    for col, ddl in [
        ("article_url_pattern",   f"ALTER TABLE sites ADD COLUMN {_ifne}article_url_pattern TEXT DEFAULT ''"),
        ("scrape_frequency_days", f"ALTER TABLE sites ADD COLUMN {_ifne}scrape_frequency_days INTEGER DEFAULT 0"),
        ("auto_enrich",           f"ALTER TABLE sites ADD COLUMN {_ifne}auto_enrich INTEGER DEFAULT 0"),
        ("last_scrape_status",    f"ALTER TABLE sites ADD COLUMN {_ifne}last_scrape_status TEXT DEFAULT ''"),
        ("last_scrape_error",     f"ALTER TABLE sites ADD COLUMN {_ifne}last_scrape_error TEXT DEFAULT ''"),
        ("last_scrape_count",     f"ALTER TABLE sites ADD COLUMN {_ifne}last_scrape_count INTEGER DEFAULT 0"),
        ("body",                  f"ALTER TABLE articles ADD COLUMN {_ifne}body TEXT DEFAULT ''"),
        ("ots",                   f"ALTER TABLE articles ADD COLUMN {_ifne}ots INTEGER DEFAULT 0"),
        ("vpe",                   f"ALTER TABLE articles ADD COLUMN {_ifne}vpe REAL DEFAULT 0"),
        ("images",                f"ALTER TABLE articles ADD COLUMN {_ifne}images TEXT DEFAULT ''"),
        ("videos",                f"ALTER TABLE articles ADD COLUMN {_ifne}videos TEXT DEFAULT ''"),
        ("last_traffic_check",    f"ALTER TABLE sites ADD COLUMN {_ifne}last_traffic_check TEXT"),
        ("monthly_visits_manual", f"ALTER TABLE sites ADD COLUMN {_ifne}monthly_visits_manual INTEGER DEFAULT 0"),
        ("traffic_frequency_days",f"ALTER TABLE sites ADD COLUMN {_ifne}traffic_frequency_days INTEGER DEFAULT 5"),
        ("category",              f"ALTER TABLE sites ADD COLUMN {_ifne}category TEXT DEFAULT ''"),
        ("country",               f"ALTER TABLE sites ADD COLUMN {_ifne}country TEXT DEFAULT 'WW'"),
        ("search_id",             f"ALTER TABLE client_brands ADD COLUMN {_ifne}search_id INTEGER REFERENCES client_searches(id) ON DELETE CASCADE"),
        ("keywords",              f"ALTER TABLE clients ADD COLUMN {_ifne}keywords TEXT DEFAULT ''"),
        ("search_keywords",       f"ALTER TABLE client_searches ADD COLUMN {_ifne}keywords TEXT DEFAULT ''"),
        ("sentiment",             f"ALTER TABLE articles ADD COLUMN {_ifne}sentiment TEXT DEFAULT ''"),
        ("cookies_json",          f"ALTER TABLE sites ADD COLUMN {_ifne}cookies_json TEXT DEFAULT ''"),
        ("local_images",          f"ALTER TABLE articles ADD COLUMN {_ifne}local_images TEXT DEFAULT ''"),
        ("local_videos",          f"ALTER TABLE articles ADD COLUMN {_ifne}local_videos TEXT DEFAULT ''"),
        ("cdp_url",               f"ALTER TABLE sites ADD COLUMN {_ifne}cdp_url TEXT DEFAULT ''"),
        ("last_working_tier",     f"ALTER TABLE sites ADD COLUMN {_ifne}last_working_tier INTEGER DEFAULT 0"),
    ]:
        try:
            conn.execute(ddl)
            if _pg:
                conn.commit()   # each ALTER needs its own transaction in PostgreSQL
        except Exception:
            if _pg:
                conn._conn.rollback()   # reset after failed ALTER (column already exists)
    # Indexes (run after ALTERs so referenced columns exist)
    for ddl in (
        "CREATE INDEX IF NOT EXISTS idx_client_searches_client ON client_searches(client_id)",
        "CREATE INDEX IF NOT EXISTS idx_client_brands_client  ON client_brands(client_id)",
        "CREATE INDEX IF NOT EXISTS idx_client_brands_search  ON client_brands(search_id)",
    ):
        try: conn.execute(ddl)
        except Exception: pass

    if conn.execute("SELECT COUNT(*) FROM sites").fetchone()[0] == 0:
        conn.execute(
            "INSERT INTO sites (name, url, article_url_pattern, category, country) VALUES "
            "('2Playbook', 'https://www.2playbook.com/', '_102.html', 'Sports B2B', 'ES')"
        )
    conn.commit()
    conn.close()


# ── countries (ISO Alpha-2 + name + emoji flag) ───────────────────────────────
COUNTRIES = [
    # ── Special ──────────────────────────────────────────────────────────────
    ("WW", "Global",                 "🌍"),
    # ── Europa ───────────────────────────────────────────────────────────────
    ("ES", "España",                 "🇪🇸"),
    ("AD", "Andorra",                "🇦🇩"),
    ("PT", "Portugal",               "🇵🇹"),
    ("FR", "Francia",                "🇫🇷"),
    ("GB", "Reino Unido",            "🇬🇧"),
    ("DE", "Alemania",               "🇩🇪"),
    ("IT", "Italia",                 "🇮🇹"),
    ("NL", "Países Bajos",           "🇳🇱"),
    ("BE", "Bélgica",                "🇧🇪"),
    ("CH", "Suiza",                  "🇨🇭"),
    ("AT", "Austria",                "🇦🇹"),
    ("SE", "Suecia",                 "🇸🇪"),
    ("NO", "Noruega",                "🇳🇴"),
    ("DK", "Dinamarca",              "🇩🇰"),
    ("FI", "Finlandia",              "🇫🇮"),
    ("PL", "Polonia",                "🇵🇱"),
    ("CZ", "República Checa",        "🇨🇿"),
    ("SK", "Eslovaquia",             "🇸🇰"),
    ("HU", "Hungría",                "🇭🇺"),
    ("RO", "Rumanía",                "🇷🇴"),
    ("BG", "Bulgaria",               "🇧🇬"),
    ("GR", "Grecia",                 "🇬🇷"),
    ("HR", "Croacia",                "🇭🇷"),
    ("RS", "Serbia",                 "🇷🇸"),
    ("SI", "Eslovenia",              "🇸🇮"),
    ("BA", "Bosnia-Herzegovina",     "🇧🇦"),
    ("ME", "Montenegro",             "🇲🇪"),
    ("MK", "Macedonia del Norte",    "🇲🇰"),
    ("AL", "Albania",                "🇦🇱"),
    ("UA", "Ucrania",                "🇺🇦"),
    ("BY", "Bielorrusia",            "🇧🇾"),
    ("MD", "Moldavia",               "🇲🇩"),
    ("RU", "Rusia",                  "🇷🇺"),
    ("EE", "Estonia",                "🇪🇪"),
    ("LV", "Letonia",                "🇱🇻"),
    ("LT", "Lituania",               "🇱🇹"),
    ("IE", "Irlanda",                "🇮🇪"),
    ("IS", "Islandia",               "🇮🇸"),
    ("LU", "Luxemburgo",             "🇱🇺"),
    ("MC", "Mónaco",                 "🇲🇨"),
    ("LI", "Liechtenstein",          "🇱🇮"),
    ("MT", "Malta",                  "🇲🇹"),
    ("CY", "Chipre",                 "🇨🇾"),
    ("TR", "Turquía",                "🇹🇷"),
    ("GE", "Georgia",                "🇬🇪"),
    ("AM", "Armenia",                "🇦🇲"),
    ("AZ", "Azerbaiyán",             "🇦🇿"),
    ("KZ", "Kazajistán",             "🇰🇿"),
    # ── América del Norte ─────────────────────────────────────────────────────
    ("US", "Estados Unidos",         "🇺🇸"),
    ("CA", "Canadá",                 "🇨🇦"),
    ("MX", "México",                 "🇲🇽"),
    ("GT", "Guatemala",              "🇬🇹"),
    ("HN", "Honduras",               "🇭🇳"),
    ("SV", "El Salvador",            "🇸🇻"),
    ("NI", "Nicaragua",              "🇳🇮"),
    ("CR", "Costa Rica",             "🇨🇷"),
    ("PA", "Panamá",                 "🇵🇦"),
    ("CU", "Cuba",                   "🇨🇺"),
    ("DO", "República Dominicana",   "🇩🇴"),
    ("PR", "Puerto Rico",            "🇵🇷"),
    ("JM", "Jamaica",                "🇯🇲"),
    ("HT", "Haití",                  "🇭🇹"),
    ("TT", "Trinidad y Tobago",      "🇹🇹"),
    # ── América del Sur ───────────────────────────────────────────────────────
    ("AR", "Argentina",              "🇦🇷"),
    ("BR", "Brasil",                 "🇧🇷"),
    ("CO", "Colombia",               "🇨🇴"),
    ("CL", "Chile",                  "🇨🇱"),
    ("PE", "Perú",                   "🇵🇪"),
    ("UY", "Uruguay",                "🇺🇾"),
    ("PY", "Paraguay",               "🇵🇾"),
    ("BO", "Bolivia",                "🇧🇴"),
    ("EC", "Ecuador",                "🇪🇨"),
    ("VE", "Venezuela",              "🇻🇪"),
    ("GY", "Guyana",                 "🇬🇾"),
    ("SR", "Surinam",                "🇸🇷"),
    # ── África ────────────────────────────────────────────────────────────────
    ("MA", "Marruecos",              "🇲🇦"),
    ("DZ", "Argelia",                "🇩🇿"),
    ("TN", "Túnez",                  "🇹🇳"),
    ("LY", "Libia",                  "🇱🇾"),
    ("EG", "Egipto",                 "🇪🇬"),
    ("NG", "Nigeria",                "🇳🇬"),
    ("ZA", "Sudáfrica",              "🇿🇦"),
    ("KE", "Kenia",                  "🇰🇪"),
    ("ET", "Etiopía",                "🇪🇹"),
    ("GH", "Ghana",                  "🇬🇭"),
    ("SN", "Senegal",                "🇸🇳"),
    ("CI", "Costa de Marfil",        "🇨🇮"),
    ("CM", "Camerún",                "🇨🇲"),
    ("TZ", "Tanzania",               "🇹🇿"),
    ("UG", "Uganda",                 "🇺🇬"),
    ("AO", "Angola",                 "🇦🇴"),
    ("MZ", "Mozambique",             "🇲🇿"),
    ("ZM", "Zambia",                 "🇿🇲"),
    ("ZW", "Zimbabue",               "🇿🇼"),
    ("MG", "Madagascar",             "🇲🇬"),
    ("MU", "Mauricio",               "🇲🇺"),
    # ── Asia ─────────────────────────────────────────────────────────────────
    ("JP", "Japón",                  "🇯🇵"),
    ("CN", "China",                  "🇨🇳"),
    ("KR", "Corea del Sur",          "🇰🇷"),
    ("IN", "India",                  "🇮🇳"),
    ("ID", "Indonesia",              "🇮🇩"),
    ("PH", "Filipinas",              "🇵🇭"),
    ("TH", "Tailandia",              "🇹🇭"),
    ("VN", "Vietnam",                "🇻🇳"),
    ("MY", "Malasia",                "🇲🇾"),
    ("SG", "Singapur",               "🇸🇬"),
    ("PK", "Pakistán",               "🇵🇰"),
    ("BD", "Bangladés",              "🇧🇩"),
    ("LK", "Sri Lanka",              "🇱🇰"),
    ("NP", "Nepal",                  "🇳🇵"),
    ("MM", "Myanmar",                "🇲🇲"),
    ("KH", "Camboya",                "🇰🇭"),
    ("SA", "Arabia Saudí",           "🇸🇦"),
    ("AE", "Emiratos Árabes",        "🇦🇪"),
    ("QA", "Catar",                  "🇶🇦"),
    ("KW", "Kuwait",                 "🇰🇼"),
    ("BH", "Baréin",                 "🇧🇭"),
    ("OM", "Omán",                   "🇴🇲"),
    ("IQ", "Irak",                   "🇮🇶"),
    ("IR", "Irán",                   "🇮🇷"),
    ("IL", "Israel",                 "🇮🇱"),
    ("JO", "Jordania",               "🇯🇴"),
    ("LB", "Líbano",                 "🇱🇧"),
    ("SY", "Siria",                  "🇸🇾"),
    ("AF", "Afganistán",             "🇦🇫"),
    ("UZ", "Uzbekistán",             "🇺🇿"),
    ("TM", "Turkmenistán",           "🇹🇲"),
    ("TJ", "Tayikistán",             "🇹🇯"),
    ("KG", "Kirguistán",             "🇰🇬"),
    ("MN", "Mongolia",               "🇲🇳"),
    ("HK", "Hong Kong",              "🇭🇰"),
    ("TW", "Taiwán",                 "🇹🇼"),
    # ── Oceanía ──────────────────────────────────────────────────────────────
    ("AU", "Australia",              "🇦🇺"),
    ("NZ", "Nueva Zelanda",          "🇳🇿"),
    ("FJ", "Fiyi",                   "🇫🇯"),
    ("PG", "Papúa Nueva Guinea",     "🇵🇬"),
]
COUNTRY_BY_CODE = {c: (n, f) for c, n, f in COUNTRIES}


init_db()
app.mount("/media", StaticFiles(directory=MEDIA_DIR), name="media")
app.mount("/static", StaticFiles(directory="static"), name="static")


# ── media download ─────────────────────────────────────────────────────────────

def _download_media(urls: list, site_id: int, article_id: int, kind: str) -> list:
    """Download media URLs and store them.
    Cloud (SUPABASE_URL set): uploads to Supabase Storage, returns public URLs.
    Local: saves to media/{site_id}/{article_id}/, returns relative paths.
    kind is 'img' or 'vid'."""
    sb = _get_supabase()
    folder = os.path.join(MEDIA_DIR, str(site_id), str(article_id))
    if not sb:
        os.makedirs(folder, exist_ok=True)
    saved = []
    for i, url in enumerate(urls, 1):
        try:
            ext = ""
            parsed_path = url.split("?")[0]
            if "." in parsed_path.rsplit("/", 1)[-1]:
                ext = "." + parsed_path.rsplit(".", 1)[-1][:5].lower()
            if not ext:
                ext = ".jpg" if kind == "img" else ".mp4"
            filename = f"{kind}_{i:03d}{ext}"
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15) as r:
                data = r.read(10 * 1024 * 1024)   # cap 10 MB
            if sb:
                storage_path = f"{site_id}/{article_id}/{filename}"
                mime = mimetypes.guess_type(filename)[0] or "application/octet-stream"
                sb.storage.from_(SUPABASE_BUCKET).upload(
                    storage_path, data,
                    {"content-type": mime, "x-upsert": "true"}
                )
                saved.append(sb.storage.from_(SUPABASE_BUCKET).get_public_url(storage_path))
            else:
                dest = os.path.join(folder, filename)
                if not os.path.exists(dest):
                    with open(dest, "wb") as f:
                        f.write(data)
                saved.append(f"{site_id}/{article_id}/{filename}")
        except Exception as e:
            print(f"[media] skip {url}: {e}")
    return saved


# ── scraping ──────────────────────────────────────────────────────────────────

def _page_signal_strength(page) -> int:
    """Return number of <a href> links — the strongest signal that we got real content."""
    if page is None:
        return 0
    try:
        return len(page.css("a[href]") or [])
    except Exception:
        return 0


def _parse_cookies(cookies_json: str) -> dict:
    """Parse cookies stored as JSON object or Netscape/header string."""
    if not cookies_json or not cookies_json.strip():
        return {}
    s = cookies_json.strip()
    try:
        v = json.loads(s)
        if isinstance(v, dict):
            return v
        if isinstance(v, list):
            return {c["name"]: c["value"] for c in v if "name" in c and "value" in c}
    except Exception:
        pass
    # fallback: "name=value; name2=value2" header format
    out = {}
    for part in s.split(";"):
        part = part.strip()
        if "=" in part:
            k, _, v = part.partition("=")
            out[k.strip()] = v.strip()
    return out


def _cookie_consent_action(page):
    """Auto-click cookie/GDPR consent accept buttons on common CMP dialogs."""
    # Wait up to 4s for any consent dialog to appear before trying selectors
    try:
        page.wait_for_timeout(2000)
    except Exception:
        pass

    # ID/class-based selectors first (fast, unambiguous)
    ID_SELECTORS = [
        "#onetrust-accept-btn-handler",          # OneTrust (Reuters, many news)
        ".onetrust-accept-btn-handler",
        "#didomi-notice-agree-button",            # Didomi
        "[data-testid='consent-accept']",
        "[data-tracking-opt-in-accept]",          # Quantcast
        "[class*='accept-all']",
        "[class*='acceptAll']",
        "[class*='cookie-accept']",
        "[id*='accept-all']",
        "[id*='acceptAll']",
        # Cookiebot (AS.com, many Spanish sites)
        "#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll",
        ".CybotCookiebotDialogBodyButton",
        # Sourcepoint / WSJ
        "[title='I Agree']",
        "[title='Accept']",
        "[title='Allow All']",
    ]
    # Text-based selectors ordered most-specific → least-specific
    TEXT_SELECTORS = [
        # Exact full-text matches first (most reliable)
        "button:text-is('Allow All')",            # Reuters
        "button:text-is('I Agree')",              # WSJ
        "button:text-is('Accept all')",
        "button:text-is('Accept All')",
        "button:text-is('Accept')",               # DAZN
        "button:text-is('Agree')",
        "button:text-is('Aceptar todo')",
        "button:text-is('Aceptar todas')",
        "button:text-is('Aceptar')",
        "button:text-is('Akkoord')",              # DPG Media
        "button:text-is('Tout accepter')",
        "button:text-is('Alle akzeptieren')",
        "button:text-is('Alle accepteren')",
        # Substring fallbacks
        "button:has-text('Allow All')",           # Reuters fallback
        "button:has-text('I Agree')",             # WSJ fallback
        "button:has-text('Accept all')",
        "button:has-text('Aceptar todo')",
        "button:has-text('Tout accepter')",
        "button:has-text('Akkoord')",
        "button:has-text('I agree')",
        "button:has-text('Accept')",
        "button:has-text('Aceptar')",
        # a-tag fallbacks (some CMPs use links)
        "a:has-text('Accept all')",
        "a:has-text('Allow All')",
        "a:has-text('I Agree')",
        "a:has-text('Akkoord')",
    ]
    for sel in ID_SELECTORS + TEXT_SELECTORS:
        try:
            btn = page.locator(sel).first
            if btn.is_visible(timeout=2000):
                btn.scroll_into_view_if_needed()
                btn.click()
                page.wait_for_timeout(1500)
                print(f"[consent] clicked: {sel}")
                return
        except Exception:
            continue


def _fetch_page(url: str, use_stealth: bool, cookies: dict | None = None,
                cdp_url: str | None = None, start_tier: int = 1):
    """Fetch a page with auto-escalation:
       Fetcher.get → DynamicFetcher.fetch → StealthyFetcher.fetch
       When cdp_url is set, skip all tiers and drive the already-running
       Chrome instance directly — full profile, cookies, and extensions included.

    Performance notes:
      - disable_resources=True skips CSS/images/fonts during browser render (faster HTML load)
      - network_idle is avoided on auto-escalation; a short fixed wait is used instead
        because heavy sites (WSJ, NYT) keep background XHR alive indefinitely
      - CDP mode uses network_idle because the user's real Chrome is already warmed up
    """
    Fetcher, DynamicFetcher, StealthyFetcher = _get_fetchers()
    ck = cookies or {}

    # CDP mode — connect to user's running Chrome; no bot detection possible
    if cdp_url:
        print(f"[scrape] CDP {cdp_url} → {url}")
        try:
            page = StealthyFetcher.fetch(url, cdp_url=cdp_url,
                                          disable_resources=False,
                                          network_idle=True, timeout=45000,
                                          page_action=_cookie_consent_action)
            return page, 0   # tier 0 = CDP
        except Exception as e:
            print(f"[scrape] CDP fetch failed: {e}")
            return None, 0

    if use_stealth:
        print(f"[scrape] forced stealthy-fetch {url}")
        try:
            page = StealthyFetcher.fetch(url, headless=True,
                                          disable_resources=True,
                                          wait=2500, timeout=40000,
                                          cookies=ck or None,
                                          page_action=_cookie_consent_action)
            return page, 3
        except Exception as e:
            print(f"[scrape] stealthy-fetch failed: {e}")
            return None, 3

    best, best_tier = None, 1

    # 1) Plain HTTP — fast, no browser overhead
    if start_tier <= 1:
        try:
            print(f"[scrape] get {url}")
            page   = Fetcher.get(url, stealthy_headers=True, timeout=12)
            status = getattr(page, "status", 200)
            links  = _page_signal_strength(page)
            print(f"[scrape]   → status={status}, links={links}")
            if status < 400 and links >= 10:
                return page, 1
            if links > 0:
                best, best_tier = page, 1
        except Exception as e:
            print(f"[scrape] get failed: {e}")

    # 2) Headless browser with JS — skip images/CSS, short fixed wait
    if start_tier <= 2:
        try:
            print(f"[scrape] fetch (browser) {url}")
            page  = DynamicFetcher.fetch(url, disable_resources=True,
                                          wait=1500, timeout=20000,
                                          cookies=ck or None,
                                          page_action=_cookie_consent_action)
            links = _page_signal_strength(page)
            print(f"[scrape]   → links={links}")
            if links >= 10:
                return page, 2
            if links > _page_signal_strength(best):
                best, best_tier = page, 2
        except Exception as e:
            print(f"[scrape] fetch failed: {e}")

    # 3) Stealth browser (anti-bot) — last resort, still skip non-essential resources
    try:
        print(f"[scrape] stealthy-fetch {url}")
        page  = StealthyFetcher.fetch(url, headless=True,
                                       disable_resources=True,
                                       wait=2000, timeout=25000,
                                       cookies=ck or None,
                                       page_action=_cookie_consent_action)
        links = _page_signal_strength(page)
        print(f"[scrape]   → links={links}")
        if links > _page_signal_strength(best):
            best, best_tier = page, 3
    except Exception as e:
        print(f"[scrape] stealthy-fetch failed: {e}")

    return best, best_tier


def _fetch_rss(url: str) -> list[dict]:
    """Parse an RSS/Atom feed and return list of {title, url}. Returns [] on failure."""
    import xml.etree.ElementTree as ET
    import urllib.request
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (compatible; Googlebot/2.1)",
            "Accept": "application/rss+xml, application/xml, text/xml, */*",
        })
        with urllib.request.urlopen(req, timeout=15) as r:
            raw = r.read()
        root = ET.fromstring(raw)
        ns = {"atom": "http://www.w3.org/2005/Atom"}
        items = []
        # RSS 2.0
        for item in root.findall(".//item"):
            title = (item.findtext("title") or "").strip()
            link  = (item.findtext("link") or "").strip()
            if title and link:
                items.append({"title": title, "url": link})
        # Atom
        if not items:
            for entry in root.findall(".//atom:entry", ns):
                title = (entry.findtext("atom:title", namespaces=ns) or "").strip()
                link_el = entry.find("atom:link", ns)
                link = (link_el.attrib.get("href") if link_el is not None else "") or ""
                if title and link:
                    items.append({"title": title, "url": link})
        print(f"[rss] {url} → {len(items)} items")
        return items
    except Exception as e:
        print(f"[rss] failed {url}: {e}")
        return []


def fetch_site_links(url: str, css_selector: str, use_stealth: bool,
                     cookies: dict | None = None, cdp_url: str | None = None,
                     start_tier: int = 1) -> tuple[list[dict], int]:
    """Return (links, tier_used). links is list of {title, url}."""
    # RSS/Atom feeds — bypass HTML scraping entirely
    _url_lower = url.lower().split("?")[0]
    if any(_url_lower.endswith(ext) for ext in (".xml", ".rss", ".atom")) or "/rss" in _url_lower or "/feed" in _url_lower:
        return _fetch_rss(url), 1

    page, tier = _fetch_page(url, use_stealth, cookies=cookies, cdp_url=cdp_url, start_tier=start_tier)
    if not page:
        return [], tier

    # Narrow scope to CSS selector if given, then collect <a> inside it
    if css_selector:
        scope = page.css(css_selector) or []
        raw_links = []
        for el in scope:
            raw_links.extend(el.css("a[href]") or [])
        if not raw_links:                          # fallback: all links
            raw_links = page.css("a[href]") or []
    else:
        raw_links = page.css("a[href]") or []

    out = []
    for a in raw_links:
        href  = (a.attrib.get("href") or "").strip()
        title = " ".join((a.get_all_text() or "").split()).strip()
        if href and title:
            out.append({"url": urljoin(url, href), "title": title})
    return out, tier


def _urllib_extract(url: str) -> dict:
    """Fast stdlib-only extractor: urllib.request + html.parser.
    No browser, no curl. Works for the majority of news sites that serve plain HTML.
    Returns dict with body, title, date, images, videos."""
    import urllib.request as _ur
    import gzip as _gz
    from html.parser import HTMLParser

    _SKIP_TAGS  = {'script', 'style', 'nav', 'header', 'footer', 'aside',
                   'noscript', 'iframe', 'template', 'svg', 'figure'}
    _SKIP_IMG   = ('/logo', '/icon', '/avatar', 'sprite', '/social',
                   '/ad/', '/ads/', '/tracking', 'blank.gif', '1x1.')

    class _P(HTMLParser):
        def __init__(self, base):
            super().__init__(convert_charrefs=True)
            self.base    = base
            self.parts   : list[str] = []
            self.imgs    : list[str] = []
            self.vids    : list[str] = []
            self.title   = ''
            self.date    = ''
            self._skip   = 0
            self._in_title = False

        def handle_starttag(self, tag, attrs):
            a = dict(attrs)
            if tag in _SKIP_TAGS:
                self._skip += 1
            if tag == 'title':
                self._in_title = True
            if tag == 'meta':
                prop    = (a.get('property') or a.get('name') or '').lower()
                content = (a.get('content') or '').strip()
                if prop in ('og:title', 'twitter:title') and not self.title:
                    self.title = content
                if prop in ('article:published_time', 'pubdate', 'date',
                            'dc.date', 'parsely-pub-date') and not self.date:
                    self.date = content[:10]
            if tag == 'time':
                dt = (a.get('datetime') or '').strip()
                if dt and not self.date:
                    self.date = dt[:10]
            if tag == 'img' and not self._skip:
                src = (a.get('src') or '').strip()
                if src and not src.startswith('data:'):
                    abs_src = urljoin(self.base, src)
                    low = abs_src.lower()
                    w = int(a.get('width') or 0)
                    h = int(a.get('height') or 0)
                    if (not any(t in low for t in _SKIP_IMG)
                            and not low.split('?')[0].endswith(('.ico', '.svg'))
                            and not (0 < w < 80 and 0 < h < 80)):
                        self.imgs.append(abs_src)
            if tag in ('video', 'source'):
                src = (a.get('src') or '').strip()
                if src:
                    self.vids.append(urljoin(self.base, src))
            if tag == 'iframe':
                src = (a.get('src') or '').strip()
                if any(x in src for x in ('youtube', 'youtu.be', 'vimeo', 'dailymotion')):
                    self.vids.append(urljoin(self.base, src))

        def handle_endtag(self, tag):
            if tag in _SKIP_TAGS:
                self._skip = max(0, self._skip - 1)
            if tag == 'title':
                self._in_title = False

        def handle_data(self, data):
            text = data.strip()
            if self._in_title and text and not self.title:
                self.title = text
            if text and not self._skip:
                self.parts.append(text)

    headers = {
        'User-Agent':      'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
        'Accept':          'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Connection':      'keep-alive',
        'Cache-Control':   'no-cache',
        'Pragma':          'no-cache',
    }
    import http.cookiejar as _cj
    cookie_jar = _cj.CookieJar()
    opener = _ur.build_opener(_ur.HTTPCookieProcessor(cookie_jar))
    req = _ur.Request(url, headers=headers)
    with opener.open(req, timeout=12) as resp:
        raw = resp.read()
        ct  = resp.headers.get('Content-Type', '')
        ce  = resp.headers.get('Content-Encoding', '')
        if ce == 'gzip' or (len(raw) > 2 and raw[:2] == b'\x1f\x8b'):
            try:
                raw = _gz.decompress(raw)
            except Exception:
                pass
        charset = 'utf-8'
        if 'charset=' in ct:
            charset = ct.split('charset=')[-1].strip().split(';')[0].split('"')[0]
        try:
            html = raw.decode(charset, errors='replace')
        except (LookupError, UnicodeDecodeError):
            html = raw.decode('utf-8', errors='replace')

    parser = _P(url)
    parser.feed(html)
    body = '\n'.join(parser.parts)
    return {
        'body':   body,
        'title':  parser.title,
        'date':   parser.date,
        'images': parser.imgs[:20],
        'videos': parser.vids[:5],
        'tier':   0,
    }


def _page_title(page) -> str:
    """Extract best title from a scraped page."""
    for sel, attr in [
        ('meta[property="og:title"]',     "content"),
        ('meta[name="twitter:title"]',    "content"),
        ('meta[name="DC.title"]',         "content"),
    ]:
        for m in (page.css(sel) or []):
            v = (m.attrib.get(attr) or "").strip()
            if v: return v
    for t in (page.css("title") or []):
        v = (t.text or "").strip()
        if v: return v
    for h in (page.css("h1") or []):
        v = (h.text or "").strip()
        if v: return v[:250]
    return ""


def _page_date(page) -> str:
    """Extract best publish date (YYYY-MM-DD) from a scraped page."""
    # 1. Meta tags
    for sel, attr in [
        ('meta[property="article:published_time"]', "content"),
        ('meta[name="pubdate"]',                    "content"),
        ('meta[name="date"]',                       "content"),
        ('meta[name="DC.date"]',                    "content"),
        ('meta[itemprop="datePublished"]',           "content"),
        ('meta[itemprop="datePublished"]',           "datetime"),
    ]:
        for m in (page.css(sel) or []):
            v = (m.attrib.get(attr) or "").strip()
            if v: return v[:10]
    # 2. <time datetime="...">
    for t in (page.css("time[datetime]") or []):
        v = (t.attrib.get("datetime") or "").strip()
        if v: return v[:10]
    # 3. JSON-LD
    import json as _j
    for script in (page.css('script[type="application/ld+json"]') or []):
        try:
            data = _j.loads(script.text or "")
            if isinstance(data, list): data = data[0]
            for key in ("datePublished", "dateCreated", "dateModified"):
                if key in data: return str(data[key])[:10]
        except Exception:
            pass
    return ""


def fetch_article_full(url: str, use_stealth: bool = False, cookies: dict | None = None,
                       cdp_url: str | None = None, start_tier: int = 1) -> dict:
    """Fetch article and return body text + title + date + image URLs + video URLs + tier_used.

    Strategy (when not using stealth/CDP):
      Tier 0  — stdlib urllib + html.parser (fast, no browser, no curl)
      Tier 1+ — Scrapling auto-escalation (plain HTTP → headless browser → stealth browser)
    Tier 0 is tried first; if it returns >= 300 chars of body we stop there.
    This avoids Scrapling's internal curl timeouts for sites that serve plain HTML.
    """
    # ── Tier 0: fast urllib fallback (skip for stealth/CDP modes) ────────────
    if not use_stealth and not cdp_url:
        try:
            fb = _urllib_extract(url)
            if len(fb.get("body", "")) >= 300:
                print(f"[scrape] urllib ok {url} ({len(fb['body'])} chars)")
                return fb
        except Exception as e:
            print(f"[scrape] urllib failed {url}: {e}")

    # ── Tiers 1-3: Scrapling ─────────────────────────────────────────────────
    page, tier = _fetch_page(url, use_stealth, cookies=cookies, cdp_url=cdp_url, start_tier=start_tier)
    if not page:
        return {"body": "", "title": "", "date": "", "images": [], "videos": [], "tier": tier}

    body = (page.get_all_text() or "").strip()

    # Junk patterns common in icons/logos/trackers
    SKIP_TOKENS = ("/logo", "/logos/", "/icon", "/icons/", "/avatar",
                   "sprite", "/social", "/badge", "/pixel", "/spinner",
                   "blank.gif", "1x1.", "/ad/", "/ads/", "tracking.")

    # ── Images ─────────────────────────────────────────────
    images, seen = [], set()
    for img in (page.css("img[src]") or []):
        src = (img.attrib.get("src") or "").strip()
        if not src or src.startswith("data:"):
            continue
        abs_url = urljoin(url, src)
        if abs_url in seen:
            continue
        low = abs_url.lower()
        if any(t in low for t in SKIP_TOKENS):
            continue
        if low.split("?", 1)[0].endswith((".svg", ".ico")):
            continue
        # If width/height provided and very small, skip (likely icon)
        try:
            w = int(img.attrib.get("width") or 0)
            h = int(img.attrib.get("height") or 0)
            if 0 < w < 80 and 0 < h < 80:
                continue
        except Exception:
            pass
        seen.add(abs_url)
        images.append(abs_url)

    # ── Videos ─────────────────────────────────────────────
    videos, seen_v = [], set()
    for v in (page.css("video[src]") or []):
        src = urljoin(url, (v.attrib.get("src") or "").strip())
        if src and src not in seen_v:
            seen_v.add(src); videos.append(src)
    for s in (page.css("video source[src]") or []):
        src = urljoin(url, (s.attrib.get("src") or "").strip())
        if src and src not in seen_v:
            seen_v.add(src); videos.append(src)
    for f in (page.css("iframe[src]") or []):
        src = (f.attrib.get("src") or "").strip()
        if any(x in src.lower() for x in
               ("youtube.com", "youtu.be", "vimeo.com", "dailymotion", "twitch.tv")):
            abs_url = urljoin(url, src)
            if abs_url not in seen_v:
                seen_v.add(abs_url); videos.append(abs_url)

    return {
        "body":   body,
        "title":  _page_title(page),
        "date":   _page_date(page),
        "images": images[:20],
        "videos": videos[:5],
        "tier":   tier,
    }


# Back-compat shim (still used if something legacy calls it)
def fetch_article_body(url: str, use_stealth: bool = False, cookies: dict | None = None,
                       cdp_url: str | None = None) -> str:
    return fetch_article_full(url, use_stealth, cookies=cookies, cdp_url=cdp_url).get("body", "")


def extract_articles(raw_links: list[dict], site) -> list[dict]:
    """Filter raw links to article candidates using the site's url_pattern or heuristics."""
    url_pattern = (site["article_url_pattern"] or "").strip()
    site_domain = urlparse(site["url"]).netloc

    SKIP = ("javascript:", "mailto:", ".pdf", ".xml",
            "facebook.", "twitter.", "linkedin.",
            "instagram.", "spotify.", "youtube.")

    # URL path segments that indicate non-article pages (policy, help, auth…)
    SKIP_PATHS = (
        "/privacy", "/cookie", "/terms", "/legal", "/aviso-legal",
        "/politica", "/política", "/datenschutz", "/impressum",
        "/help", "/support", "/faq", "/about", "/contact", "/contacto",
        "/subscribe", "/subscription", "/login", "/register", "/signup",
        "/account", "/profile", "/cart", "/checkout",
        "/sitemap", "/rss", "/feed",
        "/welcome", "/home",
    )

    seen, articles = set(), []
    for raw in raw_links:
        url   = raw["url"].strip()
        title = raw["title"].strip()

        if len(title) < 20 or url in seen:
            continue
        if any(x in url for x in SKIP):
            continue
        url_lower = url.lower()
        if any(seg in url_lower for seg in SKIP_PATHS):
            continue

        if url_pattern:
            if url_pattern not in url:
                continue
        else:
            # generic heuristic: same domain + at least 2 non-empty path segments
            parsed = urlparse(url)
            if parsed.netloc != site_domain:
                continue
            segs = [s for s in parsed.path.split("/") if s]
            if len(segs) < 2:
                continue

        parsed   = urlparse(url)
        segs     = [s for s in parsed.path.split("/") if s]
        category = segs[0] if segs else ""

        seen.add(url)
        articles.append(dict(
            title=title, article_url=url, category=category,
            author="", company_property="", company_brand="",
            company_agency="", summary="", body=""
        ))
    return articles[:60]


# ── traffic metrics ───────────────────────────────────────────────────────────

def _humanized_to_int(s: str) -> int:
    """Parse '115.8M' → 115_800_000, '12.3K' → 12_300, '3.2B' → 3_200_000_000."""
    if not s:
        return 0
    s = s.strip().upper().replace(",", "").replace(" ", "")
    m = re.match(r"([\d.]+)([KMB])?", s)
    if not m:
        return 0
    try:
        num = float(m.group(1))
    except ValueError:
        return 0
    mult = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}.get(m.group(2) or "", 1)
    return int(num * mult)


def _hms_to_sec(s: str) -> int:
    """Parse '00:02:37' → 157, '02:37' → 157."""
    if not s:
        return 0
    parts = s.strip().split(":")
    try:
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
    except ValueError:
        pass
    return 0


def fetch_similarweb_public(domain: str) -> dict:
    """Scrape the Similarweb public page using StealthyFetcher with CF solving."""
    _, _, StealthyFetcher = _get_fetchers()
    url = f"https://www.similarweb.com/website/{domain}/"
    out = {"source": "similarweb", "error": ""}
    try:
        page = StealthyFetcher.fetch(
            url, headless=True, solve_cloudflare=True,
            network_idle=True, wait=2000
        )
    except Exception as e:
        return {"source": "similarweb", "error": f"fetch: {str(e)[:200]}"}
    if not page:
        return {"source": "similarweb", "error": "no page returned"}

    text = page.get_all_text() or ""

    patterns = [
        ("monthly_visits",   r"Total\s+Visits[^0-9]+([\d.]+\s*[KMB]?)",          _humanized_to_int),
        ("bounce_rate",      r"Bounce\s+Rate[^\d.]+([\d.]+)\s*%",                lambda x: float(x)/100.0),
        ("pages_per_visit",  r"Pages\s+per\s+Visit[^\d.]+([\d.]+)",              float),
        ("avg_duration_sec", r"Avg\.?\s+(?:Visit\s+)?Duration[^\d:]+(\d+:\d+(?::\d+)?)", _hms_to_sec),
        ("global_rank",      r"Global\s+Rank[^\d#]+#?([\d,]+)",                  lambda x: int(x.replace(",", ""))),
        ("country_rank",     r"Country\s+Rank[^\d#]+#?([\d,]+)",                 lambda x: int(x.replace(",", ""))),
    ]
    for key, pat, conv in patterns:
        m = re.search(pat, text, re.I)
        if m:
            try:
                out[key] = conv(m.group(1))
            except Exception:
                pass

    if not any(out.get(k) for k in ("monthly_visits", "global_rank", "country_rank")):
        out["error"] = "no metrics extracted (CF block or no data for this domain)"
    return out


def fetch_hypestat_public(domain: str) -> dict:
    """Fallback: hypestat.com aggregates many free traffic estimators."""
    Fetcher, _, _ = _get_fetchers()
    url = f"https://hypestat.com/info/{domain}"
    try:
        page = Fetcher.get(url, stealthy_headers=True, timeout=30)
    except Exception as e:
        return {"source": "hypestat", "error": str(e)[:200]}
    if not page:
        return {"source": "hypestat", "error": "no page returned"}

    text = page.get_all_text() or ""
    out  = {"source": "hypestat", "error": ""}

    m = re.search(r"Daily\s+Unique\s+Visitors[^\d.]+([\d.,]+\s*[KMB]?)", text, re.I)
    if m:
        daily = _humanized_to_int(m.group(1))
        if daily:
            out["monthly_visits"] = daily * 30
    m = re.search(r"Daily\s+Pageviews[^\d.]+([\d.,]+\s*[KMB]?)", text, re.I)
    if m and out.get("monthly_visits"):
        daily_pv = _humanized_to_int(m.group(1))
        if daily_pv and out["monthly_visits"]:
            out["pages_per_visit"] = round(daily_pv / (out["monthly_visits"] / 30), 2)

    if not out.get("monthly_visits"):
        out["error"] = "no metrics extracted"
    return out


def fetch_traffic_metrics(domain: str) -> dict:
    """Cascade through sources, return whichever produces data."""
    domain = domain.replace("https://", "").replace("http://", "").rstrip("/").split("/")[0]
    if domain.startswith("www."):
        domain = domain[4:]

    print(f"[traffic] resolving for {domain}")
    primary = fetch_similarweb_public(domain)
    if primary.get("monthly_visits") or primary.get("global_rank"):
        print(f"[traffic]   ✓ similarweb: {primary.get('monthly_visits')} visits/mo")
        return primary

    print(f"[traffic]   similarweb failed ({primary.get('error')}), trying hypestat")
    fallback = fetch_hypestat_public(domain)
    if fallback.get("monthly_visits"):
        print(f"[traffic]   ✓ hypestat: {fallback.get('monthly_visits')} visits/mo (estimated)")
        return fallback

    print(f"[traffic]   ✗ all sources failed")
    return {"source": "unavailable",
            "error": primary.get("error") or fallback.get("error") or "all sources failed"}


def update_site_traffic(site_id: int) -> dict:
    """Fetch traffic for a site and persist a snapshot."""
    db   = get_db()
    site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
    if not site:
        db.close(); return {"ok": False, "error": "site not found"}

    domain  = urlparse(site["url"]).netloc
    metrics = fetch_traffic_metrics(domain)

    has_data = bool(metrics.get("monthly_visits") or metrics.get("global_rank")
                    or metrics.get("country_rank"))

    if has_data:
        db.execute("""
            INSERT INTO site_traffic
                (site_id, monthly_visits, bounce_rate, pages_per_visit,
                 avg_duration_sec, global_rank, country_rank, source, error)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (site_id,
              metrics.get("monthly_visits", 0),
              metrics.get("bounce_rate", 0),
              metrics.get("pages_per_visit", 0),
              metrics.get("avg_duration_sec", 0),
              metrics.get("global_rank", 0),
              metrics.get("country_rank", 0),
              metrics.get("source", ""),
              metrics.get("error", "")))
    else:
        # Failed attempt — log it but don't pollute history with empty rows.
        # Only one "failed marker" row, replacing any previous failed marker.
        db.execute("""DELETE FROM site_traffic
                      WHERE site_id=? AND monthly_visits=0 AND source='unavailable'""",
                   (site_id,))
        db.execute("""
            INSERT INTO site_traffic (site_id, source, error)
            VALUES (?,?,?)
        """, (site_id, metrics.get("source", "unavailable"),
              metrics.get("error", "")[:300]))

    db.execute("UPDATE sites SET last_traffic_check=? WHERE id=?",
               (datetime.now().isoformat(), site_id))
    db.commit(); db.close()
    return {"ok": True, **metrics}


_POSITIVE_WORDS = {
    "victoria", "gana", "ganan", "ganador", "ganadora", "campeón", "campeona", "récord",
    "éxito", "triunfo", "logro", "histórico", "histórica", "acuerdo", "alianza", "renovación",
    "amplía", "crece", "crecimiento", "expansión", "inversión", "beneficio", "beneficios",
    "positivo", "positiva", "aumenta", "supera", "apoya", "apoyo", "reconocimiento",
    "liderazgo", "líder", "lider", "innovación", "innovador", "sostenible", "sostenibilidad",
    "primera", "primero", "nuevo record", "nuevo récord", "millones", "patrocinio", "sponsor",
    "partnership", "colaboración", "debut", "buen", "buena", "excelente", "destacado",
    "remontada", "hat-trick", "doblete", "imbatible", "clasificado", "clasificada",
    "asciende", "ascenso", "sube", "subida", "refuerzo", "fichaje", "renovar",
    "sold out", "sold-out", "lleno", "récord de asistencia",
}

_NEGATIVE_WORDS = {
    "derrota", "pierde", "pierden", "perdió", "eliminado", "eliminada", "lesión", "lesionado",
    "lesionada", "sanción", "suspendido", "suspendida", "escándalo", "fraude", "pérdida",
    "pérdidas", "quiebra", "demanda", "denuncia", "rotura", "crisis", "caída", "cae",
    "baja", "descenso", "desciende", "expulsado", "expulsión", "tarjeta roja",
    "multa", "investigado", "investigada", "corrupción", "acusado", "acusada",
    "retirado", "retirada", "retirase", "fallece", "fallecido", "tragedia",
    "negativo", "negativa", "preocupante", "polémico", "polémica", "controversia",
    "problemas", "dificultades", "fracaso", "decepcionante", "frustrante",
    "cancelado", "cancelada", "aplazado", "aplazada", "boicot", "huelga",
}

def analyze_sentiment_local(title: str, summary: str = "", body: str = "") -> str:
    """Fast keyword-based sentiment: positive / negative / neutral."""
    text = f"{title} {summary} {body[:800]}".lower()
    # Remove punctuation so 'victoria,' still matches
    clean = re.sub(r"[^\w\sáéíóúüñ]", " ", text)
    tokens = set(clean.split())
    pos = sum(1 for w in _POSITIVE_WORDS if w in tokens or w in clean)
    neg = sum(1 for w in _NEGATIVE_WORDS if w in tokens or w in clean)
    if pos > neg:
        return "positive"
    if neg > pos:
        return "negative"
    return "neutral"


def enrich_with_ai_batch(articles: list[dict]) -> dict[int, dict]:
    """Enrich up to 15 articles in a single AI API call.
    articles: list of {id, title, body, site_name}
    Returns: {article_id: {company_property, company_brand, company_agency, summary, sentiment}}
    """
    if not articles:
        return {}
    items = [{"id": a["id"], "title": (a["title"] or "")[:200],
              "body": (a["body"] or "")[:1500], "site": (a["site_name"] or "")}
             for a in articles]
    prompt = (
        "Eres un analista de sports business. Para cada artículo de la lista, extrae las entidades "
        "mencionadas en el cuerpo (NO el menú/navegación). "
        "NUNCA incluyas el campo 'site' como valor en los campos de salida.\n\n"
        "Para cada artículo devuelve:\n"
        "- company_property: propiedad deportiva (liga, club, federación, evento). Vacío si no hay.\n"
        "- company_brand: marca comercial (patrocinador, fabricante, sponsor). Vacío si no hay.\n"
        "- company_agency: agencia/broadcaster/plataforma OTT. Vacío si no hay.\n"
        "- summary: resumen de 1-2 frases en español para correo profesional.\n"
        "- sentiment: 'positive' | 'neutral' | 'negative'\n\n"
        "Devuelve SÓLO un JSON array válido (sin markdown) con un objeto por artículo, "
        "incluyendo el campo 'id' original:\n\n"
        f"{json.dumps(items, ensure_ascii=False)}"
    )
    raw = None
    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if gemini_key:
        try:
            from google import genai
            from google.genai import types
            client = genai.Client(api_key=gemini_key)
            resp = client.models.generate_content(
                model="gemini-2.5-flash-lite", contents=prompt,
                config=types.GenerateContentConfig(response_mime_type="application/json"),
            )
            raw = resp.text
        except Exception as e:
            print(f"[batch-ai] Gemini error: {e}")
    if raw is None:
        anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if anthropic_key:
            try:
                import anthropic
                client = anthropic.Anthropic(api_key=anthropic_key)
                msg = client.messages.create(
                    model="claude-haiku-4-5-20251001", max_tokens=4000,
                    messages=[{"role": "user", "content": prompt}]
                )
                raw = msg.content[0].text
            except Exception as e:
                print(f"[batch-ai] Claude error: {e}")
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list):
            return {int(r["id"]): r for r in parsed if "id" in r}
    except Exception as e:
        print(f"[batch-ai] parse error: {e}\nraw: {raw[:300]}")
    return {}


def enrich_with_ai(title: str, body: str, site_name: str = "") -> dict:
    """Use Gemini 2.5 Flash to extract Propiedad / Marca / Agencia / Resumen.
    Falls back to Anthropic Claude if GEMINI_API_KEY not set but ANTHROPIC_API_KEY is."""
    publisher_note = (
        f"\n- IMPORTANTE: El artículo se publica en «{site_name}». NUNCA incluyas «{site_name}» "
        f"en ningún campo, ni siquiera si aparece varias veces en el contenido (es el menú/header)."
        if site_name else
        "\n- NO incluyas el medio publicador (2Playbook, Palco23, Marca, El Mundo, etc.) en ningún campo."
    )
    prompt = (
        "Eres un analista de sports business. Analiza el artículo y extrae las entidades "
        "MENCIONADAS dentro del CUERPO de la noticia (no del menú/navegación). "
        "Devuelve SÓLO un JSON válido sin markdown.\n\n"
        "Reglas estrictas:" + publisher_note + "\n"
        "- Sólo cita entidades que aparezcan explícitamente como SUJETO de la noticia.\n"
        "- Si hay varias en un campo, sepáralas por coma. Si no hay ninguna, devuelve cadena vacía.\n\n"
        f"Título: {title}\n\nContenido:\n{body[:4000]}\n\n"
        'JSON requerido:\n{\n'
        '  "company_property": "propiedad deportiva (liga, club, federación, evento, competición)",\n'
        '  "company_brand":    "marca comercial (patrocinador, fabricante, anunciante, sponsor)",\n'
        '  "company_agency":   "agencia/intermediario/plataforma OTT/broadcaster (ej: IMG, DAZN, Mediapro, WPP). Cadena vacía si no hay ninguno mencionado.",\n'
        '  "summary":          "resumen ejecutivo de 1-2 frases en español, listo para correo profesional",\n'
        '  "sentiment":        "positive | neutral | negative — sentimiento general de la noticia"\n}'
    )

    # Prefer Gemini (free tier — 2.5-flash-lite gives 1000 req/day vs 20 for flash)
    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if gemini_key:
        try:
            from google import genai
            from google.genai import types
            client = genai.Client(api_key=gemini_key)
            resp = client.models.generate_content(
                model="gemini-2.5-flash-lite",
                contents=prompt,
                config=types.GenerateContentConfig(response_mime_type="application/json"),
            )
            return json.loads(resp.text)
        except Exception as e:
            print(f"Gemini enrich error: {e}")

    # Fallback: Anthropic Claude
    anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if anthropic_key:
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=anthropic_key)
            msg = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=400,
                messages=[{"role": "user", "content": prompt}]
            )
            return json.loads(msg.content[0].text)
        except Exception as e:
            print(f"Claude enrich error: {e}")

    return {}


def scrape_and_store(site_id: int) -> int:
    """Scrape a site and persist results. Always updates last_scrape_status
    so the UI can show success/empty/error per site."""
    db   = get_db()
    site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
    if not site:
        db.close(); return 0

    status = "error"
    error  = ""
    count  = 0

    try:
        site_cookies = _parse_cookies(site["cookies_json"] or "")
        site_cdp    = (site["cdp_url"] or "").strip() or None
        site_tier   = int(site["last_working_tier"] or 1)
        raw_links, worked_tier = fetch_site_links(
            site["url"], site["css_selector"] or "",
            bool(site["use_stealth"]),
            cookies=site_cookies, cdp_url=site_cdp,
            start_tier=site_tier)
        if worked_tier and worked_tier != site_tier:
            db.execute("UPDATE sites SET last_working_tier=? WHERE id=?", (worked_tier, site_id))
            db.commit()
        articles  = extract_articles(raw_links, site)

        if not raw_links:
            status, error = "error", "No se pudo cargar el sitio (anti-bot, red, o página vacía)"
        elif not articles:
            status, error = "empty", f"Se cargaron {len(raw_links)} enlaces pero ninguno coincide con el patrón de artículo"
        else:
            today = date.today().isoformat()
            for a in articles:
                exists = db.execute(
                    "SELECT id FROM articles WHERE site_id=? AND title=? AND scrape_date=?",
                    (site_id, a["title"], today)
                ).fetchone()
                if not exists:
                    db.execute("""
                        INSERT INTO articles
                            (site_id,title,article_url,author,category,
                             company_property,company_brand,company_agency,summary,body,scrape_date)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)
                    """, (site_id, a["title"], a["article_url"], a["author"], a["category"],
                          a["company_property"], a["company_brand"], a["company_agency"],
                          a["summary"], a["body"], today))
                    count += 1
            status = "success"
    except Exception as e:
        status = "error"
        error  = str(e)[:300]
        print(f"[scrape] FAIL site={site_id} ({site['name']}): {error}")

    db.execute("""
        UPDATE sites SET last_scraped=?, last_scrape_status=?,
                         last_scrape_error=?, last_scrape_count=?
        WHERE id=?
    """, (datetime.now().isoformat(), status, error, count, site_id))
    db.commit(); db.close()
    return count


# ── routes ────────────────────────────────────────────────────────────────────

_BRAND_LIKE_COLS = ("a.title", "a.body", "a.summary",
                    "a.company_property", "a.company_brand", "a.company_agency")

def _search_where(query: str):
    """Build SQL WHERE clause + params.
    Splits query by spaces; all tokens must appear somewhere in the article (AND logic).
    E.g. 'FC Barcelona' → articles containing both 'FC' AND 'Barcelona'.
    """
    if not query:
        return "1=1", []
    tokens = [t for t in query.split() if t]
    if not tokens:
        return "1=1", []
    cols = [f"(' '||COALESCE({c},'')||' ')" for c in _BRAND_LIKE_COLS]
    # Each token must match in at least one column
    token_clauses, params = [], []
    for token in tokens:
        like = f"%{token}%"
        token_clauses.append("(" + " OR ".join(f"{c} LIKE ?" for c in cols) + ")")
        params.extend([like] * len(_BRAND_LIKE_COLS))
    return "(" + " AND ".join(token_clauses) + ")", params


def _split_keywords(text: str) -> list[str]:
    """Split a multi-keyword string by comma / semicolon / newline."""
    if not text: return []
    return [k.strip() for k in re.split(r"[,;\n]", text) if k.strip()]


def _search_where_multi(keywords: str):
    """Build SQL WHERE matching ANY of the comma/newline-separated keywords.
    Returns ('1=1', []) when the string is empty (no scope filter)."""
    terms = _split_keywords(keywords)
    if not terms:
        return "1=1", []
    clauses, params = [], []
    for k in terms:
        like = f"% {k} %"
        cols = [f"(' '||COALESCE({c},'')||' ')" for c in _BRAND_LIKE_COLS]
        clauses.append("(" + " OR ".join(f"{c} LIKE ?" for c in cols) + ")")
        params.extend([like] * len(_BRAND_LIKE_COLS))
    return "(" + " OR ".join(clauses) + ")", params


def compute_search_kpis(db, query: str, days: int = 14) -> dict:
    """Return KPIs (total, today, daily series, per-site breakdown, OTS, VPE)
    for a search query."""
    where, params = _search_where(query)
    today = date.today().isoformat()

    total         = db.execute(f"SELECT COUNT(*) FROM articles a WHERE {where}", params).fetchone()[0]
    today_cnt     = db.execute(f"SELECT COUNT(*) FROM articles a WHERE {where} AND a.scrape_date=?",
                               params + [today]).fetchone()[0]
    ots_sum       = db.execute(f"SELECT COALESCE(SUM(a.ots),0) FROM articles a WHERE {where}", params).fetchone()[0]
    vpe_sum       = db.execute(f"SELECT COALESCE(SUM(a.vpe),0) FROM articles a WHERE {where}", params).fetchone()[0]
    outlets_count = db.execute(f"SELECT COUNT(DISTINCT a.site_id) FROM articles a WHERE {where}", params).fetchone()[0]
    countries_count = db.execute(f"""
        SELECT COUNT(DISTINCT COALESCE(NULLIF(s.country,''),'WW'))
        FROM articles a JOIN sites s ON s.id=a.site_id WHERE {where}
    """, params).fetchone()[0]

    # Daily series for last N days (fill zeros for missing dates)
    series_rows = db.execute(f"""
        SELECT a.scrape_date AS d, COUNT(*) AS c
        FROM articles a WHERE {where}
          AND a.scrape_date >= date('now', '-{days} days')
        GROUP BY a.scrape_date ORDER BY a.scrape_date
    """, params).fetchall()
    series_map = {r["d"]: r["c"] for r in series_rows}
    from datetime import timedelta
    today_d = date.today()
    series = []
    for i in range(days - 1, -1, -1):
        d = (today_d - timedelta(days=i)).isoformat()
        series.append({"date": d, "count": series_map.get(d, 0)})

    # Top media (top 15 for comparison mode)
    top_sites = db.execute(f"""
        SELECT s.name, COUNT(*) AS c FROM articles a
        JOIN sites s ON s.id=a.site_id
        WHERE {where} GROUP BY s.id, s.name ORDER BY c DESC LIMIT 15
    """, params).fetchall()

    # Country distribution
    country_rows = db.execute(f"""
        SELECT COALESCE(NULLIF(s.country,''),'WW') AS code,
               COUNT(*) AS c, COUNT(DISTINCT s.id) AS sites
        FROM articles a JOIN sites s ON s.id=a.site_id
        WHERE {where}
        GROUP BY code ORDER BY c DESC
    """, params).fetchall()
    country_dist = []
    for r in country_rows:
        name, flag = COUNTRY_BY_CODE.get(r["code"], (r["code"], "🌐"))
        country_dist.append({
            "code": r["code"], "name": name, "flag": flag,
            "count": r["c"], "sites": r["sites"],
        })

    return {
        "total":           total,
        "today":           today_cnt,
        "ots":             ots_sum,
        "vpe":             vpe_sum,
        "outlets_count":   outlets_count,
        "countries_count": countries_count,
        "series":          series,
        "top_sites":       [{"name": r["name"], "count": r["c"]} for r in top_sites],
        "country_dist":    country_dist,
    }


VPE_RATE_PER_OTS = 0.05  # €/OTS base rate

# ── OTS/VPE factors ───────────────────────────────────────────────────────────

def _reach_rate(monthly_visits: int) -> float:
    """Fraction of daily visitors estimated to read a given article.
    Higher traffic sites spread readership over many articles → lower per-article rate."""
    if monthly_visits >= 5_000_000: return 0.010   # 1%  grandes generales
    if monthly_visits >= 1_000_000: return 0.020   # 2%  medios nacionales
    if monthly_visits >= 100_000:   return 0.050   # 5%  medios regionales
    return 0.100                                    # 10% medios pequeños/especializados

# Multiplicador por categoría del medio (coincidencia parcial, case-insensitive)
_CATEGORY_FACTORS: list[tuple[str, float]] = [
    ("televi",   2.5),
    (" tv",      2.5),
    ("radio",    2.0),
    ("agencia",  1.8),
    ("nacional", 1.5),
    ("regional", 1.3),
    ("sport",    1.2),
    ("digital",  1.0),
    ("blog",     0.6),
]

def _category_factor(category: str) -> float:
    cat = f" {(category or '').lower()} "
    for key, f in _CATEGORY_FACTORS:
        if key in cat:
            return f
    return 1.0

# Multiplicador por posición de la mención en el artículo
_POSITION_FACTORS = {"title": 2.0, "summary": 1.5, "body": 1.0}

def _mention_position(title: str, summary: str, query: str) -> str:
    """Determine where the brand query appears first: title > summary > body."""
    q = query.strip().lower()
    pad = lambda t: f" {(t or '').lower()} "
    if f" {q} " in pad(title):   return "title"
    if f" {q} " in pad(summary): return "summary"
    return "body"

def _article_ots(monthly_visits: int, category: str,
                 query: str = "", title: str = "", summary: str = "") -> tuple[float, str]:
    """OTS for one article with all factors: reach × category × position.
    Returns (ots_value, position_label)."""
    daily   = (monthly_visits or 0) / 30
    reach   = _reach_rate(monthly_visits or 0)
    cat_f   = _category_factor(category)
    pos     = _mention_position(title, summary, query) if query else "body"
    pos_f   = _POSITION_FACTORS[pos]
    return daily * reach * cat_f * pos_f, pos


def compute_universe_kpis(db, keywords: str, days: int = 30) -> dict:
    """KPIs for a client's news universe (articles matching ANY of the
    given keywords). Used to size the scope before drilling into brands."""
    from datetime import timedelta
    if not _split_keywords(keywords):
        return {"total": 0, "today": 0, "ots": 0, "vpe": 0,
                "configured": False}
    where, params = _search_where_multi(keywords)
    today_d = date.today()
    since   = (today_d - timedelta(days=days-1)).isoformat()
    today_s = today_d.isoformat()

    rows = db.execute(f"""
        SELECT a.scrape_date,
               COALESCE(NULLIF(s.category,''),'(sin categoría)') AS category,
               COALESCE(
                 (SELECT monthly_visits FROM site_traffic
                  WHERE site_id=s.id AND monthly_visits>0
                  ORDER BY measured_at DESC LIMIT 1),
                 s.monthly_visits_manual, 0
               ) AS site_visits
        FROM articles a JOIN sites s ON s.id=a.site_id
        WHERE {where} AND a.scrape_date >= ?
    """, params + [since]).fetchall()
    total = len(rows)
    today_cnt = sum(1 for r in rows if r["scrape_date"] == today_s)
    ots = sum(_article_ots(r["site_visits"], r["category"])[0] for r in rows)
    return {"total": total, "today": today_cnt, "ots": round(ots),
            "vpe": round(ots * VPE_RATE_PER_OTS, 2), "configured": True}


def compute_search_universe(db, search_kw: str, client_kw: str, days: int = 30) -> dict:
    """KPIs for articles matching búsqueda keywords AND client keywords.
    Used to size the búsqueda scope (inner universe of a search folder)."""
    from datetime import timedelta
    search_terms = _split_keywords(search_kw)
    client_terms = _split_keywords(client_kw)
    if not search_terms:
        return {"total": 0, "today": 0, "ots": 0, "vpe": 0, "configured": False}
    conditions, params = [], []
    ws, ps = _search_where_multi(search_kw)
    conditions.append(f"({ws})"); params += ps
    if client_terms:
        wc, pc = _search_where_multi(client_kw)
        conditions.append(f"({wc})"); params += pc
    where = " AND ".join(conditions)
    today_d = date.today()
    since   = (today_d - timedelta(days=days-1)).isoformat()
    today_s = today_d.isoformat()
    rows = db.execute(f"""
        SELECT a.scrape_date,
               COALESCE(NULLIF(s.category,''),'(sin categoría)') AS category,
               COALESCE(
                 (SELECT monthly_visits FROM site_traffic
                  WHERE site_id=s.id AND monthly_visits>0
                  ORDER BY measured_at DESC LIMIT 1),
                 s.monthly_visits_manual, 0
               ) AS site_visits
        FROM articles a JOIN sites s ON s.id=a.site_id
        WHERE {where} AND a.scrape_date >= ?
    """, params + [since]).fetchall()
    total = len(rows)
    today_cnt = sum(1 for r in rows if r["scrape_date"] == today_s)
    ots = sum(_article_ots(r["site_visits"], r["category"])[0] for r in rows)
    return {"total": total, "today": today_cnt, "ots": round(ots),
            "vpe": round(ots * VPE_RATE_PER_OTS, 2), "configured": True}


def get_search_articles(db, search_kw: str, client_kw: str, days: int = 30) -> list:
    """Full article list for a búsqueda scope (search keywords + client universe)."""
    from datetime import timedelta
    if not _split_keywords(search_kw):
        return []
    conditions, params = [], []
    ws, ps = _search_where_multi(search_kw)
    conditions.append(f"({ws})"); params += ps
    if _split_keywords(client_kw):
        wc, pc = _search_where_multi(client_kw)
        conditions.append(f"({wc})"); params += pc
    where = " AND ".join(conditions)
    since = (date.today() - timedelta(days=days-1)).isoformat()
    rows = db.execute(f"""
        SELECT a.id, a.title, a.article_url, a.scrape_date,
               s.name AS site_name,
               COALESCE(NULLIF(s.country,''),'WW') AS country,
               COALESCE(NULLIF(s.category,''),'(sin categoría)') AS category,
               COALESCE(
                 (SELECT monthly_visits FROM site_traffic
                  WHERE site_id=s.id AND monthly_visits>0
                  ORDER BY measured_at DESC LIMIT 1),
                 s.monthly_visits_manual, 0
               ) AS site_visits
        FROM articles a JOIN sites s ON s.id=a.site_id
        WHERE {where} AND a.scrape_date >= ?
        ORDER BY a.scrape_date DESC
    """, params + [since]).fetchall()
    result = []
    for r in rows:
        ots_a, _ = _article_ots(r["site_visits"], r["category"])
        _, flag = COUNTRY_BY_CODE.get(r["country"] or "WW", ("", "🌐"))
        result.append({"id": r["id"], "title": r["title"],
            "article_url": r["article_url"], "scrape_date": r["scrape_date"],
            "site_name": r["site_name"], "country": r["country"], "flag": flag,
            "category": r["category"], "ots": round(ots_a),
            "vpe": round(ots_a * VPE_RATE_PER_OTS, 2)})
    return result


def compute_brand_impact(db, query: str, days: int = 30,
                         scope_query: str = "", search_scope: str = "") -> dict:
    """Full impact breakdown for a brand query: totals, distribution by
    site.country / site.category, top sites, top articles, daily series.
    OTS = monthly_visits/30 per article; VPE = OTS × VPE_RATE_PER_OTS.
    `scope_query` restricts to client universe; `search_scope` further
    restricts to the búsqueda's own keyword universe."""
    from datetime import timedelta
    where_b, p_b = _search_where(query)
    conditions = [f"({where_b})"]
    params = list(p_b)
    if scope_query:
        wc, pc = _search_where_multi(scope_query)
        conditions.append(f"({wc})"); params += pc
    if search_scope:
        ws, ps = _search_where_multi(search_scope)
        conditions.append(f"({ws})"); params += ps
    where = " AND ".join(conditions)
    today_d = date.today()
    since   = (today_d - timedelta(days=days-1)).isoformat()
    today_s = today_d.isoformat()

    rows = db.execute(f"""
        SELECT a.id, a.title, a.summary, a.article_url, a.scrape_date,
               s.id AS site_id, s.name AS site_name,
               COALESCE(NULLIF(s.country,''),'WW') AS country,
               COALESCE(NULLIF(s.category,''),'(sin categoría)') AS category,
               COALESCE(
                 (SELECT monthly_visits FROM site_traffic
                  WHERE site_id=s.id AND monthly_visits>0
                  ORDER BY measured_at DESC LIMIT 1),
                 s.monthly_visits_manual, 0
               ) AS site_visits
        FROM articles a JOIN sites s ON s.id=a.site_id
        WHERE {where} AND a.scrape_date >= ?
    """, params + [since]).fetchall()

    total = len(rows); today_cnt = 0; ots_sum = 0.0
    by_country_map: dict = {}
    by_category_map: dict = {}
    by_site: dict = {}
    series_map: dict = {}

    for r in rows:
        ots_a, pos = _article_ots(r["site_visits"], r["category"],
                                   query, r["title"], r["summary"])
        vpe_a = ots_a * VPE_RATE_PER_OTS
        ots_sum += ots_a
        if r["scrape_date"] == today_s: today_cnt += 1
        series_map[r["scrape_date"]] = series_map.get(r["scrape_date"], 0) + 1

        c = by_country_map.setdefault(r["country"], {"count":0,"ots":0.0,"vpe":0.0,"sites":set()})
        c["count"] += 1; c["ots"] += ots_a; c["vpe"] += vpe_a; c["sites"].add(r["site_id"])

        cat = by_category_map.setdefault(r["category"], {"count":0,"ots":0.0,"vpe":0.0,"sites":set()})
        cat["count"] += 1; cat["ots"] += ots_a; cat["vpe"] += vpe_a; cat["sites"].add(r["site_id"])

        st = by_site.setdefault(r["site_id"], {"name":r["site_name"],"country":r["country"],
            "category":r["category"],"count":0,"ots":0.0,"vpe":0.0})
        st["count"] += 1; st["ots"] += ots_a; st["vpe"] += vpe_a

    vpe_sum = ots_sum * VPE_RATE_PER_OTS

    by_country = []
    for code, v in by_country_map.items():
        name, flag = COUNTRY_BY_CODE.get(code, (code, "🌐"))
        by_country.append({"code": code, "name": name, "flag": flag,
            "count": v["count"], "ots": v["ots"], "vpe": v["vpe"],
            "sites": len(v["sites"])})
    by_country.sort(key=lambda x: x["count"], reverse=True)

    by_category = [{"name": k, "count": v["count"], "ots": v["ots"],
        "vpe": v["vpe"], "sites": len(v["sites"])}
        for k, v in by_category_map.items()]
    by_category.sort(key=lambda x: x["count"], reverse=True)

    top_sites = []
    for sid, s in by_site.items():
        _, flag = COUNTRY_BY_CODE.get(s["country"] or "WW", ("", "🌐"))
        top_sites.append({"id": sid, "name": s["name"], "country": s["country"],
            "flag": flag, "category": s["category"],
            "count": s["count"], "ots": s["ots"], "vpe": s["vpe"]})
    top_sites.sort(key=lambda x: x["ots"], reverse=True)
    top_sites = top_sites[:10]

    top_articles = []
    for r in rows:
        ots_a, pos = _article_ots(r["site_visits"], r["category"],
                                   query, r["title"], r["summary"])
        top_articles.append({"id": r["id"], "title": r["title"],
            "article_url": r["article_url"], "scrape_date": r["scrape_date"],
            "site_name": r["site_name"], "country": r["country"],
            "position": pos,
            "ots": round(ots_a), "vpe": round(ots_a * VPE_RATE_PER_OTS, 2)})
    top_articles.sort(key=lambda x: (x["ots"], x["scrape_date"]), reverse=True)
    top_articles = top_articles[:8]

    from datetime import timedelta as _td
    series = [{"date": (today_d - _td(days=i)).isoformat(),
               "count": series_map.get((today_d - _td(days=i)).isoformat(), 0)}
              for i in range(days-1, -1, -1)]

    return {
        "total": total, "today": today_cnt,
        "ots": round(ots_sum), "vpe": round(vpe_sum, 2),
        "by_country": [{**c, "ots": round(c["ots"]), "vpe": round(c["vpe"], 2)} for c in by_country],
        "by_category": [{**c, "ots": round(c["ots"]), "vpe": round(c["vpe"], 2)} for c in by_category],
        "top_sites": top_sites, "top_articles": top_articles,
        "series": series, "days": days,
    }


@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request):
    db    = get_db()
    today = date.today().isoformat()
    stats = {
        "sites":  db.execute("SELECT COUNT(*) FROM sites WHERE active=1").fetchone()[0],
        "today":  db.execute("SELECT COUNT(*) FROM articles WHERE scrape_date=?", (today,)).fetchone()[0],
        "total":  db.execute("SELECT COUNT(*) FROM articles").fetchone()[0],
        "emails": db.execute("SELECT COUNT(*) FROM email_logs").fetchone()[0],
    }
    sites = db.execute("""
        SELECT s.*,
               COALESCE(td.c, 0) AS today_count,
               COALESCE(tt.c, 0) AS total_count
        FROM sites s
        LEFT JOIN (SELECT site_id, COUNT(*) AS c FROM articles WHERE scrape_date=? GROUP BY site_id) td ON td.site_id=s.id
        LEFT JOIN (SELECT site_id, COUNT(*) AS c FROM articles GROUP BY site_id) tt ON tt.site_id=s.id
        WHERE s.active=1 ORDER BY s.last_scraped DESC LIMIT 30
    """, (today,)).fetchall()
    recent = db.execute("""
        SELECT a.*, s.name AS site_name FROM articles a
        JOIN sites s ON s.id=a.site_id
        WHERE a.scrape_date=? ORDER BY a.scraped_at DESC LIMIT 8
    """, (today,)).fetchall()

    # Daily totals across all articles (last 30 days) for global chart
    daily_rows = db.execute("""
        SELECT scrape_date AS d, COUNT(*) AS c FROM articles
        WHERE scrape_date >= date('now','-30 days')
        GROUP BY scrape_date ORDER BY scrape_date
    """).fetchall()
    from datetime import timedelta
    today_d = date.today()
    daily_map = {r["d"]: r["c"] for r in daily_rows}
    daily_series = []
    for i in range(29, -1, -1):
        d = (today_d - timedelta(days=i)).isoformat()
        daily_series.append({"date": d, "count": daily_map.get(d, 0)})

    # Country distribution: articles per site.country (last 30 days)
    country_rows = db.execute("""
        SELECT COALESCE(NULLIF(s.country,''),'WW') AS code,
               COUNT(*) AS c, COUNT(DISTINCT s.id) AS sites
        FROM articles a
        JOIN sites s ON s.id=a.site_id
        WHERE a.scrape_date >= date('now','-30 days')
        GROUP BY code ORDER BY c DESC
    """).fetchall()
    country_dist = []
    for r in country_rows:
        name, flag = COUNTRY_BY_CODE.get(r["code"], (r["code"], "🌐"))
        country_dist.append({
            "code": r["code"], "name": name, "flag": flag,
            "count": r["c"], "sites": r["sites"],
        })

    db.close()
    # saved_searches and dashboard_clients are loaded async via /dashboard/widgets
    return templates.TemplateResponse(request=request, name="dashboard.html", context=dict(
        request=request, active_page="dashboard",
        stats=stats, sites=sites, recent=recent, today=today,
        daily_series=daily_series, country_dist=country_dist))


@app.get("/dashboard/widgets", response_class=JSONResponse)
async def dashboard_widgets():
    """Heavy widgets loaded asynchronously after the dashboard renders."""
    db = get_db()
    saved = db.execute("SELECT * FROM saved_searches ORDER BY created_at DESC").fetchall()
    saved_kpis = []
    for s in saved:
        kpis = compute_search_kpis(db, s["query"])
        saved_kpis.append({"id": s["id"], "name": s["name"], "query": s["query"], "kpis": kpis})

    client_rows = db.execute("""
        SELECT c.id, c.name, c.notes,
               (SELECT COUNT(*) FROM client_searches WHERE client_id=c.id) AS search_count,
               (SELECT COUNT(*) FROM client_brands  WHERE client_id=c.id) AS brand_count
        FROM clients c ORDER BY c.created_at DESC
    """).fetchall()
    dashboard_clients = []
    for c in client_rows:
        brands = db.execute("SELECT query FROM client_brands WHERE client_id=?", (c["id"],)).fetchall()
        impacts = [compute_brand_impact(db, b["query"], days=30) for b in brands]
        agg = {"total": sum(k["total"] for k in impacts),
               "ots":   sum(k["ots"]   for k in impacts),
               "vpe":   sum(k["vpe"]   for k in impacts)}
        dashboard_clients.append({"id": c["id"], "name": c["name"],
                                   "search_count": c["search_count"],
                                   "brand_count":  c["brand_count"],
                                   "impact": agg})
    db.close()
    return {"saved_searches": saved_kpis, "clients": dashboard_clients}


@app.post("/searches")
async def create_saved_search(name: str = Form(...), query: str = Form(...)):
    db = get_db()
    db.execute("INSERT INTO saved_searches (name, query) VALUES (?, ?)",
               (name.strip(), query.strip()))
    db.commit()
    sid = db.execute("SELECT id FROM saved_searches ORDER BY id DESC LIMIT 1").fetchone()[0]
    db.close()
    return JSONResponse({"ok": True, "id": sid})


@app.post("/searches/{search_id}/delete")
async def delete_saved_search(search_id: int):
    db = get_db()
    db.execute("DELETE FROM saved_searches WHERE id=?", (search_id,))
    db.commit(); db.close()
    return JSONResponse({"ok": True})


# ── clients · searches · brands ───────────────────────────────────────────────

def _aggregate_impacts(impacts: list[dict]) -> dict:
    """Sum a list of compute_brand_impact() results into one combined view."""
    agg = {"total": 0, "today": 0, "ots": 0, "vpe": 0,
           "by_country": {}, "by_category": {}, "series_map": {}}
    for k in impacts:
        agg["total"] += k["total"]; agg["today"] += k["today"]
        agg["ots"]   += k["ots"];   agg["vpe"]   += k["vpe"]
        for c in k["by_country"]:
            row = agg["by_country"].setdefault(c["code"], {**c, "count":0,"ots":0,"vpe":0,"sites":0})
            row["count"] += c["count"]; row["ots"] += c["ots"]
            row["vpe"]   += c["vpe"];   row["sites"] += c["sites"]
        for cat in k["by_category"]:
            row = agg["by_category"].setdefault(cat["name"], {**cat,"count":0,"ots":0,"vpe":0,"sites":0})
            row["count"] += cat["count"]; row["ots"] += cat["ots"]
            row["vpe"]   += cat["vpe"];   row["sites"] += cat["sites"]
        for p in k["series"]:
            agg["series_map"][p["date"]] = agg["series_map"].get(p["date"], 0) + p["count"]
    by_country  = sorted(agg["by_country"].values(),  key=lambda r: r["count"], reverse=True)
    by_category = sorted(agg["by_category"].values(), key=lambda r: r["count"], reverse=True)
    series = [{"date": d, "count": agg["series_map"][d]} for d in sorted(agg["series_map"].keys())]
    return {"total": agg["total"], "today": agg["today"], "ots": agg["ots"],
            "vpe": agg["vpe"], "by_country": by_country,
            "by_category": by_category, "series": series}


@app.get("/clients", response_class=HTMLResponse)
async def clients_list(request: Request):
    db = get_db()
    clients = db.execute("""
        SELECT c.*,
               COALESCE(cs.n, 0) AS search_count,
               COALESCE(cb.n, 0) AS brand_count
        FROM clients c
        LEFT JOIN (SELECT client_id, COUNT(*) AS n FROM client_searches GROUP BY client_id) cs ON cs.client_id=c.id
        LEFT JOIN (SELECT client_id, COUNT(*) AS n FROM client_brands   GROUP BY client_id) cb ON cb.client_id=c.id
        ORDER BY c.created_at DESC
    """).fetchall()
    db.close()
    # KPI totals and impact are loaded async via /clients/widgets
    return templates.TemplateResponse(request=request, name="clients.html",
        context=dict(request=request, active_page="clients", clients=clients))


@app.get("/clients/widgets", response_class=JSONResponse)
async def clients_widgets():
    """Heavy per-client KPIs loaded asynchronously."""
    db = get_db()
    rows = db.execute("""
        SELECT c.id, c.keywords,
               COALESCE(cs.n, 0) AS search_count,
               COALESCE(cb.n, 0) AS brand_count
        FROM clients c
        LEFT JOIN (SELECT client_id, COUNT(*) AS n FROM client_searches GROUP BY client_id) cs ON cs.client_id=c.id
        LEFT JOIN (SELECT client_id, COUNT(*) AS n FROM client_brands   GROUP BY client_id) cb ON cb.client_id=c.id
        ORDER BY c.created_at DESC
    """).fetchall()
    result = []
    total_mentions = total_ots = total_vpe = total_universe = 0
    for c in rows:
        scope = c["keywords"] or ""
        universe = compute_universe_kpis(db, scope, days=30)
        brands = db.execute("SELECT query FROM client_brands WHERE client_id=?", (c["id"],)).fetchall()
        impacts = [compute_brand_impact(db, b["query"], days=30, scope_query=scope) for b in brands]
        agg = {"total": sum(k["total"] for k in impacts),
               "ots":   sum(k["ots"]   for k in impacts),
               "vpe":   sum(k["vpe"]   for k in impacts)}
        total_mentions += agg["total"]; total_ots += agg["ots"]; total_vpe += agg["vpe"]
        total_universe += universe["total"]
        result.append({"id": c["id"], "impact": agg, "universe": universe})
    db.close()
    totals = {"mentions": total_mentions, "ots": total_ots, "vpe": total_vpe,
              "universe": total_universe}
    return {"clients": result, "totals": totals}


@app.post("/clients")
async def create_client(name: str = Form(...), notes: str = Form(""),
                        keywords: str = Form("")):
    db = get_db()
    try:
        db.execute("INSERT INTO clients (name, notes, keywords) VALUES (?,?,?)",
                   (name.strip(), notes.strip(), keywords.strip()))
        db.commit()
    except Exception as _ie:
        if "unique" not in str(_ie).lower() and "duplicate" not in str(_ie).lower():
            raise
        db.close()
        return JSONResponse({"ok": False, "error": "Ya existe un cliente con ese nombre"}, status_code=400)
    db.close()
    return RedirectResponse("/clients", status_code=303)


@app.post("/clients/{cid}/edit")
async def edit_client(cid: int, name: str = Form(...), notes: str = Form(""),
                      keywords: str = Form("")):
    db = get_db()
    db.execute("UPDATE clients SET name=?, notes=?, keywords=? WHERE id=?",
               (name.strip(), notes.strip(), keywords.strip(), cid))
    db.commit(); db.close()
    return RedirectResponse(f"/clients/{cid}", status_code=303)


@app.post("/clients/{cid}/delete")
async def delete_client(cid: int):
    db = get_db()
    db.execute("DELETE FROM client_brands   WHERE client_id=?", (cid,))
    db.execute("DELETE FROM client_searches WHERE client_id=?", (cid,))
    db.execute("DELETE FROM clients         WHERE id=?",        (cid,))
    db.commit(); db.close()
    return RedirectResponse("/clients", status_code=303)


@app.get("/clients/{cid}", response_class=HTMLResponse)
async def client_detail(request: Request, cid: int, days: int = 30):
    """Show all búsquedas (folders) belonging to the client with KPIs per folder.
    All brand matching is scoped to the client's keywords universe."""
    db = get_db()
    client = db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
    if not client:
        db.close()
        return HTMLResponse("Cliente no encontrado", status_code=404)

    scope = client["keywords"] or ""
    universe = compute_universe_kpis(db, scope, days=days)

    searches_rows = db.execute(
        "SELECT * FROM client_searches WHERE client_id=? ORDER BY created_at DESC",
        (cid,)).fetchall()

    searches = []
    all_impacts: list[dict] = []
    for s in searches_rows:
        s_kw = s["keywords"] or "" if "keywords" in s.keys() else ""
        brands = db.execute(
            "SELECT * FROM client_brands WHERE search_id=? ORDER BY created_at",
            (s["id"],)).fetchall()
        impacts = [compute_brand_impact(db, b["query"], days=days,
                                        scope_query=scope, search_scope=s_kw)
                   for b in brands]
        all_impacts.extend(impacts)
        agg = _aggregate_impacts(impacts)
        searches.append({**dict(s),
            "brand_count": len(brands),
            "brands": [{**dict(b), "kpis": k} for b, k in zip(brands, impacts)],
            "agg": agg})

    client_agg = _aggregate_impacts(all_impacts)

    db.close()
    return templates.TemplateResponse(request=request, name="client_detail.html",
        context=dict(request=request, active_page="clients",
                     client=dict(client), searches=searches,
                     universe=universe, agg=client_agg, days=days))


@app.post("/clients/{cid}/searches")
async def create_search(cid: int, name: str = Form(...),
                        keywords: str = Form(""), notes: str = Form("")):
    db = get_db()
    db.execute("INSERT INTO client_searches (client_id, name, keywords, notes) VALUES (?,?,?,?)",
               (cid, name.strip(), keywords.strip(), notes.strip()))
    db.commit(); db.close()
    return RedirectResponse(f"/clients/{cid}", status_code=303)


@app.post("/clients/{cid}/searches/{sid}/edit")
async def edit_search(cid: int, sid: int, name: str = Form(...),
                      keywords: str = Form(""), notes: str = Form("")):
    db = get_db()
    db.execute("UPDATE client_searches SET name=?, keywords=?, notes=? WHERE id=? AND client_id=?",
               (name.strip(), keywords.strip(), notes.strip(), sid, cid))
    db.commit(); db.close()
    return RedirectResponse(f"/clients/{cid}/searches/{sid}", status_code=303)


@app.post("/clients/{cid}/searches/{sid}/delete")
async def delete_search(cid: int, sid: int):
    db = get_db()
    db.execute("DELETE FROM client_brands   WHERE search_id=?", (sid,))
    db.execute("DELETE FROM client_searches WHERE id=? AND client_id=?", (sid, cid))
    db.commit(); db.close()
    return RedirectResponse(f"/clients/{cid}", status_code=303)


@app.get("/clients/{cid}/searches/{sid}", response_class=HTMLResponse)
async def search_detail(request: Request, cid: int, sid: int, days: int = 30):
    db = get_db()
    client = db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
    search = db.execute("SELECT * FROM client_searches WHERE id=? AND client_id=?",
                        (sid, cid)).fetchone()
    if not client or not search:
        db.close()
        return HTMLResponse("No encontrado", status_code=404)
    client_kw = client["keywords"] or ""
    search_kw = search["keywords"] or "" if "keywords" in search.keys() else ""
    # Client-level universe (for banner reference)
    client_universe = compute_universe_kpis(db, client_kw, days=days)
    # Búsqueda-level universe: articles matching búsqueda keywords within client universe
    search_universe = compute_search_universe(db, search_kw, client_kw, days=days)
    # All articles for this búsqueda's keywords (shown at the bottom)
    search_articles = get_search_articles(db, search_kw, client_kw, days=days)
    # Brands scoped to both client universe AND búsqueda keywords
    brands = db.execute(
        "SELECT * FROM client_brands WHERE search_id=? ORDER BY created_at DESC",
        (sid,)).fetchall()
    impacts = [compute_brand_impact(db, b["query"], days=days,
                                    scope_query=client_kw, search_scope=search_kw)
               for b in brands]
    brand_data = [{**dict(b), "kpis": k} for b, k in zip(brands, impacts)]
    agg = _aggregate_impacts(impacts)
    db.close()
    return templates.TemplateResponse(request=request, name="client_search.html",
        context=dict(request=request, active_page="clients",
                     client=dict(client), search=dict(search),
                     brands=brand_data, agg=agg,
                     client_universe=client_universe, search_universe=search_universe,
                     search_articles=search_articles, days=days))


@app.get("/clients/{cid}/searches/{sid}/brands/{bid}", response_class=HTMLResponse)
async def brand_detail(request: Request, cid: int, sid: int, bid: int, days: int = 30):
    db = get_db()
    client = db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
    search = db.execute("SELECT * FROM client_searches WHERE id=? AND client_id=?",
                        (sid, cid)).fetchone()
    brand  = db.execute("SELECT * FROM client_brands WHERE id=? AND search_id=?",
                        (bid, sid)).fetchone()
    if not client or not search or not brand:
        db.close()
        return HTMLResponse("No encontrado", status_code=404)
    client_kw = client["keywords"] or ""
    search_kw = search["keywords"] or "" if "keywords" in search.keys() else ""
    universe = compute_universe_kpis(db, client_kw, days=days)
    impact  = compute_brand_impact(db, brand["query"], days=days,
                                   scope_query=client_kw, search_scope=search_kw)
    # Full article list for this brand scoped to client + búsqueda keywords
    articles = get_search_articles.__wrapped__ if hasattr(get_search_articles, '__wrapped__') else None
    # Build full article list with brand filter + scopes
    from datetime import timedelta
    where_b, p_b = _search_where(brand["query"])
    conditions = [f"({where_b})"]
    params = list(p_b)
    if client_kw:
        wc, pc = _search_where_multi(client_kw)
        conditions.append(f"({wc})"); params += pc
    if search_kw:
        ws, ps = _search_where_multi(search_kw)
        conditions.append(f"({ws})"); params += ps
    where = " AND ".join(conditions)
    since = (date.today() - timedelta(days=days-1)).isoformat()
    all_rows = db.execute(f"""
        SELECT a.id, a.title, a.summary, a.article_url, a.scrape_date,
               s.name AS site_name, COALESCE(NULLIF(s.country,''),'WW') AS country,
               COALESCE(NULLIF(s.category,''),'(sin categoría)') AS category,
               COALESCE(
                 (SELECT monthly_visits FROM site_traffic
                  WHERE site_id=s.id AND monthly_visits>0
                  ORDER BY measured_at DESC LIMIT 1),
                 s.monthly_visits_manual, 0
               ) AS site_visits
        FROM articles a JOIN sites s ON s.id=a.site_id
        WHERE {where} AND a.scrape_date >= ?
        ORDER BY a.scrape_date DESC
    """, params + [since]).fetchall()
    articles = []
    for r in all_rows:
        ots_a, pos = _article_ots(r["site_visits"], r["category"],
                                   brand["query"], r["title"], r["summary"])
        _, flag = COUNTRY_BY_CODE.get(r["country"] or "WW", ("", "🌐"))
        articles.append({"id": r["id"], "title": r["title"],
            "article_url": r["article_url"], "scrape_date": r["scrape_date"],
            "site_name": r["site_name"], "country": r["country"],
            "flag": flag, "category": r["category"], "position": pos,
            "ots": round(ots_a), "vpe": round(ots_a * VPE_RATE_PER_OTS, 2)})
    db.close()
    return templates.TemplateResponse(request=request, name="client_brand.html",
        context=dict(request=request, active_page="clients",
                     client=dict(client), search=dict(search), brand=dict(brand),
                     impact=impact, articles=articles, universe=universe, days=days))


@app.post("/clients/{cid}/searches/{sid}/brands")
async def create_brand(cid: int, sid: int, name: str = Form(...),
                       query: str = Form(...), notes: str = Form("")):
    db = get_db()
    db.execute("INSERT INTO client_brands (client_id, search_id, name, query, notes) "
               "VALUES (?,?,?,?,?)",
               (cid, sid, name.strip(), query.strip(), notes.strip()))
    db.commit(); db.close()
    return RedirectResponse(f"/clients/{cid}/searches/{sid}", status_code=303)


@app.post("/clients/{cid}/searches/{sid}/brands/{bid}/edit")
async def edit_brand(cid: int, sid: int, bid: int, name: str = Form(...),
                     query: str = Form(...), notes: str = Form("")):
    db = get_db()
    db.execute("UPDATE client_brands SET name=?, query=?, notes=? "
               "WHERE id=? AND search_id=?",
               (name.strip(), query.strip(), notes.strip(), bid, sid))
    db.commit(); db.close()
    return RedirectResponse(f"/clients/{cid}/searches/{sid}", status_code=303)


@app.post("/clients/{cid}/searches/{sid}/brands/{bid}/delete")
async def delete_brand(cid: int, sid: int, bid: int):
    db = get_db()
    db.execute("DELETE FROM client_brands WHERE id=? AND search_id=?", (bid, sid))
    db.commit(); db.close()
    return RedirectResponse(f"/clients/{cid}/searches/{sid}", status_code=303)


@app.get("/sites", response_class=HTMLResponse)
async def sites_list(request: Request):
    db    = get_db()
    today = date.today().isoformat()
    sites = db.execute("""
        SELECT s.*,
               COALESCE(td.c, 0) AS today_count,
               COALESCE(tt.c, 0) AS total_count,
               st.monthly_visits  AS latest_visits,
               st.source          AS latest_traffic_source
        FROM sites s
        LEFT JOIN (SELECT site_id, COUNT(*) AS c FROM articles WHERE scrape_date=? GROUP BY site_id) td ON td.site_id=s.id
        LEFT JOIN (SELECT site_id, COUNT(*) AS c FROM articles GROUP BY site_id) tt ON tt.site_id=s.id
        LEFT JOIN (
            SELECT t.site_id, t.monthly_visits, t.source
            FROM site_traffic t
            JOIN (SELECT site_id, MAX(measured_at) AS ma FROM site_traffic WHERE monthly_visits>0 GROUP BY site_id) m
                 ON t.site_id=m.site_id AND t.measured_at=m.ma
        ) st ON st.site_id=s.id
        ORDER BY s.name
    """, (today,)).fetchall()

    # Aggregate stats for KPI cards
    total = len(sites)
    active = sum(1 for s in sites if s["active"])
    errors = sum(1 for s in sites if s["last_scrape_status"] == "error")
    empty  = sum(1 for s in sites if s["last_scrape_status"] == "empty")
    with_traffic = sum(1 for s in sites if (s["latest_visits"] or s["monthly_visits_manual"]))
    total_visits = sum((s["latest_visits"] or s["monthly_visits_manual"] or 0) for s in sites)

    # By country
    country_map: dict = {}
    for s in sites:
        code = s["country"] or "WW"
        name, flag = COUNTRY_BY_CODE.get(code, (code, "🌐"))
        e = country_map.setdefault(code, {"code": code, "name": name, "flag": flag, "total": 0,
                                          "active": 0, "with_traffic": 0, "visits": 0})
        e["total"] += 1
        if s["active"]: e["active"] += 1
        v = s["latest_visits"] or s["monthly_visits_manual"] or 0
        if v: e["with_traffic"] += 1; e["visits"] += v
    by_country = sorted(country_map.values(), key=lambda x: x["total"], reverse=True)

    # By category
    cat_map: dict = {}
    for s in sites:
        cat = s["category"] or "(sin categoría)"
        e = cat_map.setdefault(cat, {"name": cat, "total": 0, "active": 0, "visits": 0})
        e["total"] += 1
        if s["active"]: e["active"] += 1
        e["visits"] += s["latest_visits"] or s["monthly_visits_manual"] or 0
    by_category = sorted(cat_map.values(), key=lambda x: x["total"], reverse=True)

    db.close()
    stats = {"total": total, "active": active, "paused": total - active,
             "errors": errors, "empty": empty,
             "with_traffic": with_traffic, "without_traffic": total - with_traffic,
             "traffic_pct": round(with_traffic / total * 100) if total else 0,
             "total_visits": total_visits}
    return templates.TemplateResponse(request=request, name="sites.html", context=dict(
        request=request, active_page="sites", sites=sites,
        countries=COUNTRIES, country_by_code=COUNTRY_BY_CODE,
        stats=stats, by_country=by_country, by_category=by_category))


@app.post("/sites")
async def add_site(name: str = Form(...), url: str = Form(...),
                   css_selector: str = Form(""), use_stealth: int = Form(0),
                   article_url_pattern: str = Form(""),
                   category: str = Form(""), country: str = Form("WW")):
    db = get_db()
    try:
        db.execute(
            "INSERT INTO sites (name,url,css_selector,use_stealth,article_url_pattern,category,country) "
            "VALUES (?,?,?,?,?,?,?)",
            (name, url.strip(), css_selector.strip(), use_stealth,
             article_url_pattern.strip(), category.strip(), country.strip() or "WW")
        )
        db.commit()
    except Exception as _ie:
        if "unique" not in str(_ie).lower() and "duplicate" not in str(_ie).lower():
            raise
        pass
    finally:
        db.close()
    return RedirectResponse("/sites", status_code=303)


def resolve_media_url(name: str) -> dict:
    """Use Gemini to resolve a media outlet name → homepage URL."""
    api_key = os.environ.get("GEMINI_API_KEY", "")
    if not api_key:
        return {"name": name, "url": "", "found": False, "reason": "no API key"}
    try:
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=api_key)
        prompt = (
            "Eres experto en medios de comunicación. Para el siguiente medio, "
            "devuelve SÓLO un JSON válido sin markdown.\n\n"
            f"Medio: {name}\n\n"
            "JSON requerido:\n"
            "{\n"
            '  "official_name": "nombre oficial del medio",\n'
            '  "url": "https://www.dominio.com/" (homepage exacta, con https y trailing slash),\n'
            '  "found": true si conoces el medio con certeza, false si no estás seguro\n'
            "}"
        )
        resp = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=prompt,
            config=types.GenerateContentConfig(response_mime_type="application/json"),
        )
        data = json.loads(resp.text)
        return {
            "name":  data.get("official_name") or name,
            "url":   data.get("url", ""),
            "found": bool(data.get("found")),
        }
    except Exception as e:
        return {"name": name, "url": "", "found": False, "reason": str(e)}


_VALID_COUNTRY_CODES = {c for c, _, _ in COUNTRIES}
_COUNTRY_NAME_TO_CODE = {n.lower(): c for c, n, _ in COUNTRIES}

def _normalize_country(raw: str) -> str:
    """Accept ISO code (es, ES) or country name (España, Spain → ES). Default WW."""
    if not raw: return "WW"
    raw = raw.strip()
    up = raw.upper()
    if up in _VALID_COUNTRY_CODES: return up
    if raw.lower() in _COUNTRY_NAME_TO_CODE: return _COUNTRY_NAME_TO_CODE[raw.lower()]
    aliases = {"spain":"ES", "españa":"ES", "andorra":"AD", "portugal":"PT",
               "france":"FR", "francia":"FR", "uk":"GB", "united kingdom":"GB",
               "usa":"US", "us":"US", "estados unidos":"US"}
    return aliases.get(raw.lower(), "WW")

def parse_media_list(text: str) -> list[dict]:
    """Parse a text/CSV blob into {name, url, category, country} dicts.
    Accepted line formats (CSV, tab or `;` separated):
      - Just a name:                            "Marca"
      - Name + URL:                             "Marca, https://www.marca.com/"
      - Name + URL + category:                  "Marca, https://www.marca.com/, Sports News"
      - Name + URL + category + country:        "Marca, https://www.marca.com/, Sports News, ES"
    Country can be an ISO code (ES) or name (España). Empty country → WW.
    """
    out = []
    seen = set()
    for line in text.splitlines():
        line = line.strip().strip('"')
        if not line or line.startswith("#"):
            continue
        if line.lower().startswith(("name,", "nombre,", "medio,")):
            continue
        # Tabs / semicolons → commas, then split
        norm = line.replace("\t", ",").replace(";", ",")
        parts = [p.strip().strip('"') for p in norm.split(",")]
        name     = parts[0] if len(parts) > 0 else ""
        url      = parts[1] if len(parts) > 1 and parts[1].startswith("http") else ""
        category = parts[2] if len(parts) > 2 else ""
        country  = _normalize_country(parts[3]) if len(parts) > 3 else "WW"
        if name and name.lower() not in seen:
            seen.add(name.lower())
            out.append({"name": name, "url": url,
                        "category": category, "country": country})
    return out


@app.post("/sites/import/preview")
async def import_preview(file: Optional[UploadFile] = File(None),
                         text: str = Form("")):
    """Parse uploaded file or text, resolve missing URLs via Gemini.
    Returns a list of {name, url, found} for the user to review."""
    blob = ""
    if file is not None and file.filename:
        try:
            blob = (await file.read()).decode("utf-8", errors="replace")
        except Exception as e:
            return JSONResponse({"ok": False, "error": str(e)}, status_code=400)
    if text:
        blob = (blob + "\n" + text) if blob else text

    if not blob.strip():
        return JSONResponse({"ok": False, "error": "empty input"}, status_code=400)

    items = parse_media_list(blob)

    # Already-existing URLs: mark them so the user knows
    db = get_db()
    existing_names = {r["name"].lower() for r in db.execute("SELECT name FROM sites").fetchall()}
    existing_urls  = {r["url"]          for r in db.execute("SELECT url  FROM sites").fetchall()}
    db.close()

    results = []
    for item in items:
        name = item["name"]
        url  = item["url"]
        if not url:
            resolved = resolve_media_url(name)
            url      = resolved["url"]
            found    = resolved["found"]
            official = resolved["name"]
        else:
            found    = True
            official = name
        results.append({
            "name":     official,
            "url":      url,
            "category": item.get("category", ""),
            "country":  item.get("country", "WW"),
            "found":    found,
            "duplicate": (official.lower() in existing_names) or (url in existing_urls),
        })
    return JSONResponse({"ok": True, "items": results,
                         "ai_used": bool(os.environ.get("GEMINI_API_KEY"))})


@app.post("/sites/import/confirm")
async def import_confirm(request: Request):
    """Accept JSON list of sites and add them all."""
    payload = await request.json()
    items   = payload.get("items", [])
    db      = get_db()
    added   = 0
    skipped = 0
    for it in items:
        name = (it.get("name") or "").strip()
        url  = (it.get("url")  or "").strip()
        if not name or not url:
            skipped += 1; continue
        category = (it.get("category") or "").strip()
        country  = _normalize_country(it.get("country") or "WW")
        try:
            db.execute(
                "INSERT INTO sites (name,url,article_url_pattern,category,country) "
                "VALUES (?,?,?,?,?)",
                (name, url, (it.get("article_url_pattern") or "").strip(),
                 category, country))
            added += 1
        except Exception as _ie:
            if "unique" not in str(_ie).lower() and "duplicate" not in str(_ie).lower():
                raise
            skipped += 1
    db.commit(); db.close()
    return JSONResponse({"ok": True, "added": added, "skipped": skipped})


@app.post("/sites/{site_id}/edit")
async def edit_site(site_id: int,
                    name: str = Form(...),
                    url: str = Form(...),
                    css_selector: str = Form(""),
                    use_stealth: int = Form(0),
                    article_url_pattern: str = Form(""),
                    category: str = Form(""),
                    country: str = Form("WW"),
                    cookies_json: str = Form(""),
                    cdp_url: str = Form("")):
    db = get_db()
    try:
        db.execute("""UPDATE sites
                      SET name=?, url=?, css_selector=?, use_stealth=?,
                          article_url_pattern=?, category=?, country=?, cookies_json=?, cdp_url=?
                      WHERE id=?""",
                   (name, url.strip(), css_selector.strip(), use_stealth,
                    article_url_pattern.strip(), category.strip(),
                    country.strip() or "WW", cookies_json.strip(), cdp_url.strip(), site_id))
        db.commit()
    except Exception as _ie:
        if "unique" not in str(_ie).lower() and "duplicate" not in str(_ie).lower():
            raise
        pass
    finally:
        db.close()
    return RedirectResponse("/sites", status_code=303)


@app.post("/sites/{site_id}/delete")
async def delete_site(site_id: int):
    db = get_db()
    db.execute("DELETE FROM sites WHERE id=?", (site_id,))
    db.commit(); db.close()
    return RedirectResponse("/sites", status_code=303)


@app.post("/sites/{site_id}/toggle")
async def toggle_site(site_id: int):
    db = get_db()
    db.execute("UPDATE sites SET active=1-active WHERE id=?", (site_id,))
    db.commit(); db.close()
    return RedirectResponse("/sites", status_code=303)


@app.post("/sites/{site_id}/scrape")
def scrape_site_now(site_id: int):
    """Sync def → FastAPI runs in threadpool so Playwright sync API works."""
    scrape_and_store(site_id)
    return RedirectResponse("/news", status_code=303)


@app.post("/scrape-all")
def scrape_all():
    """Sync def → FastAPI runs in threadpool so Playwright sync API works."""
    db    = get_db()
    sites = db.execute("SELECT id FROM sites WHERE active=1").fetchall()
    db.close()
    for s in sites:
        scrape_and_store(s["id"])
    return RedirectResponse("/news", status_code=303)


PAGE_SIZE = 200

@app.get("/news", response_class=HTMLResponse)
async def news(request: Request, date_filter: Optional[str] = None,
               date_from: Optional[str] = None, date_to: Optional[str] = None,
               site_id: Optional[str] = None, q: Optional[str] = None,
               q2: Optional[str] = None, page: int = 1):
    db     = get_db()
    where  = []
    params = []

    site_id_int = int(site_id) if site_id and site_id.isdigit() else None

    if q:
        # Each word in the query must appear in at least one field (AND between words)
        _news_cols = ("a.title", "a.body", "a.summary",
                      "a.company_property", "a.company_brand", "a.company_agency")
        for _token in (t for t in q.split() if t):
            _like = f"%{_token}%"
            where.append("(" + " OR ".join(f"{c} LIKE ?" for c in _news_cols) + ")")
            params.extend([_like] * len(_news_cols))
        filter_date = ""
    elif date_from or date_to:
        filter_date = ""
        if date_from and date_to:
            where.append("a.scrape_date BETWEEN ? AND ?")
            params.extend([date_from, date_to])
        elif date_from:
            where.append("a.scrape_date >= ?")
            params.append(date_from)
        else:
            where.append("a.scrape_date <= ?")
            params.append(date_to)
    else:
        filter_date = date_filter or date.today().isoformat()
        where.append("a.scrape_date=?")
        params.append(filter_date)

    if site_id_int:
        where.append("a.site_id=?")
        params.append(site_id_int)

    page = max(1, page)
    offset = (page - 1) * PAGE_SIZE
    where_sql = " AND ".join(where) if where else "1=1"

    total_count = db.execute(f"SELECT COUNT(*) FROM articles a JOIN sites s ON s.id=a.site_id WHERE {where_sql}", params).fetchone()[0]

    articles = db.execute(f"""
        SELECT a.*, s.name AS site_name,
            COALESCE(
                (SELECT monthly_visits FROM site_traffic
                 WHERE site_id=a.site_id AND monthly_visits>0
                 ORDER BY measured_at DESC LIMIT 1),
                s.monthly_visits_manual, 0
            ) AS site_monthly_visits
        FROM articles a
        JOIN sites s ON s.id=a.site_id
        WHERE {where_sql}
        ORDER BY a.scrape_date DESC, a.id DESC
        LIMIT ? OFFSET ?
    """, params + [PAGE_SIZE, offset]).fetchall()

    total_pages = max(1, (total_count + PAGE_SIZE - 1) // PAGE_SIZE)

    sites = db.execute("SELECT id,name FROM sites WHERE active=1 ORDER BY name").fetchall()
    dates = [r["scrape_date"] for r in
             db.execute("SELECT DISTINCT scrape_date FROM articles ORDER BY scrape_date DESC LIMIT 60").fetchall()]

    search_kpis  = compute_search_kpis(db, q,  days=30) if q  else None
    search_kpis2 = compute_search_kpis(db, q2, days=30) if q2 else None
    db.close()
    return templates.TemplateResponse(request=request, name="news.html", context=dict(
        request=request, active_page="news", articles=articles,
        sites=sites, filter_date=filter_date, filter_site=site_id_int,
        search_q=q or "", search_q2=q2 or "", available_dates=dates,
        date_from=date_from or "", date_to=date_to or "",
        search_kpis=search_kpis, search_kpis2=search_kpis2,
        page=page, total_pages=total_pages, total_count=total_count))


@app.post("/articles/{article_id}")
async def update_article(article_id: int,
                         company_property: str = Form(""),
                         company_brand:    str = Form(""),
                         company_agency:   str = Form(""),
                         summary:          str = Form("")):
    db = get_db()
    db.execute("""
        UPDATE articles SET company_property=?,company_brand=?,company_agency=?,summary=?
        WHERE id=?
    """, (company_property, company_brand, company_agency, summary, article_id))
    db.commit(); db.close()
    return JSONResponse({"ok": True})


def _enrich_article(article_id: int) -> dict:
    """Reusable enrichment logic shared by HTTP endpoint and scheduler."""
    db  = get_db()
    art = db.execute(
        "SELECT a.*, s.use_stealth, s.name AS site_name, s.cookies_json, s.cdp_url, "
        "s.last_working_tier FROM articles a "
        "JOIN sites s ON s.id=a.site_id WHERE a.id=?",
        (article_id,)
    ).fetchone()
    if not art:
        db.close()
        return {"ok": False, "error": "not found"}

    site_cookies = _parse_cookies(art["cookies_json"] or "")
    site_cdp     = (art["cdp_url"] or "").strip() or None
    site_tier    = int(art["last_working_tier"] or 1)
    full = fetch_article_full(art["article_url"], bool(art["use_stealth"]),
                               cookies=site_cookies, cdp_url=site_cdp,
                               start_tier=site_tier)
    body, images, videos = full["body"], full["images"], full["videos"]
    worked_tier = full.get("tier", site_tier)
    if worked_tier and worked_tier != site_tier:
        db.execute("UPDATE sites SET last_working_tier=? WHERE id=?",
                   (worked_tier, art["site_id"]))
        db.commit()

    # Download media to disk in the background — doesn't block the enrich result
    _site_id_snap = art["site_id"]
    def _bg_media():
        li = _download_media(images, _site_id_snap, article_id, "img")
        lv = _download_media(videos, _site_id_snap, article_id, "vid")
        if li or lv:
            try:
                _db = get_db()
                _db.execute("UPDATE articles SET local_images=?, local_videos=? WHERE id=?",
                            (json.dumps(li), json.dumps(lv), article_id))
                _db.commit(); _db.close()
            except Exception as _e:
                print(f"[media] db update failed: {_e}")
    threading.Thread(target=_bg_media, daemon=True, name=f"media-{article_id}").start()

    ai   = enrich_with_ai(art["title"], body, art["site_name"])

    publishers = [r["name"] for r in db.execute("SELECT name FROM sites").fetchall()]
    def _strip_publishers(value: str) -> str:
        if not value:
            return ""
        items = [x.strip() for x in value.split(",")]
        items = [x for x in items if x and not any(p.lower() in x.lower() for p in publishers)]
        return ", ".join(items)

    if ai:
        new_prop      = _strip_publishers(ai.get("company_property", ""))
        new_brand     = _strip_publishers(ai.get("company_brand", ""))
        new_agency    = _strip_publishers(ai.get("company_agency", ""))
        new_sum       = ai.get("summary", "") or ""
        ai_sentiment  = ai.get("sentiment", "").strip().lower()
        new_sentiment = ai_sentiment if ai_sentiment in ("positive", "negative", "neutral") \
                        else analyze_sentiment_local(art["title"], new_sum, body)
    else:
        new_prop      = art["company_property"] or ""
        new_brand     = art["company_brand"]    or ""
        new_agency    = art["company_agency"]   or ""
        new_sum       = art["summary"]          or ""
        new_sentiment = analyze_sentiment_local(art["title"], new_sum, body)

    db.execute("""
        UPDATE articles SET body=?, company_property=?, company_brand=?,
            company_agency=?, summary=?, images=?, videos=?, sentiment=? WHERE id=?
    """, (body, new_prop, new_brand, new_agency, new_sum,
          json.dumps(images), json.dumps(videos), new_sentiment, article_id))
    db.commit(); db.close()

    return {
        "ok":               True,
        "company_property": new_prop,
        "company_brand":    new_brand,
        "company_agency":   new_agency,
        "summary":          new_sum,
        "sentiment":        new_sentiment,
        "has_body":         bool(body),
        "body_preview":     body[:400] if body else "",
        "images_count":     len(images),
        "videos_count":     len(videos),
        "ai_used":          bool(os.environ.get("GEMINI_API_KEY") or os.environ.get("ANTHROPIC_API_KEY")),
    }


@app.get("/articles/{article_id}/body")
def get_article_body(article_id: int):
    """Return full body text + images + videos of an article."""
    db = get_db()
    art = db.execute(
        "SELECT id, title, article_url, body, images, videos FROM articles WHERE id=?",
        (article_id,)
    ).fetchone()
    db.close()
    if not art:
        return JSONResponse({"ok": False}, status_code=404)

    def _decode(blob):
        if not blob:
            return []
        try:
            v = json.loads(blob)
            return v if isinstance(v, list) else []
        except Exception:
            return []

    return JSONResponse({
        "ok":     True,
        "id":     art["id"],
        "title":  art["title"],
        "url":    art["article_url"],
        "body":   art["body"] or "",
        "images": _decode(art["images"]),
        "videos": _decode(art["videos"]),
    })


@app.post("/articles/{article_id}/enrich")
def enrich_article(article_id: int):
    """Scrape article body and run AI extraction. Sync def → runs in threadpool."""
    result = _enrich_article(article_id)
    if not result.get("ok"):
        return JSONResponse(result, status_code=404)
    return JSONResponse(result)


@app.post("/articles/enrich-stream")
async def enrich_stream(request: Request):
    """Parallel bulk enrichment with SSE progress stream.
    Body: {"ids": [...], "batch_size": 12, "workers": 5}
    Streams: data: {JSON}\\n\\n  per article completed.
    """
    body_data   = await request.json()
    ids         = body_data.get("ids", [])
    batch_size  = min(int(body_data.get("batch_size", 12)), 20)
    n_workers   = min(int(body_data.get("workers", 5)), 8)

    if not ids:
        return JSONResponse({"error": "no ids"}, status_code=400)

    # Preload article meta (title, url, site info) in one query
    db = get_db()
    ph = ",".join(["?"] * len(ids))
    rows = db.execute(
        f"SELECT a.id, a.title, a.article_url, a.company_property, a.company_brand, "
        f"a.company_agency, a.summary, s.name AS site_name, s.use_stealth, "
        f"s.cookies_json, s.cdp_url, s.last_working_tier, s.id AS site_id "
        f"FROM articles a JOIN sites s ON s.id=a.site_id WHERE a.id IN ({ph})",
        ids
    ).fetchall()
    db.close()
    art_map = {r["id"]: dict(r) for r in rows}

    publishers = None  # loaded lazily inside thread

    def _scrape_one(art_id: int) -> dict:
        """Scrape a single article page. Returns article dict enriched with body/images/videos."""
        art = art_map.get(art_id)
        if not art:
            return {"id": art_id, "ok": False, "error": "not found"}
        try:
            cookies   = _parse_cookies(art["cookies_json"] or "")
            cdp_url   = (art["cdp_url"] or "").strip() or None
            start_tier = int(art["last_working_tier"] or 1)
            full = fetch_article_full(art["article_url"], bool(art["use_stealth"]),
                                      cookies=cookies, cdp_url=cdp_url, start_tier=start_tier)
            return {"id": art_id, "ok": True,
                    "title": art["title"], "site_name": art["site_name"],
                    "site_id": art["site_id"],
                    "body": full["body"], "images": full["images"], "videos": full["videos"],
                    "tier": full.get("tier", start_tier)}
        except Exception as e:
            return {"id": art_id, "ok": False, "error": str(e)}

    def _save_and_build_result(art_id: int, scraped: dict, ai_result: dict) -> dict:
        nonlocal publishers
        art  = art_map.get(art_id, {})
        body = scraped.get("body", "")

        if ai_result:
            if publishers is None:
                _db = get_db()
                publishers = [r["name"] for r in _db.execute("SELECT name FROM sites").fetchall()]
                _db.close()
            def _strip(val):
                items = [x.strip() for x in (val or "").split(",")]
                return ", ".join(x for x in items if x and not any(p.lower() in x.lower() for p in publishers))
            new_prop  = _strip(ai_result.get("company_property", ""))
            new_brand = _strip(ai_result.get("company_brand", ""))
            new_agency= _strip(ai_result.get("company_agency", ""))
            new_sum   = ai_result.get("summary", "") or ""
            ai_sent   = (ai_result.get("sentiment") or "").strip().lower()
            new_sent  = ai_sent if ai_sent in ("positive", "negative", "neutral") \
                        else analyze_sentiment_local(art.get("title",""), new_sum, body)
        else:
            new_prop  = art.get("company_property", "") or ""
            new_brand = art.get("company_brand", "")    or ""
            new_agency= art.get("company_agency", "")   or ""
            new_sum   = art.get("summary", "")          or ""
            new_sent  = analyze_sentiment_local(art.get("title",""), new_sum, body)

        images = scraped.get("images", [])
        videos = scraped.get("videos", [])
        _db = get_db()
        _db.execute("""UPDATE articles SET body=?, company_property=?, company_brand=?,
                       company_agency=?, summary=?, images=?, videos=?, sentiment=? WHERE id=?""",
                    (body, new_prop, new_brand, new_agency, new_sum,
                     json.dumps(images), json.dumps(videos), new_sent, art_id))
        _db.commit(); _db.close()

        # Background media download
        site_id = scraped.get("site_id") or art.get("site_id")
        if (images or videos) and site_id:
            def _bg():
                li = _download_media(images, site_id, art_id, "img")
                lv = _download_media(videos, site_id, art_id, "vid")
                if li or lv:
                    try:
                        _d = get_db()
                        _d.execute("UPDATE articles SET local_images=?, local_videos=? WHERE id=?",
                                   (json.dumps(li), json.dumps(lv), art_id))
                        _d.commit(); _d.close()
                    except Exception: pass
            threading.Thread(target=_bg, daemon=True).start()

        return {"ok": True, "id": art_id,
                "company_property": new_prop, "company_brand": new_brand,
                "company_agency": new_agency, "summary": new_sum,
                "sentiment": new_sent, "has_body": bool(body),
                "body_preview": body[:400] if body else "",
                "images_count": len(images), "videos_count": len(videos)}

    async def event_stream():
        total   = len(ids)
        done    = 0
        batches = [ids[i:i+batch_size] for i in range(0, len(ids), batch_size)]

        for batch in batches:
            # ── Phase 1: parallel scrape ──────────────────────────────────────
            loop      = asyncio.get_event_loop()
            futures   = [loop.run_in_executor(None, _scrape_one, aid) for aid in batch]
            scraped_list = await asyncio.gather(*futures)
            scraped_ok   = [s for s in scraped_list if s.get("ok")]
            scraped_fail = [s for s in scraped_list if not s.get("ok")]

            # ── Phase 2: batch AI ─────────────────────────────────────────────
            ai_results: dict[int, dict] = {}
            if scraped_ok:
                ai_results = await loop.run_in_executor(None, enrich_with_ai_batch, scraped_ok)

            # ── Phase 3: save + stream results ───────────────────────────────
            for s in scraped_ok:
                ai  = ai_results.get(s["id"], {})
                res = await loop.run_in_executor(None, _save_and_build_result, s["id"], s, ai)
                done += 1
                res["done"]  = done
                res["total"] = total
                yield f"data: {json.dumps(res)}\n\n"

            for s in scraped_fail:
                done += 1
                yield f"data: {json.dumps({'ok': False, 'id': s['id'], 'error': s.get('error',''), 'done': done, 'total': total})}\n\n"

        yield f"data: {json.dumps({'type': 'done', 'done': done, 'total': total})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/articles/{article_id}/sentiment")
def article_sentiment(article_id: int):
    """Compute sentiment for a single article using local keyword analysis (instant, no AI quota)."""
    db  = get_db()
    art = db.execute("SELECT id, title, summary, body FROM articles WHERE id=?", (article_id,)).fetchone()
    if not art:
        db.close()
        return JSONResponse({"ok": False, "error": "not found"}, status_code=404)
    sentiment = analyze_sentiment_local(art["title"], art["summary"] or "", art["body"] or "")
    db.execute("UPDATE articles SET sentiment=? WHERE id=?", (sentiment, article_id))
    db.commit(); db.close()
    return JSONResponse({"ok": True, "id": article_id, "sentiment": sentiment})


@app.post("/articles/sentiment-bulk")
async def articles_sentiment_bulk(request: Request):
    """Compute sentiment for a list of article IDs (or all if ids=[])."""
    body = await request.json()
    ids  = body.get("ids", [])
    db   = get_db()
    if ids:
        rows = db.execute(
            f"SELECT id,title,summary,body FROM articles WHERE id IN ({','.join('?'*len(ids))})", ids
        ).fetchall()
    else:
        rows = db.execute("SELECT id,title,summary,body FROM articles").fetchall()
    updated = 0
    for r in rows:
        s = analyze_sentiment_local(r["title"], r["summary"] or "", r["body"] or "")
        db.execute("UPDATE articles SET sentiment=? WHERE id=?", (s, r["id"]))
        updated += 1
    db.commit(); db.close()
    return JSONResponse({"ok": True, "updated": updated})


def auto_enrich_site(site_id: int, max_articles: int = 80, timeout_mins: int = 8):
    """Enrich articles for this site that don't have a body yet.
    Stops after timeout_mins to prevent blocking the worker indefinitely."""
    db = get_db()
    arts = db.execute("""
        SELECT id FROM articles
        WHERE site_id=? AND (body='' OR body IS NULL)
        ORDER BY id DESC LIMIT ?
    """, (site_id, max_articles)).fetchall()
    db.close()
    print(f"[auto-enrich] site={site_id}, {len(arts)} articles to enrich (max {timeout_mins}min)")
    deadline = time.time() + timeout_mins * 60
    for a in arts:
        if time.time() > deadline:
            print(f"[auto-enrich] site={site_id} timeout after {timeout_mins}min, stopping")
            break
        try:
            _enrich_article(a["id"])
            time.sleep(0.5)
        except Exception as e:
            print(f"[auto-enrich] failed for article {a['id']}: {e}")


@app.get("/aggregated", response_class=HTMLResponse)
async def aggregated(request: Request, site_id: Optional[str] = None,
                     category: Optional[str] = None, q: Optional[str] = None,
                     page: int = 1):
    db     = get_db()
    where  = ["1=1"]
    params = []

    site_id_int = int(site_id) if site_id and site_id.isdigit() else None

    if site_id_int:
        where.append("a.site_id=?"); params.append(site_id_int)
    if category:
        where.append("a.category=?"); params.append(category)
    if q:
        where.append("(a.title LIKE ? OR a.summary LIKE ? OR a.company_brand LIKE ? OR a.body LIKE ?)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"]
    base   = f"FROM articles a JOIN sites s ON s.id=a.site_id WHERE {' AND '.join(where)}"
    total  = db.execute(f"SELECT COUNT(*) {base}", params).fetchone()[0]
    per_pg = 50
    offset = (page - 1) * per_pg
    rows   = db.execute(
        f"""SELECT a.*, s.name AS site_name,
            COALESCE(
                (SELECT monthly_visits FROM site_traffic
                 WHERE site_id=a.site_id AND monthly_visits>0
                 ORDER BY measured_at DESC LIMIT 1),
                s.monthly_visits_manual, 0
            ) AS site_monthly_visits
            {base} ORDER BY a.scrape_date DESC,a.id DESC
            LIMIT {per_pg} OFFSET {offset}""", params
    ).fetchall()
    sites  = db.execute("SELECT id,name FROM sites ORDER BY name").fetchall()
    cats   = [r["category"] for r in
              db.execute("SELECT DISTINCT category FROM articles WHERE category!='' ORDER BY category").fetchall()]
    db.close()
    return templates.TemplateResponse(request=request, name="aggregated.html", context=dict(
        request=request, active_page="aggregated", articles=rows,
        sites=sites, categories=cats, filter_site=site_id_int,
        filter_category=category, search_q=q or "",
        total=total, page=page, per_page=per_pg,
        total_pages=max(1, (total + per_pg - 1) // per_pg)))


@app.get("/email", response_class=HTMLResponse)
async def email_page(request: Request, article_ids: Optional[str] = None):
    db       = get_db()
    selected = []
    if article_ids:
        ids = [int(i) for i in article_ids.split(",") if i.strip().isdigit()]
        if ids:
            ph       = ",".join("?" * len(ids))
            selected = db.execute(f"""
                SELECT a.*,s.name AS site_name FROM articles a
                JOIN sites s ON s.id=a.site_id WHERE a.id IN ({ph})
            """, ids).fetchall()
    today_arts = db.execute("""
        SELECT a.*,s.name AS site_name FROM articles a
        JOIN sites s ON s.id=a.site_id
        WHERE a.scrape_date=date('now') ORDER BY a.site_id, a.id
    """).fetchall()
    logs = db.execute("SELECT * FROM email_logs ORDER BY sent_at DESC LIMIT 10").fetchall()
    db.close()
    return templates.TemplateResponse(request=request, name="email.html", context=dict(
        request=request, active_page="email",
        selected_articles=selected, today_articles=today_arts, email_logs=logs))


@app.post("/email/log")
async def log_email(subject: str = Form(""), recipients: str = Form(""),
                    article_ids: str = Form(""), body: str = Form("")):
    db = get_db()
    db.execute("INSERT INTO email_logs (subject,recipients,article_ids,body) VALUES (?,?,?,?)",
               (subject, recipients, article_ids, body))
    db.commit(); db.close()
    return JSONResponse({"ok": True})


# ── settings & scheduler ──────────────────────────────────────────────────────

@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    db = get_db()
    today_s = date.today().isoformat()
    sites = db.execute("""
        SELECT s.*,
               COALESCE(tt.c, 0) AS total_count,
               COALESCE(td.c, 0) AS today_count,
               st.monthly_visits  AS latest_visits,
               st.source          AS latest_source
        FROM sites s
        LEFT JOIN (SELECT site_id, COUNT(*) AS c FROM articles GROUP BY site_id) tt ON tt.site_id=s.id
        LEFT JOIN (SELECT site_id, COUNT(*) AS c FROM articles WHERE scrape_date=? GROUP BY site_id) td ON td.site_id=s.id
        LEFT JOIN (
            SELECT t.site_id, t.monthly_visits, t.source
            FROM site_traffic t
            JOIN (SELECT site_id, MAX(measured_at) AS ma FROM site_traffic WHERE monthly_visits>0 GROUP BY site_id) m
                 ON t.site_id=m.site_id AND t.measured_at=m.ma
        ) st ON st.site_id=s.id
        ORDER BY s.name
    """, (today_s,)).fetchall()
    db.close()
    # Mask stored keys for display (show last 4 chars only)
    def _mask(key: str) -> str:
        return f"{'•' * (len(key) - 4)}{key[-4:]}" if len(key) > 4 else "•" * len(key)
    gemini_key     = os.environ.get("GEMINI_API_KEY", "")
    anthropic_key  = os.environ.get("ANTHROPIC_API_KEY", "")
    return templates.TemplateResponse(request=request, name="settings.html", context=dict(
        request=request, active_page="settings", sites=sites,
        countries=COUNTRIES, country_by_code=COUNTRY_BY_CODE,
        ai_configured=bool(gemini_key or anthropic_key),
        gemini_key_set=bool(gemini_key),
        gemini_key_masked=_mask(gemini_key) if gemini_key else "",
        anthropic_key_set=bool(anthropic_key),
        anthropic_key_masked=_mask(anthropic_key) if anthropic_key else "",
        dotenv_available=_DOTENV_AVAILABLE))


@app.post("/settings/api-keys")
async def save_api_keys(gemini_key: str = Form(""), anthropic_key: str = Form("")):
    """Save API keys to .env file and update os.environ in-process (no restart needed)."""
    gemini_key    = gemini_key.strip()
    anthropic_key = anthropic_key.strip()

    def _set(key: str, value: str):
        if value:
            os.environ[key] = value
            if _DOTENV_AVAILABLE and _ENV_FILE:
                set_key(str(_ENV_FILE), key, value)
        # Empty string = "leave unchanged" (don't delete)

    _set("GEMINI_API_KEY",    gemini_key)
    _set("ANTHROPIC_API_KEY", anthropic_key)
    return RedirectResponse("/settings?saved=1", status_code=303)


@app.post("/settings/api-keys/delete")
async def delete_api_key(key_name: str = Form(...)):
    """Remove an API key from os.environ and .env."""
    if key_name in ("GEMINI_API_KEY", "ANTHROPIC_API_KEY"):
        os.environ.pop(key_name, None)
        if _DOTENV_AVAILABLE and _ENV_FILE:
            set_key(str(_ENV_FILE), key_name, "")
    return RedirectResponse("/settings?deleted=1", status_code=303)


@app.post("/sites/{site_id}/schedule")
def update_schedule(site_id: int,
                    scrape_frequency_days:  int = Form(0),
                    auto_enrich:            int = Form(0),
                    traffic_frequency_days: int = Form(5),
                    monthly_visits_manual:  int = Form(0),
                    country:                str = Form("WW")):
    db = get_db()
    db.execute("""UPDATE sites
                  SET scrape_frequency_days=?, auto_enrich=?,
                      traffic_frequency_days=?, monthly_visits_manual=?, country=?
                  WHERE id=?""",
               (scrape_frequency_days, auto_enrich,
                traffic_frequency_days, monthly_visits_manual, country or "WW", site_id))
    db.commit()
    row = db.execute("""SELECT scrape_frequency_days, auto_enrich, last_scraped,
                               traffic_frequency_days, monthly_visits_manual,
                               last_traffic_check
                        FROM sites WHERE id=?""", (site_id,)).fetchone()
    db.close()
    return JSONResponse({
        "ok":                     True,
        "scrape_frequency_days":  row["scrape_frequency_days"],
        "auto_enrich":            bool(row["auto_enrich"]),
        "last_scraped":           row["last_scraped"],
        "traffic_frequency_days": row["traffic_frequency_days"],
        "monthly_visits_manual":  row["monthly_visits_manual"],
        "last_traffic_check":     row["last_traffic_check"],
    })


@app.post("/sites/{site_id}/traffic/refresh")
def refresh_site_traffic(site_id: int):
    """Trigger a manual traffic refresh for one site (sync, may take 30s)."""
    result = update_site_traffic(site_id)
    return JSONResponse(result)


@app.get("/scheduler/status")
def scheduler_status():
    """Real-time scheduler stats + today's queue breakdown."""
    today = date.today().isoformat()
    db    = get_db()
    rows  = db.execute("""
        SELECT status, COUNT(*) AS n, SUM(new_articles) AS arts, COUNT(error) FILTER (WHERE error!='') AS errs
        FROM scrape_queue WHERE job_date=?
        GROUP BY status
    """, (today,)).fetchall()
    total_active = db.execute("SELECT COUNT(*) FROM sites WHERE active=1 AND scrape_frequency_days>0").fetchone()[0]
    recent_errors = db.execute("""
        SELECT s.name, q.error, q.completed_at
        FROM scrape_queue q JOIN sites s ON s.id=q.site_id
        WHERE q.job_date=? AND q.status='error' AND q.error!=''
        ORDER BY q.completed_at DESC LIMIT 20
    """, (today,)).fetchall()
    db.close()

    by_status = {r["status"]: {"count": r["n"], "articles": r["arts"] or 0} for r in rows}
    total_queued = sum(v["count"] for v in by_status.values())
    done         = by_status.get("done",    {}).get("count", 0)
    pending      = by_status.get("pending", {}).get("count", 0)
    running      = by_status.get("running", {}).get("count", 0)
    error_count  = by_status.get("error",   {}).get("count", 0)
    total_new    = sum(v.get("articles", 0) for v in by_status.values())
    pct          = round(done / total_queued * 100) if total_queued else 0

    with _sched_lock:
        stats = dict(_sched_stats)

    return JSONResponse({
        "ok":            True,
        "today":         today,
        "total_active_sites": total_active,
        "queue": {
            "total":   total_queued,
            "pending": pending,
            "running": running,
            "done":    done,
            "error":   error_count,
            "pct_done": pct,
            "new_articles": total_new,
        },
        "workers": {
            "scrape_workers":  SCRAPE_WORKERS,
            "traffic_workers": TRAFFIC_WORKERS,
            "active":          stats["active_workers"],
        },
        "process": {
            "started_at":  stats["started_at"],
            "total_done":  stats["total_done"],
            "total_ok":    stats["total_ok"],
            "total_error": stats["total_error"],
            "total_new":   stats["total_new"],
            "last_fill_at": stats["last_fill_at"],
            "last_fill_n":  stats["last_fill_n"],
        },
        "recent_errors": [{"site": r["name"], "error": r["error"][:120], "at": r["completed_at"]}
                          for r in recent_errors],
    })


@app.post("/scheduler/queue/fill")
def scheduler_force_fill():
    """Manually trigger a queue fill (useful for testing or after adding many sites)."""
    added = _queue_fill()
    return JSONResponse({"ok": True, "added": added})


@app.post("/scheduler/queue/site/{site_id}")
def scheduler_enqueue_site(site_id: int, priority: int = 10):
    """Force-enqueue a specific site with high priority (skips frequency check)."""
    today = date.today().isoformat()
    db    = get_db()
    site  = db.execute("SELECT id, name FROM sites WHERE id=?", (site_id,)).fetchone()
    if not site:
        db.close(); return JSONResponse({"ok": False, "error": "site not found"}, status_code=404)
    db.execute("""
        INSERT INTO scrape_queue (site_id, job_date, status, priority)
        VALUES (?, ?, 'pending', ?)
        ON CONFLICT(site_id, job_date) DO UPDATE SET status='pending', priority=excluded.priority
    """, (site_id, today, priority))
    db.commit(); db.close()
    return JSONResponse({"ok": True, "site": site["name"], "priority": priority})


@app.post("/settings/traffic/refresh-all")
async def refresh_all_traffic(background_tasks: BackgroundTasks):
    """Queue a traffic refresh for all active sites (runs in background)."""
    db = get_db()
    site_ids = [r["id"] for r in db.execute(
        "SELECT id FROM sites WHERE active=1 ORDER BY name"
    ).fetchall()]
    db.close()

    def run_all():
        for sid in site_ids:
            try:
                update_site_traffic(sid)
            except Exception as e:
                print(f"[bulk-traffic] site {sid} error: {e}")

    background_tasks.add_task(run_all)
    return JSONResponse({"ok": True, "queued": len(site_ids)})


@app.post("/sites/{site_id}/traffic/manual")
async def set_manual_traffic(site_id: int, monthly_visits: int = Form(0)):
    """Set the manual override visits/month for sites where Similarweb has no data."""
    db = get_db()
    db.execute("UPDATE sites SET monthly_visits_manual=? WHERE id=?",
               (monthly_visits, site_id))
    db.commit(); db.close()
    return JSONResponse({"ok": True, "monthly_visits_manual": monthly_visits})


@app.get("/sites/{site_id}/traffic")
def get_site_traffic(site_id: int):
    """Return time-series of traffic snapshots for one site."""
    db   = get_db()
    site = db.execute("SELECT id, name, url, last_traffic_check, monthly_visits_manual FROM sites WHERE id=?",
                      (site_id,)).fetchone()
    if not site:
        db.close(); return JSONResponse({"ok": False}, status_code=404)
    rows = db.execute("""
        SELECT measured_at, monthly_visits, bounce_rate, pages_per_visit,
               avg_duration_sec, global_rank, country_rank, source, error
        FROM site_traffic WHERE site_id=? ORDER BY measured_at ASC
    """, (site_id,)).fetchall()
    # Latest = most recent row that actually got data
    latest = next((r for r in reversed(rows) if r["monthly_visits"]), rows[-1] if rows else None)
    db.close()
    return JSONResponse({
        "ok":               True,
        "site":             {"id": site["id"], "name": site["name"], "url": site["url"]},
        "last_check":       site["last_traffic_check"],
        "manual_visits":    site["monthly_visits_manual"] or 0,
        "history":          [dict(r) for r in rows],
        "latest":           dict(latest) if latest else None,
    })


# ── Parallel scheduler ────────────────────────────────────────────────────────
# Configurable via environment variables:
#   SCRAPE_WORKERS      number of parallel scrape threads   (default 20)
#   TRAFFIC_WORKERS     number of parallel traffic threads  (default 5)
#   QUEUE_FILL_INTERVAL seconds between queue-fill checks   (default 60)

SCRAPE_WORKERS      = int(os.environ.get("SCRAPE_WORKERS",      20))
TRAFFIC_WORKERS     = int(os.environ.get("TRAFFIC_WORKERS",      5))
QUEUE_FILL_INTERVAL = int(os.environ.get("QUEUE_FILL_INTERVAL", 60))

# In-process stats (reset on restart)
_sched_stats: dict = {
    "started_at":    None,
    "total_done":    0,
    "total_ok":      0,
    "total_error":   0,
    "total_new":     0,
    "active_workers": 0,
    "last_fill_at":  None,
    "last_fill_n":   0,
}
_sched_lock = threading.Lock()


def _queue_fill():
    """Insert pending jobs for every site that is due today and not yet queued."""
    today = date.today().isoformat()
    db    = get_db()
    due   = db.execute("""
        SELECT id, name, scrape_frequency_days FROM sites
        WHERE active=1 AND scrape_frequency_days > 0
          AND (last_scraped IS NULL OR
               julianday('now') - julianday(last_scraped) >= scrape_frequency_days)
          AND id NOT IN (
              SELECT site_id FROM scrape_queue
              WHERE job_date=? AND status IN ('pending','running','done')
          )
    """, (today,)).fetchall()

    added = 0
    for s in due:
        try:
            db.execute(
                "INSERT OR IGNORE INTO scrape_queue (site_id, job_date, status, priority) VALUES (?,?,?,?)",
                (s["id"], today, "pending", 5)
            )
            added += 1
        except Exception:
            pass
    if added:
        db.commit()
    db.close()

    with _sched_lock:
        _sched_stats["last_fill_at"] = datetime.now().isoformat()
        _sched_stats["last_fill_n"]  = added
    if added:
        print(f"[scheduler] queued {added} sites for {today}")
    return added


def _claim_job(worker_id: str) -> Optional[dict]:
    """Atomically claim one pending job from the queue."""
    db  = get_db()
    row = db.execute("""
        SELECT q.id, q.site_id, s.name, s.auto_enrich
        FROM scrape_queue q
        JOIN sites s ON s.id = q.site_id
        WHERE q.status='pending'
        ORDER BY q.priority DESC, q.enqueued_at ASC
        LIMIT 1
    """).fetchone()
    if not row:
        db.close(); return None
    db.execute("""
        UPDATE scrape_queue
        SET status='running', started_at=?, worker_id=?
        WHERE id=? AND status='pending'
    """, (datetime.now().isoformat(), worker_id, row["id"]))
    db.commit()
    db.close()
    return dict(row)


def _finish_job(job_id: int, new_articles: int, error: str = ""):
    status = "error" if error else "done"
    db = get_db()
    db.execute("""
        UPDATE scrape_queue
        SET status=?, completed_at=?, new_articles=?, error=?
        WHERE id=?
    """, (status, datetime.now().isoformat(), new_articles, error[:400], job_id))
    db.commit(); db.close()
    with _sched_lock:
        _sched_stats["total_done"]  += 1
        _sched_stats["total_new"]   += new_articles
        if error: _sched_stats["total_error"] += 1
        else:     _sched_stats["total_ok"]    += 1


def _scrape_worker(worker_id: str):
    """Long-running worker thread: continuously pulls jobs from the queue."""
    with _sched_lock:
        _sched_stats["active_workers"] += 1
    try:
        while True:
            job = _claim_job(worker_id)
            if not job:
                time.sleep(5)  # queue empty, wait and retry
                continue
            error = ""
            new_c = 0
            try:
                new_c = scrape_and_store(job["site_id"])
                if job["auto_enrich"] and new_c > 0:
                    auto_enrich_site(job["site_id"])
            except Exception as e:
                error = str(e)[:400]
                print(f"[worker:{worker_id}] error site={job['site_id']} ({job['name']}): {error[:120]}")
            _finish_job(job["id"], new_c, error)
    except Exception as e:
        print(f"[worker:{worker_id}] fatal: {e}")
    finally:
        with _sched_lock:
            _sched_stats["active_workers"] -= 1


def _traffic_worker(worker_id: str):
    """Separate pool for traffic refreshes (lighter, fewer workers needed)."""
    while True:
        db  = get_db()
        due = db.execute("""
            SELECT id, name FROM sites
            WHERE active=1 AND traffic_frequency_days > 0
              AND (last_traffic_check IS NULL OR
                   julianday('now') - julianday(last_traffic_check) >= traffic_frequency_days)
            LIMIT 1
        """).fetchone()
        db.close()

        if not due:
            time.sleep(60); continue
        try:
            update_site_traffic(due["id"])
        except Exception as e:
            print(f"[traffic:{worker_id}] error site={due['id']}: {e}")
        time.sleep(2)  # small gap between traffic requests


def scheduler_loop():
    """Coordinator thread: fills the scrape queue every minute, keeps workers alive."""
    print(f"[scheduler] started · {SCRAPE_WORKERS} scrape workers · {TRAFFIC_WORKERS} traffic workers")
    with _sched_lock:
        _sched_stats["started_at"] = datetime.now().isoformat()

    # Start scrape worker pool
    for i in range(SCRAPE_WORKERS):
        t = threading.Thread(target=_scrape_worker, args=(f"S{i:02d}",), daemon=True, name=f"scrape-{i:02d}")
        t.start()

    # Start traffic worker pool
    for i in range(TRAFFIC_WORKERS):
        t = threading.Thread(target=_traffic_worker, args=(f"T{i:02d}",), daemon=True, name=f"traffic-{i:02d}")
        t.start()

    # Coordinator: fill queue periodically
    while True:
        try:
            _queue_fill()
        except Exception as e:
            print(f"[scheduler] fill error: {e}")
        time.sleep(QUEUE_FILL_INTERVAL)


# ── Export helpers ─────────────────────────────────────────────────────────────

def _xl_workbook():
    """Return openpyxl + common style objects."""
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    return wb, openpyxl, PatternFill, Font, Alignment, Border, Side, get_column_letter

BRAND_COLOR  = "0D9488"   # teal-600
HEADER_COLOR = "0F172A"   # slate-900
SUBH_COLOR   = "1E293B"   # slate-800
ACCENT_COLOR = "134E4A"   # dark teal
LIGHT_FILL   = "F0FDFA"   # teal-50
ALT_FILL     = "F8FAFC"   # slate-50
REPORT_FONT  = "Calibri"

def _style_header_row(ws, row_idx, ncols, PatternFill, Font, Alignment, Border, Side,
                      bg=HEADER_COLOR, fg="FFFFFF", size=10):
    fill   = PatternFill("solid", fgColor=bg)
    font   = Font(name=REPORT_FONT, bold=True, color=fg, size=size)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill   = fill
        cell.font   = font
        cell.alignment = align
        cell.border = border

def _style_data_row(ws, row_idx, ncols, PatternFill, Font, Alignment, even=True):
    fill  = PatternFill("solid", fgColor=ALT_FILL if even else "FFFFFF")
    font  = Font(name=REPORT_FONT, size=9)
    align = Alignment(vertical="center")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align

def _xl_cover_sheet(wb, title, subtitle, PatternFill, Font, Alignment, get_column_letter):
    ws = wb.active
    ws.title = "Portada"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 14
    ws.row_dimensions[1].height  = 6
    ws.row_dimensions[2].height  = 60
    ws.row_dimensions[3].height  = 30
    ws.row_dimensions[4].height  = 20
    ws.row_dimensions[5].height  = 14
    ws.row_dimensions[6].height  = 14
    ws.row_dimensions[7].height  = 14
    ws.row_dimensions[8].height  = 60

    # Header banner
    ws.merge_cells("B2:J2")
    cell = ws["B2"]
    cell.value     = "SPORTS INTEL"
    cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
    cell.font      = Font(name=REPORT_FONT, bold=True, color="FFFFFF", size=28)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("B3:J3")
    cell = ws["B3"]
    cell.value     = title
    cell.fill      = PatternFill("solid", fgColor=BRAND_COLOR)
    cell.font      = Font(name=REPORT_FONT, bold=True, color="FFFFFF", size=16)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("B4:J4")
    cell = ws["B4"]
    cell.value     = subtitle
    cell.fill      = PatternFill("solid", fgColor=BRAND_COLOR)
    cell.font      = Font(name=REPORT_FONT, color="D1FAE5", size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("B5:J5")
    ws["B5"].fill = PatternFill("solid", fgColor=BRAND_COLOR)

    ws.merge_cells("B6:J6")
    ws["B6"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["B6"].value = f"Generado el {date.today().strftime('%d/%m/%Y')}"
    ws["B6"].font  = Font(name=REPORT_FONT, color="6B7280", size=9, italic=True)
    ws["B6"].alignment = Alignment(horizontal="left", vertical="center", indent=2)

    ws.merge_cells("B8:J8")
    ws["B8"].fill  = PatternFill("solid", fgColor=LIGHT_FILL)
    ws["B8"].value = "Confidencial · Sports Intel / Scraper Hub"
    ws["B8"].font  = Font(name=REPORT_FONT, color="9CA3AF", size=8, italic=True)
    ws["B8"].alignment = Alignment(horizontal="center", vertical="center")
    return ws


# ── Excel: Sites ───────────────────────────────────────────────────────────────

@app.get("/export/sites.xlsx")
def export_sites_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb, *_ = _xl_workbook()

    _xl_cover_sheet(wb, "Informe de Medios", "Listado completo · métricas de tráfico, cobertura y scraping",
                    PatternFill, Font, Alignment, get_column_letter)

    db    = get_db()
    today = date.today().isoformat()
    sites = db.execute("""
        SELECT s.*,
               COALESCE(st.monthly_visits, s.monthly_visits_manual, 0) AS visits,
               COALESCE(st.source, '')       AS traffic_source,
               COALESCE(st.bounce_rate, 0)   AS bounce_rate,
               COALESCE(st.pages_per_visit, 0) AS pages_per_visit,
               COALESCE(st.global_rank, 0)   AS global_rank,
               COALESCE(st.country_rank, 0)  AS country_rank,
               COALESCE(td.c, 0) AS today_c,
               COALESCE(tt.c, 0) AS total_c
        FROM sites s
        LEFT JOIN (SELECT t.site_id, t.monthly_visits, t.source, t.bounce_rate,
                          t.pages_per_visit, t.global_rank, t.country_rank
                   FROM site_traffic t
                   JOIN (SELECT site_id, MAX(measured_at) AS ma FROM site_traffic
                         WHERE monthly_visits>0 GROUP BY site_id) m
                        ON t.site_id=m.site_id AND t.measured_at=m.ma) st ON st.site_id=s.id
        LEFT JOIN (SELECT site_id, COUNT(*) AS c FROM articles WHERE scrape_date=? GROUP BY site_id) td ON td.site_id=s.id
        LEFT JOIN (SELECT site_id, COUNT(*) AS c FROM articles GROUP BY site_id) tt ON tt.site_id=s.id
        ORDER BY s.name
    """, (today,)).fetchall()
    db.close()

    # ── Sheet 1: Resumen ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Resumen")
    ws_sum.sheet_view.showGridLines = False

    active_sites  = [s for s in sites if s["active"]]
    total_visits  = sum(s["visits"] or 0 for s in sites)
    total_arts    = sum(s["total_c"] for s in sites)
    today_arts    = sum(s["today_c"] for s in sites)
    with_traffic  = sum(1 for s in sites if (s["visits"] or 0) > 0)

    # KPI cards
    kpis = [
        ("Medios totales",         len(sites),        None),
        ("Medios activos",         len(active_sites),  "16A34A"),
        ("Con datos de tráfico",   with_traffic,       None),
        ("Visitas/mes totales",    total_visits,       "2563EB"),
        ("Artículos totales",      total_arts,         None),
        ("Artículos hoy",          today_arts,         "0D9488"),
    ]
    ws_sum.column_dimensions["A"].width = 3
    ws_sum.column_dimensions["B"].width = 28
    ws_sum.column_dimensions["C"].width = 20
    ws_sum.column_dimensions["D"].width = 3
    ws_sum.column_dimensions["E"].width = 28
    ws_sum.column_dimensions["F"].width = 20

    ws_sum.merge_cells("B2:F2")
    ws_sum["B2"].value = "RESUMEN DE MEDIOS"
    ws_sum["B2"].fill  = PatternFill("solid", fgColor=HEADER_COLOR)
    ws_sum["B2"].font  = Font(name=REPORT_FONT, bold=True, color="FFFFFF", size=14)
    ws_sum["B2"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws_sum.row_dimensions[2].height = 36

    for i, (label, value, color) in enumerate(kpis):
        row = 4 + (i // 2) * 3
        col_l = "B" if i % 2 == 0 else "E"
        col_v = "C" if i % 2 == 0 else "F"
        ws_sum[f"{col_l}{row}"].value = label
        ws_sum[f"{col_l}{row}"].font  = Font(name=REPORT_FONT, color="6B7280", size=9)
        ws_sum[f"{col_l}{row}"].fill  = PatternFill("solid", fgColor=ALT_FILL)
        ws_sum[f"{col_l}{row+1}"].value = value
        ws_sum[f"{col_l}{row+1}"].font  = Font(name=REPORT_FONT, bold=True,
                                               color=color or HEADER_COLOR, size=18)
        ws_sum[f"{col_l}{row+1}"].number_format = "#,##0"
        ws_sum[f"{col_l}{row+1}"].fill = PatternFill("solid", fgColor=ALT_FILL)
        ws_sum[f"{col_v}{row}"].fill   = PatternFill("solid", fgColor=ALT_FILL)
        ws_sum[f"{col_v}{row+1}"].fill = PatternFill("solid", fgColor=ALT_FILL)
        ws_sum.row_dimensions[row].height   = 16
        ws_sum.row_dimensions[row+1].height = 28

    # By country table
    country_map: dict = {}
    for s in sites:
        code = s["country"] or "WW"
        name, flag = COUNTRY_BY_CODE.get(code, (code, ""))
        e = country_map.setdefault(code, {"name": f"{flag} {name}".strip(), "total": 0,
                                          "active": 0, "visits": 0, "arts": 0})
        e["total"]  += 1
        if s["active"]: e["active"] += 1
        e["visits"] += s["visits"] or 0
        e["arts"]   += s["total_c"]
    by_country = sorted(country_map.values(), key=lambda x: x["visits"], reverse=True)

    start_row = 4 + ((len(kpis) + 1) // 2) * 3 + 2
    ws_sum.merge_cells(f"B{start_row}:F{start_row}")
    ws_sum[f"B{start_row}"].value = "Por país"
    ws_sum[f"B{start_row}"].fill  = PatternFill("solid", fgColor=BRAND_COLOR)
    ws_sum[f"B{start_row}"].font  = Font(name=REPORT_FONT, bold=True, color="FFFFFF", size=10)
    ws_sum[f"B{start_row}"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
    ws_sum.row_dimensions[start_row].height = 22

    ch_headers = ["País", "Medios", "Activos", "Visitas/Mes", "Artículos"]
    ch_cols    = ["B", "C", "D", "E", "F"]
    for col, hdr in zip(ch_cols, ch_headers):
        cell = ws_sum[f"{col}{start_row+1}"]
        cell.value = hdr
        cell.fill  = PatternFill("solid", fgColor=SUBH_COLOR)
        cell.font  = Font(name=REPORT_FONT, bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center" if hdr != "País" else "left",
                                   vertical="center", indent=1 if hdr == "País" else 0)
    ws_sum.row_dimensions[start_row+1].height = 20

    for i, cr in enumerate(by_country):
        r  = start_row + 2 + i
        ev = i % 2 == 0
        fill = PatternFill("solid", fgColor=ALT_FILL if ev else "FFFFFF")
        for col in ch_cols:
            ws_sum[f"{col}{r}"].fill = fill
            ws_sum[f"{col}{r}"].font = Font(name=REPORT_FONT, size=9)
        ws_sum[f"B{r}"].value = cr["name"]
        ws_sum[f"C{r}"].value = cr["total"];  ws_sum[f"C{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum[f"D{r}"].value = cr["active"]; ws_sum[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum[f"E{r}"].value = cr["visits"]; ws_sum[f"E{r}"].number_format = "#,##0"
        ws_sum[f"F{r}"].value = cr["arts"];   ws_sum[f"F{r}"].number_format = "#,##0"
        ws_sum.row_dimensions[r].height = 17

    # By category table (to the right)
    cat_map: dict = {}
    for s in sites:
        cat = s["category"] or "(sin categoría)"
        e = cat_map.setdefault(cat, {"total": 0, "active": 0, "visits": 0})
        e["total"] += 1
        if s["active"]: e["active"] += 1
        e["visits"] += s["visits"] or 0
    by_cat = sorted(cat_map.values(), key=lambda x: x["visits"], reverse=True)

    cat_col_start = 8  # column H
    ws_sum.column_dimensions[get_column_letter(cat_col_start)].width = 3
    ws_sum.column_dimensions[get_column_letter(cat_col_start+1)].width = 26
    ws_sum.column_dimensions[get_column_letter(cat_col_start+2)].width = 12
    ws_sum.column_dimensions[get_column_letter(cat_col_start+3)].width = 18

    cat_start_row = start_row
    ws_sum.merge_cells(start_row=cat_start_row, start_column=cat_col_start+1,
                       end_row=cat_start_row, end_column=cat_col_start+3)
    cell = ws_sum.cell(row=cat_start_row, column=cat_col_start+1, value="Por categoría")
    cell.fill = PatternFill("solid", fgColor=BRAND_COLOR)
    cell.font = Font(name=REPORT_FONT, bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    for j, (ch, off) in enumerate([("Categoría",0),("Medios",1),("Visitas/Mes",2)]):
        cell = ws_sum.cell(row=cat_start_row+1, column=cat_col_start+1+j, value=ch)
        cell.fill = PatternFill("solid", fgColor=SUBH_COLOR)
        cell.font = Font(name=REPORT_FONT, bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center" if j>0 else "left",
                                   vertical="center", indent=1 if j==0 else 0)

    for i, cr in enumerate(by_cat):
        r  = cat_start_row + 2 + i
        ev = i % 2 == 0
        fill = PatternFill("solid", fgColor=ALT_FILL if ev else "FFFFFF")
        for off in range(3):
            c = ws_sum.cell(row=r, column=cat_col_start+1+off)
            c.fill = fill; c.font = Font(name=REPORT_FONT, size=9)
        name_key = [k for k,v in cat_map.items() if v is cr]
        ws_sum.cell(row=r, column=cat_col_start+1).value = name_key[0] if name_key else ""
        ws_sum.cell(row=r, column=cat_col_start+2).value = cr["total"]
        ws_sum.cell(row=r, column=cat_col_start+2).alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.cell(row=r, column=cat_col_start+3).value = cr["visits"]
        ws_sum.cell(row=r, column=cat_col_start+3).number_format = "#,##0"
        ws_sum.row_dimensions[r].height = 17

    # ── Sheet 2: Listado completo ─────────────────────────────────────────────
    ws = wb.create_sheet("Medios")
    ws.sheet_view.showGridLines = False

    headers = [
        "Nombre", "URL", "País", "Categoría",
        "Activo", "Stealth", "CDP",
        "Artículos Hoy", "Artículos Total",
        "Visitas/Mes", "Fuente tráfico", "Bounce Rate", "Páginas/Visita",
        "Rank Global", "Rank País",
        "Frecuencia Scrape", "Auto-Enrich", "Frecuencia Tráfico",
        "Último Scrape", "Estado", "Nuevos último scrape", "Error",
        "CSS Selector", "URL Pattern",
    ]
    widths = [
        30, 42, 12, 18,
        8, 8, 8,
        12, 12,
        16, 14, 12, 14,
        14, 12,
        16, 12, 18,
        20, 12, 18, 40,
        22, 22,
    ]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=1, column=i).value = h
    ws.row_dimensions[1].height = 30
    _style_header_row(ws, 1, len(headers), PatternFill, Font, Alignment, Border, Side)
    ws.freeze_panes = "A2"

    FREQ_LABELS = {0: "Manual", 1: "Cada día", 2: "Cada 2 días", 3: "Cada 3 días",
                   7: "Semanal", 14: "Quincenal", 30: "Mensual"}

    for r, s in enumerate(sites, 2):
        even = r % 2 == 0
        _style_data_row(ws, r, len(headers), PatternFill, Font, Alignment, even)
        ws.row_dimensions[r].height = 17

        freq  = s["scrape_frequency_days"] or 0
        freq_label = FREQ_LABELS.get(freq, f"Cada {freq} días" if freq else "Manual")
        tfreq = s["traffic_frequency_days"] or 5
        tfreq_label = f"Cada {tfreq} días"

        vals = [
            s["name"], s["url"],
            s["country"] or "WW", s["category"] or "",
            "Sí" if s["active"] else "No",
            "Sí" if s["use_stealth"] else "No",
            "Sí" if (s["cdp_url"] or "").strip() else "No",
            s["today_c"], s["total_c"],
            s["visits"] or 0,
            s["traffic_source"] or "",
            round(s["bounce_rate"] or 0, 1),
            round(s["pages_per_visit"] or 0, 2),
            s["global_rank"] or 0,
            s["country_rank"] or 0,
            freq_label,
            "Sí" if s["auto_enrich"] else "No",
            tfreq_label,
            (s["last_scraped"] or "")[:16].replace("T", " "),
            s["last_scrape_status"] or "",
            s["last_scrape_count"] or 0,
            (s["last_scrape_error"] or "")[:120],
            s["css_selector"] or "",
            s["article_url_pattern"] or "",
        ]

        bool_cols   = {5, 6, 7, 17}   # 1-indexed: Activo, Stealth, CDP, Auto-Enrich
        num_cols    = {8, 9, 14, 15, 21}
        money_cols  = {10}

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in bool_cols:
                is_yes = v == "Sí"
                cell.font = Font(name=REPORT_FONT, size=9,
                                 color="16A34A" if is_yes else "6B7280", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in num_cols:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in money_cols:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c == 20:  # status
                status_colors = {"ok": "16A34A", "error": "DC2626", "empty": "D97706"}
                color = status_colors.get(v, "6B7280")
                cell.font = Font(name=REPORT_FONT, size=9, color=color, bold=bool(v))
                cell.alignment = Alignment(horizontal="center", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"sports_intel_medios_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Excel: Articles ────────────────────────────────────────────────────────────

@app.get("/export/articles.xlsx")
def export_articles_excel(date_filter: Optional[str] = None, q: Optional[str] = None):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb, *_ = _xl_workbook()
    _, PatternFill, Font, Alignment, Border, Side, get_column_letter = (
        openpyxl, PatternFill, Font, Alignment, Border, Side, get_column_letter)

    subtitle = f"Búsqueda: «{q}»" if q else f"Fecha: {date_filter or 'todas'}"
    _xl_cover_sheet(wb, "Informe de Noticias", subtitle, PatternFill, Font, Alignment, get_column_letter)

    db    = get_db()
    where = ["1=1"]
    params: list = []
    if date_filter:
        where.append("a.scrape_date=?"); params.append(date_filter)
    if q:
        clause, p = _search_where(q)
        where.append(clause); params.extend(p)
    rows = db.execute(f"""
        SELECT a.*, s.name AS site_name, s.category, s.country,
               COALESCE((SELECT monthly_visits FROM site_traffic
                          WHERE site_id=s.id ORDER BY measured_at DESC LIMIT 1),
                         s.monthly_visits_manual, 0) AS site_visits
        FROM articles a JOIN sites s ON s.id=a.site_id
        WHERE {' AND '.join(where)}
        ORDER BY a.scrape_date DESC, a.scraped_at DESC
    """, params).fetchall()

    # ── Aggregate data ─────────────────────────────────────────────────────────
    from collections import defaultdict

    def _ots(visits): return round((visits or 0) / 30)

    agg_medio: dict = defaultdict(lambda: {"articles": 0, "ots": 0, "country": "", "category": ""})
    agg_country: dict = defaultdict(lambda: {"articles": 0, "ots": 0})
    agg_category: dict = defaultdict(lambda: {"articles": 0, "ots": 0})
    agg_date: dict = defaultdict(lambda: {"articles": 0, "ots": 0})

    for a in rows:
        v = a["site_visits"] or 0
        ots_day = _ots(v)
        medio = a["site_name"] or "—"
        country = a["country"] or "—"
        cat = a["category"] or "—"
        d = a["scrape_date"] or "—"
        agg_medio[medio]["articles"] += 1
        agg_medio[medio]["ots"] += ots_day
        agg_medio[medio]["country"] = country
        agg_medio[medio]["category"] = cat
        agg_country[country]["articles"] += 1
        agg_country[country]["ots"] += ots_day
        agg_category[cat]["articles"] += 1
        agg_category[cat]["ots"] += ots_day
        agg_date[d]["articles"] += 1
        agg_date[d]["ots"] += ots_day

    db.close()

    # ── Sheet: Noticias (all rows) ─────────────────────────────────────────────
    ws = wb.create_sheet("Noticias")
    ws.sheet_view.showGridLines = False

    headers = ["Fecha", "Titular", "Medio", "País", "Categoría",
               "Visitas/Mes", "OTS/día", "Propiedad", "Marca", "Agencia", "URL"]
    widths  = [12, 55, 22, 12, 16, 14, 10, 18, 18, 18, 50]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        ws.cell(row=1, column=i).value = h
    ws.row_dimensions[1].height = 30
    _style_header_row(ws, 1, len(headers), PatternFill, Font, Alignment, Border, Side)
    ws.freeze_panes = "A2"

    for r, a in enumerate(rows, 2):
        even = r % 2 == 0
        _style_data_row(ws, r, len(headers), PatternFill, Font, Alignment, even)
        ws.row_dimensions[r].height = 18
        visits = a["site_visits"] or 0
        ots    = _ots(visits)
        vals = [
            a["scrape_date"], a["title"], a["site_name"],
            a["country"] or "", a["category"] or "",
            visits, ots,
            a["company_property"] or "", a["company_brand"] or "",
            a["company_agency"] or "", a["article_url"] or "",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (6, 7):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if c == 11 and v:
                cell.hyperlink = v
                cell.font = Font(name=REPORT_FONT, size=9, color="0D9488", underline="single")

    # ── Sheet: Por Medio ───────────────────────────────────────────────────────
    def _write_agg_sheet(wb_obj, sheet_name, col_label, data_dict, extra_cols=None,
                         PF=PatternFill, F=Font, Al=Alignment, B=Border, S=Side, GCL=get_column_letter):
        ws2 = wb_obj.create_sheet(sheet_name)
        ws2.sheet_view.showGridLines = False
        hdrs = [col_label] + (extra_cols or []) + ["Artículos", "OTS/día"]
        wds  = [30] + [14] * len(extra_cols or []) + [12, 12]
        for i, (h, w) in enumerate(zip(hdrs, wds), 1):
            ws2.column_dimensions[GCL(i)].width = w
            ws2.cell(row=1, column=i).value = h
        ws2.row_dimensions[1].height = 30
        _style_header_row(ws2, 1, len(hdrs), PF, F, Al, B, S)
        ws2.freeze_panes = "A2"
        sorted_rows = sorted(data_dict.items(), key=lambda x: x[1]["articles"], reverse=True)
        for r, (key, vals) in enumerate(sorted_rows, 2):
            _style_data_row(ws2, r, len(hdrs), PF, F, Al, r % 2 == 0)
            ws2.row_dimensions[r].height = 18
            row_vals = [key]
            for ec in (extra_cols or []):
                row_vals.append(vals.get(ec.lower(), ""))
            row_vals += [vals["articles"], vals["ots"]]
            for c, v in enumerate(row_vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                if c >= len(hdrs) - 1:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        return ws2

    ws_medio = wb.create_sheet("Por Medio")
    ws_medio.sheet_view.showGridLines = False
    h_medio = ["Medio", "País", "Categoría", "Artículos", "OTS/día acum."]
    w_medio = [30, 14, 18, 12, 14]
    for i, (h, w) in enumerate(zip(h_medio, w_medio), 1):
        ws_medio.column_dimensions[get_column_letter(i)].width = w
        ws_medio.cell(row=1, column=i).value = h
    ws_medio.row_dimensions[1].height = 30
    _style_header_row(ws_medio, 1, len(h_medio), PatternFill, Font, Alignment, Border, Side)
    ws_medio.freeze_panes = "A2"
    for r, (medio, d) in enumerate(sorted(agg_medio.items(), key=lambda x: x[1]["articles"], reverse=True), 2):
        _style_data_row(ws_medio, r, len(h_medio), PatternFill, Font, Alignment, r % 2 == 0)
        ws_medio.row_dimensions[r].height = 18
        for c, v in enumerate([medio, d["country"], d["category"], d["articles"], d["ots"]], 1):
            cell = ws_medio.cell(row=r, column=c, value=v)
            if c in (4, 5):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── Sheet: Por País ────────────────────────────────────────────────────────
    ws_pais = wb.create_sheet("Por País")
    ws_pais.sheet_view.showGridLines = False
    h_pais = ["País", "Artículos", "OTS/día acum."]
    w_pais = [18, 12, 14]
    for i, (h, w) in enumerate(zip(h_pais, w_pais), 1):
        ws_pais.column_dimensions[get_column_letter(i)].width = w
        ws_pais.cell(row=1, column=i).value = h
    ws_pais.row_dimensions[1].height = 30
    _style_header_row(ws_pais, 1, len(h_pais), PatternFill, Font, Alignment, Border, Side)
    ws_pais.freeze_panes = "A2"
    for r, (country, d) in enumerate(sorted(agg_country.items(), key=lambda x: x[1]["articles"], reverse=True), 2):
        _style_data_row(ws_pais, r, len(h_pais), PatternFill, Font, Alignment, r % 2 == 0)
        ws_pais.row_dimensions[r].height = 18
        for c, v in enumerate([country, d["articles"], d["ots"]], 1):
            cell = ws_pais.cell(row=r, column=c, value=v)
            if c in (2, 3):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── Sheet: Por Categoría ───────────────────────────────────────────────────
    ws_cat = wb.create_sheet("Por Categoría")
    ws_cat.sheet_view.showGridLines = False
    h_cat = ["Categoría", "Artículos", "OTS/día acum."]
    w_cat = [22, 12, 14]
    for i, (h, w) in enumerate(zip(h_cat, w_cat), 1):
        ws_cat.column_dimensions[get_column_letter(i)].width = w
        ws_cat.cell(row=1, column=i).value = h
    ws_cat.row_dimensions[1].height = 30
    _style_header_row(ws_cat, 1, len(h_cat), PatternFill, Font, Alignment, Border, Side)
    ws_cat.freeze_panes = "A2"
    for r, (cat, d) in enumerate(sorted(agg_category.items(), key=lambda x: x[1]["articles"], reverse=True), 2):
        _style_data_row(ws_cat, r, len(h_cat), PatternFill, Font, Alignment, r % 2 == 0)
        ws_cat.row_dimensions[r].height = 18
        for c, v in enumerate([cat, d["articles"], d["ots"]], 1):
            cell = ws_cat.cell(row=r, column=c, value=v)
            if c in (2, 3):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # ── Sheet: Por Fecha ───────────────────────────────────────────────────────
    ws_fecha = wb.create_sheet("Por Fecha")
    ws_fecha.sheet_view.showGridLines = False
    h_fecha = ["Fecha", "Artículos", "OTS/día acum."]
    w_fecha = [14, 12, 14]
    for i, (h, w) in enumerate(zip(h_fecha, w_fecha), 1):
        ws_fecha.column_dimensions[get_column_letter(i)].width = w
        ws_fecha.cell(row=1, column=i).value = h
    ws_fecha.row_dimensions[1].height = 30
    _style_header_row(ws_fecha, 1, len(h_fecha), PatternFill, Font, Alignment, Border, Side)
    ws_fecha.freeze_panes = "A2"
    for r, (d, vals) in enumerate(sorted(agg_date.items(), reverse=True), 2):
        _style_data_row(ws_fecha, r, len(h_fecha), PatternFill, Font, Alignment, r % 2 == 0)
        ws_fecha.row_dimensions[r].height = 18
        for c, v in enumerate([d, vals["articles"], vals["ots"]], 1):
            cell = ws_fecha.cell(row=r, column=c, value=v)
            if c in (2, 3):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"sports_intel_noticias_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Excel: Client report ───────────────────────────────────────────────────────

@app.get("/clients/{cid}/export.xlsx")
def export_client_excel(cid: int, days: int = 30):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb, *_ = _xl_workbook()
    _, PatternFill, Font, Alignment, Border, Side, get_column_letter = (
        openpyxl, PatternFill, Font, Alignment, Border, Side, get_column_letter)

    db = get_db()
    client = db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
    if not client:
        db.close()
        return JSONResponse({"error": "not found"}, status_code=404)

    searches = db.execute("SELECT * FROM client_searches WHERE client_id=?", (cid,)).fetchall()
    brands   = db.execute("SELECT * FROM client_brands WHERE client_id=?", (cid,)).fetchall()

    _xl_cover_sheet(wb, f"Informe de Cliente · {client['name']}",
                    f"Monitorización de medios · últimos {days} días",
                    PatternFill, Font, Alignment, get_column_letter)

    # ── Sheet 1: Resumen de marcas ──────────────────────────────────────────
    ws_brands = wb.create_sheet("Marcas")
    ws_brands.sheet_view.showGridLines = False
    bh = ["Marca", "Búsqueda", "Menciones", "OTS", "VPE (€)", "Países", "Medios distintos"]
    bw = [22, 22, 14, 14, 14, 10, 14]
    for i, (h, w) in enumerate(zip(bh, bw), 1):
        ws_brands.column_dimensions[get_column_letter(i)].width = w
        ws_brands.cell(row=1, column=i).value = h
    ws_brands.row_dimensions[1].height = 30
    _style_header_row(ws_brands, 1, len(bh), PatternFill, Font, Alignment, Border, Side)
    ws_brands.freeze_panes = "A2"

    r = 2
    for s in searches:
        skw = s["keywords"] or ""
        ckw = client["keywords"] or ""
        for b in brands:
            impact = compute_brand_impact(db, b["query"], days=days,
                                         scope_query=ckw, search_scope=skw)
            even = r % 2 == 0
            _style_data_row(ws_brands, r, len(bh), PatternFill, Font, Alignment, even)
            ws_brands.row_dimensions[r].height = 18
            vals = [b["name"], s["name"], impact["total"], impact["ots"],
                    round(impact["vpe"], 2),
                    len(impact.get("by_country", [])), len(impact.get("top_sites", []))]
            for c, v in enumerate(vals, 1):
                cell = ws_brands.cell(row=r, column=c, value=v)
                if c in (3, 4): cell.number_format = "#,##0"
                if c == 5:      cell.number_format = '#,##0.00 "€"'
                if c in (3, 4, 5, 6, 7):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            r += 1

    # ── Sheet 2: Noticias ───────────────────────────────────────────────────
    ws_art = wb.create_sheet("Noticias")
    ws_art.sheet_view.showGridLines = False
    ah = ["Fecha", "Titular", "Marca", "Medio", "País", "Posición", "OTS", "VPE (€)", "URL"]
    aw = [12, 52, 18, 22, 12, 12, 12, 12, 48]
    for i, (h, w) in enumerate(zip(ah, aw), 1):
        ws_art.column_dimensions[get_column_letter(i)].width = w
        ws_art.cell(row=1, column=i).value = h
    ws_art.row_dimensions[1].height = 30
    _style_header_row(ws_art, 1, len(ah), PatternFill, Font, Alignment, Border, Side)
    ws_art.freeze_panes = "A2"

    r = 2
    ckw = client["keywords"] or ""
    for b in brands:
        articles = get_search_articles(db, b["query"], ckw, days=days)
        for a in articles:
            ots_v, pos = _article_ots(a.get("monthly_visits", 0) or 0,
                                      a.get("category", ""), b["query"],
                                      a.get("title", ""), a.get("summary", ""))
            vpe_v = round(ots_v * VPE_RATE_PER_OTS, 2)
            even = r % 2 == 0
            _style_data_row(ws_art, r, len(ah), PatternFill, Font, Alignment, even)
            ws_art.row_dimensions[r].height = 18
            vals = [a.get("scrape_date", ""), a.get("title", ""), b["name"],
                    a.get("site_name", ""), a.get("country", ""), pos,
                    round(ots_v), vpe_v, a.get("article_url", "")]
            for c, v in enumerate(vals, 1):
                cell = ws_art.cell(row=r, column=c, value=v)
                if c == 7: cell.number_format = "#,##0"
                if c == 8: cell.number_format = '#,##0.00 "€"'
                if c == 9 and v:
                    cell.hyperlink = v
                    cell.font = Font(name=REPORT_FONT, size=9, color="0D9488", underline="single")
            r += 1

    db.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"sports_intel_{client['name'].lower().replace(' ','_')}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── PDF: Brand report ──────────────────────────────────────────────────────────

@app.get("/clients/{cid}/searches/{sid}/brands/{bid}/report.pdf")
def export_brand_pdf(cid: int, sid: int, bid: int, days: int = 30):
    from fpdf import FPDF

    db     = get_db()
    client = db.execute("SELECT * FROM clients WHERE id=?", (cid,)).fetchone()
    search = db.execute("SELECT * FROM client_searches WHERE id=? AND client_id=?", (sid, cid)).fetchone()
    brand  = db.execute("SELECT * FROM client_brands  WHERE id=? AND client_id=?", (bid, cid)).fetchone()
    if not (client and search and brand):
        db.close()
        return JSONResponse({"error": "not found"}, status_code=404)

    ckw    = client["keywords"] or ""
    skw    = search["keywords"] or brand["query"]
    impact = compute_brand_impact(db, brand["query"], days=days,
                                  scope_query=ckw, search_scope=skw)
    articles = get_search_articles(db, brand["query"], ckw, days=days)
    db.close()

    # ── Build PDF ──────────────────────────────────────────────────────────
    TEAL   = (13, 148, 136)
    DARK   = (15, 23, 42)
    GRAY   = (107, 114, 128)
    LGRAY  = (241, 245, 249)
    WHITE  = (255, 255, 255)

    _FONT_PATHS = [
        "/Library/Fonts/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
    ]
    _FONT_PATH = next((p for p in _FONT_PATHS if os.path.exists(p)), None)

    class PDF(FPDF):
        def header(self):
            if self.page_no() == 1:
                return
            self.set_fill_color(*DARK)
            self.rect(0, 0, 210, 10, "F")
            self.set_font("uni", size=7)
            self.set_text_color(*WHITE)
            self.set_xy(10, 2)
            self.cell(0, 6, "SPORTS INTEL · INFORME DE MONITORIZACIÓN", align="L")
            self.set_xy(0, 2)
            self.cell(200, 6, f"Pág. {self.page_no()}", align="R")
            self.set_text_color(*DARK)

        def footer(self):
            if self.page_no() == 1:
                return
            self.set_y(-12)
            self.set_font("uni", size=7)
            self.set_text_color(*GRAY)
            self.cell(0, 5, f"Generado el {date.today().strftime('%d/%m/%Y')} · Confidencial · Sports Intel", align="C")

    pdf = PDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.set_margins(15, 14, 15)
    if _FONT_PATH:
        pdf.add_font("uni", style="", fname=_FONT_PATH)
        pdf.add_font("uni", style="B", fname=_FONT_PATH)
    else:
        pdf.add_font("uni", style="", fname="Helvetica")
        pdf.add_font("uni", style="B", fname="Helvetica")

    # ── Cover page ─────────────────────────────────────────────────────────
    pdf.add_page()

    # Dark banner top
    pdf.set_fill_color(*DARK)
    pdf.rect(0, 0, 210, 60, "F")
    pdf.set_fill_color(*TEAL)
    pdf.rect(0, 60, 210, 3, "F")

    pdf.set_font("uni", "B", 26)
    pdf.set_text_color(*WHITE)
    pdf.set_xy(15, 15)
    pdf.cell(0, 12, "SPORTS INTEL", align="L")

    pdf.set_font("uni", "", 11)
    pdf.set_text_color(161, 201, 198)
    pdf.set_xy(15, 30)
    pdf.cell(0, 7, "Plataforma de Monitorización de Medios", align="L")

    pdf.set_font("uni", "B", 18)
    pdf.set_text_color(*WHITE)
    pdf.set_xy(15, 42)
    pdf.cell(0, 10, "INFORME DE MARCA", align="L")

    pdf.set_text_color(*DARK)

    # Brand name big
    pdf.set_font("uni", "B", 32)
    pdf.set_text_color(*TEAL)
    pdf.set_xy(15, 75)
    pdf.cell(0, 16, brand["name"].upper(), align="L")

    pdf.set_font("uni", "", 12)
    pdf.set_text_color(*GRAY)
    pdf.set_xy(15, 94)
    pdf.cell(0, 7, f"Cliente: {client['name']}  ·  Búsqueda: {search['name']}  ·  Período: {days} días", align="L")

    pdf.set_xy(15, 102)
    pdf.cell(0, 6, f"Fecha del informe: {date.today().strftime('%d de %B de %Y')}", align="L")

    # KPI boxes
    def kpi_box(x, y, label, value, sub=""):
        pdf.set_fill_color(*LGRAY)
        pdf.rect(x, y, 52, 30, "F")
        pdf.set_fill_color(*TEAL)
        pdf.rect(x, y, 52, 2, "F")
        pdf.set_font("uni", "", 7)
        pdf.set_text_color(*GRAY)
        pdf.set_xy(x + 2, y + 4)
        pdf.cell(48, 5, label.upper(), align="L")
        pdf.set_font("uni", "B", 16)
        pdf.set_text_color(*TEAL)
        pdf.set_xy(x + 2, y + 10)
        pdf.cell(48, 10, str(value), align="L")
        if sub:
            pdf.set_font("uni", "", 7)
            pdf.set_text_color(*GRAY)
            pdf.set_xy(x + 2, y + 22)
            pdf.cell(48, 5, sub, align="L")

    def fmt_num(n):
        if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
        if n >= 1_000:     return f"{n/1_000:.0f}K"
        return str(int(n))

    kpi_box(15,  118, "Menciones", fmt_num(impact["total"]), f"{days} días")
    kpi_box(70,  118, "OTS",       fmt_num(impact["ots"]),  "impresiones pot.")
    kpi_box(125, 118, "VPE",
            (f"{impact['vpe']/1000:.1f}K€" if impact["vpe"] >= 1000 else f"{impact['vpe']:.0f}€"),
            "€ pub. equivalente")

    kpi_box(15,  155, "Países",    len(impact.get("by_country", [])), "mercados")
    kpi_box(70,  155, "Medios",    len(impact.get("top_sites",  [])), "fuentes distintas")

    # Divider
    pdf.set_draw_color(*TEAL)
    pdf.set_line_width(0.5)
    pdf.line(15, 193, 195, 193)
    pdf.set_font("uni", "", 8)
    pdf.set_text_color(*GRAY)
    pdf.set_xy(15, 196)
    pdf.cell(0, 5, "Información confidencial preparada exclusivamente para uso interno del cliente.", align="L")

    # ── Page 2: Top media ──────────────────────────────────────────────────
    pdf.add_page()
    pdf.set_y(14)

    def section_title(title):
        pdf.set_font("uni", "B", 12)
        pdf.set_text_color(*TEAL)
        pdf.cell(0, 8, title, ln=True)
        pdf.set_fill_color(*TEAL)
        pdf.rect(15, pdf.get_y(), 180, 0.5, "F")
        pdf.ln(4)
        pdf.set_text_color(*DARK)

    def table_header(cols, widths):
        pdf.set_fill_color(*DARK)
        pdf.set_font("uni", "B", 8)
        pdf.set_text_color(*WHITE)
        pdf.set_fill_color(*DARK)
        for label, w in zip(cols, widths):
            pdf.cell(w, 7, label, border=0, fill=True, align="C")
        pdf.ln()
        pdf.set_text_color(*DARK)

    def table_row(vals, widths, aligns, even):
        pdf.set_fill_color(*(LGRAY if even else WHITE))
        pdf.set_font("uni", "", 8)
        pdf.set_text_color(*DARK)
        for v, w, a in zip(vals, widths, aligns):
            pdf.cell(w, 6, str(v)[:40], border=0, fill=True, align=a)
        pdf.ln()

    section_title("Top Medios por Menciones")
    table_header(["Medio", "País", "Menciones", "OTS", "VPE"],
                 [70, 25, 25, 25, 25])
    for i, site in enumerate(impact.get("top_sites", [])[:20]):
        table_row([site.get("name",""), site.get("country",""),
                   site.get("count",""), fmt_num(site.get("ots",0)),
                   f'{site.get("vpe",0):.0f}€'],
                  [70, 25, 25, 25, 25], ["L","C","C","R","R"], i % 2 == 0)
    pdf.ln(6)

    # Country distribution
    if impact.get("by_country"):
        section_title("Distribución por País")
        table_header(["País", "Menciones", "OTS", "VPE"],
                     [80, 30, 35, 35])
        for i, c in enumerate(impact["by_country"][:15]):
            table_row([c.get("name",""), c.get("count",""),
                       fmt_num(c.get("ots",0)), f'{c.get("vpe",0):.0f}€'],
                      [80, 30, 35, 35], ["L","C","R","R"], i % 2 == 0)
        pdf.ln(6)

    # ── Page 3+: Articles ──────────────────────────────────────────────────
    if articles:
        pdf.add_page()
        pdf.set_y(14)
        section_title(f"Noticias con Mención ({len(articles)} artículos)")
        table_header(["Fecha", "Titular", "Medio", "Posición", "OTS"],
                     [20, 90, 38, 18, 14])
        for i, a in enumerate(articles[:200]):
            ots_v, pos = _article_ots(a.get("monthly_visits", 0) or 0,
                                      a.get("category", ""), brand["query"],
                                      a.get("title", ""), a.get("summary", ""))
            pos_label = {"title": "Titular", "summary": "Resumen", "body": "Cuerpo"}.get(pos, pos)
            table_row([a.get("scrape_date",""),
                       (a.get("title","") or "")[:70],
                       (a.get("site_name","") or "")[:28],
                       pos_label, fmt_num(ots_v)],
                      [20, 90, 38, 18, 14], ["C","L","L","C","R"], i % 2 == 0)

    pdf.set_text_color(*DARK)
    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    fname = f"sports_intel_{brand['name'].lower().replace(' ','_')}_{date.today().isoformat()}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── Bulk CSV scrape ────────────────────────────────────────────────────────────

@app.get("/bulk", response_class=HTMLResponse)
async def bulk_page(request: Request):
    return templates.TemplateResponse(request=request, name="bulk.html", context={"active_page": "bulk"})


@app.get("/bulk/test-url")
async def bulk_test_url(url: str):
    """Debug: test what _urllib_extract returns for a given URL."""
    try:
        result = _urllib_extract(url)
        return JSONResponse({
            "ok": True,
            "body_len": len(result.get("body", "")),
            "body_preview": result.get("body", "")[:500],
            "title": result.get("title", ""),
            "date": result.get("date", ""),
            "images": len(result.get("images", [])),
            "videos": len(result.get("videos", [])),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)})


@app.post("/bulk/scrape")
async def bulk_scrape(request: Request, file: UploadFile = File(None),
                      use_stealth: int = Form(0)):
    """Accept a CSV file with a 'url' column (optionally title, date).
    Scrape each URL and return JSON list of results."""
    import csv, io as _io

    content = await file.read()
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(_io.StringIO(text))
    rows = list(reader)

    # Normalise header names
    def _col(row, *names):
        for n in names:
            for k in row:
                if k.strip().lower() == n.lower():
                    return row[k].strip()
        return ""

    results = []
    for row in rows:
        url = _col(row, "url", "URL", "link", "Link", "href")
        if not url:
            continue
        title_hint = _col(row, "title", "Title", "titulo", "Título")
        date_hint  = _col(row, "date", "Date", "fecha", "Fecha")
        try:
            full = fetch_article_full(url, bool(use_stealth))
            title = title_hint or full.get("title") or ""
            date  = date_hint  or full.get("date")  or ""
            sentiment = analyze_sentiment_local(title or url, "", full["body"])
            results.append({
                "url":        url,
                "title":      title,
                "date":       date,
                "body":       full["body"][:15000],
                "body_len":   len(full["body"]),
                "images":     full["images"],
                "videos":     full["videos"],
                "sentiment":  sentiment,
                "ok":         True,
                "error":      "",
            })
        except Exception as e:
            results.append({
                "url": url, "title": title_hint, "date": date_hint,
                "body": "", "body_len": 0, "images": [], "videos": [],
                "sentiment": "", "ok": False, "error": str(e)[:200],
            })

    return JSONResponse({"results": results})


@app.post("/bulk/extract-one")
async def bulk_extract_one(request: Request):
    """Extract full content (body, images, videos) for a single URL.
    Uses run_in_executor so fetch_article_full runs in a thread — multiple
    concurrent requests run in parallel instead of queuing the event loop."""
    data = await request.json()
    url = (data.get("url") or "").strip()
    stealth = bool(data.get("stealth", False))
    if not url:
        return JSONResponse({"ok": False, "error": "No URL provided"})
    try:
        import asyncio, functools
        loop = asyncio.get_event_loop()
        full = await loop.run_in_executor(None, functools.partial(fetch_article_full, url, stealth))
        return JSONResponse({
            "ok":     True,
            "title":  full.get("title", ""),
            "date":   full.get("date", ""),
            "body":   full["body"][:15000],
            "body_len": len(full["body"]),
            "images": full["images"],
            "videos": full["videos"],
            "tier":   full.get("tier", ""),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)[:300]})


@app.post("/bulk/download-zip")
async def bulk_download_zip(request: Request):
    """Create a ZIP with articles.csv + images_manifest.csv + videos_manifest.csv.
    Images are NOT downloaded server-side to avoid timeouts; the manifest lists every
    image URL with its article ID so the user can batch-download locally."""
    import zipfile, csv as _csv, io as _io

    body_bytes = await request.body()
    data = json.loads(body_bytes)
    rows = data.get("results", [])

    # ── articles.csv ────────────────────────────────────────────────────────
    art_buf = _io.StringIO()
    art_writer = _csv.DictWriter(
        art_buf,
        fieldnames=["id", "url", "title", "date", "sentiment", "chars", "imgs", "vids", "body", "error"],
        extrasaction="ignore",
    )
    art_writer.writeheader()

    # ── images_manifest.csv ─────────────────────────────────────────────────
    img_buf = _io.StringIO()
    img_writer = _csv.writer(img_buf)
    img_writer.writerow(["article_id", "img_n", "filename", "image_url", "article_url", "title"])

    # ── videos_manifest.csv ─────────────────────────────────────────────────
    vid_buf = _io.StringIO()
    vid_writer = _csv.writer(vid_buf)
    vid_writer.writerow(["article_id", "vid_n", "video_url", "article_url", "title"])

    for idx, r in enumerate(rows):
        aid    = f"{idx + 1:05d}"
        images = r.get("images") or []
        videos = r.get("videos") or []
        art_writer.writerow({
            "id":        aid,
            "url":       r.get("url", ""),
            "title":     r.get("title", ""),
            "date":      r.get("date", ""),
            "sentiment": r.get("sentiment", ""),
            "chars":     r.get("body_len", 0),
            "imgs":      len(images),
            "vids":      len(videos),
            "body":      r.get("body", ""),
            "error":     r.get("error", ""),
        })
        for n, img_url in enumerate(images[:20]):
            _, ext = os.path.splitext(urlparse(img_url).path)
            ext = ext.lower() if ext.lower() in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".avif"} else ".jpg"
            img_writer.writerow([aid, n, f"{aid}_{n:02d}{ext}", img_url, r.get("url", ""), r.get("title", "")])
        for n, vid_url in enumerate(videos[:5]):
            vid_writer.writerow([aid, n, vid_url, r.get("url", ""), r.get("title", "")])

    readme = (
        "CAMALEONIC CLIPPING — Bulk Export\n"
        "==================================\n\n"
        "Archivos incluidos:\n"
        "  articles.csv         — todos los artículos con ID, título, fecha, sentiment y texto\n"
        "  images_manifest.csv  — URLs de imágenes con article_id y filename sugerido\n"
        "  videos_manifest.csv  — URLs de vídeos con article_id\n\n"
        "Para descargar las imágenes localmente, usa el script Python incluido:\n"
        "  python download_images.py\n"
        "(requiere: pip install requests)\n"
    )

    dl_script = (
        "import csv, os, requests\n\n"
        "os.makedirs('images', exist_ok=True)\n"
        "with open('images_manifest.csv', encoding='utf-8-sig') as f:\n"
        "    for row in csv.DictReader(f):\n"
        "        try:\n"
        "            r = requests.get(row['image_url'], timeout=10, headers={'User-Agent': 'Mozilla/5.0'})\n"
        "            r.raise_for_status()\n"
        "            with open(f\"images/{row['filename']}\", 'wb') as out:\n"
        "                out.write(r.content)\n"
        "            print(f\"OK {row['filename']}\")\n"
        "        except Exception as e:\n"
        "            print(f\"SKIP {row['filename']}: {e}\")\n"
    )

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("articles.csv",        art_buf.getvalue().encode("utf-8-sig"))
        zf.writestr("images_manifest.csv", img_buf.getvalue().encode("utf-8-sig"))
        zf.writestr("videos_manifest.csv", vid_buf.getvalue().encode("utf-8-sig"))
        zf.writestr("README.txt",          readme.encode("utf-8"))
        zf.writestr("download_images.py",  dl_script.encode("utf-8"))

    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="camaleonic_clipping.zip"'},
    )


@app.post("/bulk/export")
async def bulk_export(request: Request, fmt: str = Form("excel")):
    """Accept JSON results body, return Excel or CSV download."""
    body = await request.body()
    data = json.loads(body)
    rows = data.get("results", [])

    if fmt == "csv":
        import csv as _csv, io as _io
        buf = _io.StringIO()
        fields = ["url", "title", "date", "sentiment", "body_len", "images", "videos", "body", "error"]
        w = _csv.DictWriter(buf, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            r2 = dict(r)
            r2["images"] = " | ".join(r2.get("images") or [])
            r2["videos"] = " | ".join(r2.get("videos") or [])
            w.writerow(r2)
        buf.seek(0)
        return StreamingResponse(iter([buf.getvalue().encode("utf-8-sig")]),
                                 media_type="text/csv",
                                 headers={"Content-Disposition": 'attachment; filename="bulk_scrape.csv"'})

    # Excel — write_only streams rows directly to disk: constant RAM regardless of row count
    import openpyxl, tempfile as _tf, os as _os

    tmp = _tf.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()

    BODY_MAX = 2_000

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Scrape results")
    ws.append(["URL", "Título", "Fecha", "Sentiment", "Chars body",
               "Imágenes", "Vídeos", "Body (2k)", "Error"])
    for r in rows:
        ws.append([
            r.get("url", ""),
            r.get("title", ""),
            r.get("date", ""),
            r.get("sentiment", ""),
            r.get("body_len", 0),
            " | ".join(r.get("images") or []),
            " | ".join(r.get("videos") or []),
            (r.get("body") or "")[:BODY_MAX],
            r.get("error", ""),
        ])
    wb.save(tmp.name)

    def _iter_and_delete(path):
        try:
            with open(path, "rb") as f:
                while True:
                    chunk = f.read(65536)
                    if not chunk:
                        break
                    yield chunk
        finally:
            _os.unlink(path)

    return StreamingResponse(
        _iter_and_delete(tmp.name),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="bulk_scrape.xlsx"'},
    )


# ── Server-side media download jobs ──────────────────────────────────────────
_MEDIA_JOBS: dict = {}          # job_id -> job dict
_MEDIA_DIR  = Path(tempfile.gettempdir()) / "cam_media"
_MEDIA_DIR.mkdir(exist_ok=True)


def _media_download_worker(job_id: str, img_tasks: list, vid_rows: list,
                            art_rows: list) -> None:
    """Download images server-side (no CORS), build ZIP, mark job done."""
    import urllib.request as _ur
    import csv as _csv
    import io as _io

    job = _MEDIA_JOBS[job_id]
    downloaded: dict[str, bytes] = {}

    def _fetch(img_url: str) -> bytes:
        req = _ur.Request(img_url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer":    img_url.split("/")[0] + "//" + img_url.split("/")[2] + "/",
        })
        with _ur.urlopen(req, timeout=10) as r:
            data = r.read(10 * 1024 * 1024)   # max 10 MB per image
        return data

    with ThreadPoolExecutor(max_workers=20) as pool:
        futures = {pool.submit(_fetch, url): (filename,) for _, _, url, filename in img_tasks}
        for fut in as_completed(futures):
            (filename,) = futures[fut]
            try:
                downloaded[filename] = fut.result()
                job["ok"] += 1
            except Exception:
                job["fail"] += 1
            job["done"] += 1

    # Build ZIP
    zip_path = _MEDIA_DIR / f"{job_id}.zip"
    with _zipfile_mod.ZipFile(zip_path, "w", _zipfile_mod.ZIP_DEFLATED) as zf:
        for fname, data in downloaded.items():
            zf.writestr(f"images/{fname}", data)

        # images_manifest.csv
        img_buf = _io.StringIO()
        img_w = _csv.writer(img_buf)
        img_w.writerow(["article_id", "img_n", "filename", "image_url"])
        for aid, n, url, fname in img_tasks:
            img_w.writerow([aid, n, fname, url])
        zf.writestr("images_manifest.csv", img_buf.getvalue().encode("utf-8-sig"))

        # videos_manifest.csv
        vid_buf = _io.StringIO()
        vid_w = _csv.writer(vid_buf)
        vid_w.writerow(["article_id", "vid_n", "video_url", "article_url", "title"])
        for row in vid_rows:
            vid_w.writerow(row)
        zf.writestr("videos_manifest.csv", vid_buf.getvalue().encode("utf-8-sig"))

        # articles_index.csv
        art_buf = _io.StringIO()
        art_w = _csv.writer(art_buf)
        art_w.writerow(["article_id", "url", "title", "date"])
        for row in art_rows:
            art_w.writerow(row)
        zf.writestr("articles_index.csv", art_buf.getvalue().encode("utf-8-sig"))

    job["zip_path"] = str(zip_path)
    job["status"]   = "done"


@app.post("/bulk/start-media-download")
async def start_media_download(request: Request):
    """Kick off a background server-side image download job."""
    body_bytes = await request.body()
    data = json.loads(body_bytes)
    results = data.get("results", [])

    img_tasks: list[tuple] = []   # (aid, n, url, filename)
    vid_rows:  list[list]  = []
    art_rows:  list[list]  = []

    for idx, r in enumerate(results):
        aid    = f"{idx + 1:05d}"
        images = r.get("images") or []
        videos = r.get("videos") or []
        title  = r.get("title", "")
        url    = r.get("url", "")
        date   = r.get("date", "")
        art_rows.append([aid, url, title, date])
        for n, img_url in enumerate(images[:20]):
            _, ext = os.path.splitext(urlparse(img_url).path)
            ext = ext.lower() if ext.lower() in {".jpg",".jpeg",".png",".gif",".webp",".avif"} else ".jpg"
            img_tasks.append((aid, n, img_url, f"{aid}_{n:02d}{ext}"))
        for n, vid_url in enumerate(videos[:5]):
            vid_rows.append([aid, n, vid_url, url, title])

    job_id = uuid.uuid4().hex[:10]
    _MEDIA_JOBS[job_id] = {
        "status": "running",
        "total":  len(img_tasks),
        "done":   0,
        "ok":     0,
        "fail":   0,
        "zip_path": None,
    }
    threading.Thread(
        target=_media_download_worker,
        args=(job_id, img_tasks, vid_rows, art_rows),
        daemon=True,
        name=f"media-{job_id}",
    ).start()

    return JSONResponse({"job_id": job_id, "total": len(img_tasks)})


@app.get("/bulk/media-status/{job_id}")
async def media_status(job_id: str):
    job = _MEDIA_JOBS.get(job_id)
    if not job:
        raise HTTPException(404, "Job not found")
    return JSONResponse(job)


@app.get("/bulk/media-zip/{job_id}")
async def media_zip_download(job_id: str):
    from fastapi.responses import FileResponse
    job = _MEDIA_JOBS.get(job_id)
    if not job or job["status"] != "done" or not job.get("zip_path"):
        raise HTTPException(404, "ZIP not ready")
    return FileResponse(
        job["zip_path"],
        media_type="application/zip",
        filename="camaleonic_media.zip",
    )


# Start the scheduler thread (daemon so it dies with the process)
threading.Thread(target=scheduler_loop, daemon=True, name="scheduler").start()


# ── Keep-alive self-ping (prevents Render free tier cold starts) ─────────────
def _keepalive_loop():
    """Ping own /health every 9 minutes so Render never spins the service down."""
    import urllib.request as _ur
    time.sleep(30)   # wait for server to be ready
    base = os.environ.get("RENDER_EXTERNAL_URL", "http://localhost:8000")
    url  = f"{base}/health"
    while True:
        try:
            _ur.urlopen(url, timeout=10)
        except Exception:
            pass
        time.sleep(540)   # 9 minutes


threading.Thread(target=_keepalive_loop, daemon=True, name="keepalive").start()


@app.get("/health")
async def health():
    return JSONResponse({"status": "ok"})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("webapp:app", host="0.0.0.0", port=8000, reload=True)
